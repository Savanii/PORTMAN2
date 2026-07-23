"""
Report-13 : Vessel Movement Report  (module RP01)
File location: modules/RP01/RP01/report13/report13.py
Template location: modules/RP01/RP01/report13.html

Rebuilt against the confirmed real schema (information_schema dump +
sample data check on 2026-07-22):

    - mis_vessel_master : fin_year (text) + month (text, e.g. "April") --
      NOT month_idx. Legacy rows are filtered by matching the month name.
    - ldud_header.vcn_id (int)          -> vcn_header.id            (FK)
    - vcn_header.vessel_master_doc(text)-> vessels.doc_num (text)   (FK)
    - ldud_parcel_ops.ldud_id (int)     -> ldud_header.id            (FK)
    - vessel_cargo is a static cargo-taxonomy lookup table with NO link
      to a specific voyage -- dropped from the join entirely. Real
      per-voyage cargo name comes from ldud_parcel_ops.cargo_name.
    - vcn_header.cargo_type and ldud_parcel_ops.terminal_name are real
      columns -- no longer hardcoded post-cutover.
    - vessels.imo_num (NOT imo_no).
    - Every date/time column in the new schema is stored as `text` in
      plain ISO format ("2026-07-06"), confirmed via live sample data.
      Cast with `NULLIF(col, '')::date` / `::timestamp` to handle blank
      strings safely.
    - Post-cutover operational columns (anchored_datetime, pilot times,
      discharge times, cast-off) are all still blank in current data --
      that's expected for a system that just cut over, not a bug in
      this code. Report-13 will just show blank cells for those until
      real operational data is entered.
    - All date/time fields returned to the client (and exported) are
      formatted as 'DD-MM-YYYY HH24:MI' (e.g. "06-07-2026 14:30"),
      never raw ISO with a 'T' separator.
    - The export (.xlsx) reproduces the full 30-column reference
      workbook layout (`13_Vessel_wise_report_June-2026.xlsx`), even
      though the live queries below only ever populate 19 of those
      columns. Columns with no matching source field (Window Schedule,
      Readiness Time, Cargo Commenced/Completed 2 & 3, Total
      Containers/TEUs/Moves/Crane Hours) are written out blank -- see
      EXPORT_FIELD_MAP below for exactly which UI keys plug into which
      workbook columns.

ASSUMPTIONS still open (not yet confirmed):
    1. mis_vessel_master.month values are assumed to literally match
       MONTH_LABELS strings below ("April", "May", ...). If your data
       uses abbreviations or numbers instead, the legacy WHERE clause
       needs adjusting -- run:
           SELECT DISTINCT month FROM mis_vessel_master;
       to confirm.
    2. get_db()/get_cursor(conn) convention (dict-like fetchall results,
       manual cur.close()) -- taken from report4.py's import line, not
       yet verified against report4's actual query-running code.
    3. Legacy date/time columns (anchorage_time, pilot_pickup, alongside,
       ops_commenced, cargo_completion, cast_off, pilot_board_departure)
       are assumed to already be native timestamp/date columns (not
       text), so to_char() is applied directly without a NULLIF/::cast
       step. If legacy also stores these as text, add the same
       NULLIF(col,'')::timestamp cast used in the new-schema query.
    4. This build's single CARGO COMMENCED/COMPLETED pair is exported
       into the workbook's "...1" columns (Cargo Commenced 1 / Cargo
       Completed 1); the "...2"/"...3" pairs are left blank since the
       current queries don't produce a second or third parcel window.
    5. `quantity` from the queries is exported into TOTAL TONNES, since
       that's the only quantity figure the current code fetches.
"""

from datetime import date
from io import BytesIO

from flask import jsonify, request, send_file, render_template

from database import get_db, get_cursor

# Shared blueprint, defined in modules/RP01/RP01/__init__.py.
from .. import bp



CUTOFF_CALENDAR = date(2026, 7, 1)

MONTH_LABELS = [
    "April", "May", "June", "July", "August", "September",
    "October", "November", "December", "January", "February", "March",
]

# Standard display format for every date/time value returned by this
# report: "DD-MM-YYYY HH24:MI" -- no ISO 'T' separator, no seconds.
DATETIME_FMT = "DD-MM-YYYY HH24:MI"
DATE_ONLY_FMT = "DD-MM-YYYY"



class ReportDataError(Exception):
    """Raised for any data-layer problem while building Report-13."""
    pass


def _month_idx_to_date(fin_year: str, month_idx: int) -> date:
    """(fin_year, month_idx) -> calendar date (1st of month). month_idx 0 = April."""
    start_year = int(fin_year.split("-")[0])
    if month_idx <= 8:  # Apr(0)..Dec(8)
        return date(start_year, 4 + month_idx, 1)
    return date(start_year + 1, month_idx - 8, 1)  # Jan(9)..Mar(11)


def _date_to_fin_year_month_idx(d: date) -> tuple[str, int]:
    """Calendar date -> (financial year, month index)."""

    if d.month >= 4:
        fin_year = f"{d.year}-{str(d.year + 1)[-2:]}"
        month_idx = d.month - 4
    else:
        fin_year = f"{d.year - 1}-{str(d.year)[-2:]}"
        month_idx = d.month + 8

    return fin_year, month_idx


def _period_bounds(fin_year: str, month_idx: int) -> tuple[date, date]:
    """Calendar [start, end) for the given fin_year/month_idx, end exclusive."""
    start = _month_idx_to_date(fin_year, month_idx)
    end = _month_idx_to_date(fin_year, month_idx + 1) if month_idx < 11 else date(start.year + 1, 4, 1)
    return start, end


def _uses_legacy_schema(fin_year: str, month_idx: int) -> bool:
    return _month_idx_to_date(fin_year, month_idx) < CUTOFF_CALENDAR


# ---------------------------------------------------------------------------
# Page route
# ---------------------------------------------------------------------------
@bp.route("/module/RP01/report13/")
def report13_page():
    return render_template("report13.html", port_name="JAWAHARLAL NEHRU PORT / JJLTPL")


# ---------------------------------------------------------------------------
# Meta endpoint -- available years / months, combining both eras
# ---------------------------------------------------------------------------
@bp.route("/api/module/RP01/report13/meta", methods=["GET"])
def report13_meta():
    conn = get_db()
    cur = get_cursor(conn)
    try:
        # Legacy years/months: real columns on mis_vessel_master.
        cur.execute("SELECT DISTINCT fin_year, month FROM mis_vessel_master")
        legacy_rows = cur.fetchall()

        # Post-cutover years/months: derived from ldud_header.created_date,
        # since the new schema has no fin_year/month columns at all.
        cur.execute("""
            SELECT DISTINCT NULLIF(created_date, '')::date AS d
            FROM ldud_header
            WHERE NULLIF(created_date, '') IS NOT NULL
        """)
        new_rows = cur.fetchall()

        month_lookup = {
            "apr": 0,
            "may": 1,
            "jun": 2,
            "jul": 3,
            "aug": 4,
            "sep": 5,
            "oct": 6,
            "nov": 7,
            "dec": 8,
            "jan": 9,
            "feb": 10,
            "mar": 11,
        }

        year_month_idx: dict[str, set[int]] = {}

        # Legacy months (before July 2026)
        for r in legacy_rows:
            fy = r["fin_year"]
            m = (r["month"] or "").strip()[:3].lower()
            idx = month_lookup.get(m)

            if fy and idx is not None:
                year_month_idx.setdefault(fy, set()).add(idx)

        # New schema months (July 2026 onwards)
        for r in new_rows:
            d = r["d"]
            if d is None:
                continue

            fy, idx = _date_to_fin_year_month_idx(d)
            year_month_idx.setdefault(fy, set()).add(idx)

        print("year_month_idx =", year_month_idx)
        years = sorted(year_month_idx.keys())
        months = {
            y: [{"idx": i, "label": MONTH_LABELS[i]} for i in sorted(year_month_idx[y])]
            for y in years
        }
        return jsonify({"years": years, "months": months})
    except Exception as exc:
        return jsonify({"error": f"Failed to load year/month options: {exc}"}), 500
    finally:
        cur.close()


# ---------------------------------------------------------------------------
# Report endpoint
# ---------------------------------------------------------------------------
@bp.route("/api/module/RP01/report13/report", methods=["GET"])
def report13_report():
    fin_year = request.args.get("fin_year")
    month_idx = request.args.get("month_idx", type=int)

    if not fin_year or month_idx is None:
        return jsonify({"error": "fin_year and month_idx are required"}), 400

    print("=" * 80)
    print("REPORT13 API CALLED")
    print("__file__ =", __file__)
    print("fin_year =", fin_year)
    print("month_idx =", month_idx)
    print("month =", MONTH_LABELS[month_idx])
    print("uses_legacy =", _uses_legacy_schema(fin_year, month_idx))
    print("=" * 80)

    try:
        if _uses_legacy_schema(fin_year, month_idx):
            print(">>> CALLING _fetch_legacy()")
            rows = _fetch_legacy(fin_year, month_idx)
        else:
            print(">>> CALLING _fetch_new_schema()")
            rows = _fetch_new_schema(fin_year, month_idx)

        return jsonify({
            "month_label": f"{MONTH_LABELS[month_idx]} {fin_year}",
            "rows": rows
        })

    except ReportDataError as exc:
        return jsonify({"error": str(exc)}), 422
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(exc)}), 500


def _fetch_legacy(fin_year: str, month_idx: int):
    print(">>> INSIDE _fetch_legacy")

    conn = get_db()
    cur = get_cursor(conn)

    try:
        cur.execute(f"""
            SELECT
                fin_year AS month,
                COALESCE(TRIM(vcn_no), '') AS via_no,
                imo_no,
                vessel_name,
                agent,
                grt,
                loa,
                'LIQUID' AS cargo_type,
                cargo,
                'JJLTPL' AS terminal,
                berth_no AS berth,
                to_char(NULLIF(anchorage_time,'')::timestamp,'{DATETIME_FMT}') AS anchored_time,
                to_char(NULLIF(pilot_pickup,'')::timestamp,'{DATETIME_FMT}') AS pilot_boarded,
                to_char(NULLIF(alongside,'')::timestamp,'{DATETIME_FMT}') AS alongside_time,
                to_char(NULLIF(ops_commenced,'')::timestamp,'{DATETIME_FMT}') AS cargo_commenced,
                to_char(NULLIF(cargo_completion,'')::timestamp,'{DATETIME_FMT}') AS cargo_completed,
                to_char(NULLIF(cast_off,'')::timestamp,'{DATETIME_FMT}') AS cast_off_time,
                to_char(NULLIF(pilot_board_departure,'')::timestamp,'{DATETIME_FMT}') AS pilot_disembarked,
                quantity
            FROM mis_vessel_master
            WHERE COALESCE(is_deleted, FALSE) = FALSE
              AND fin_year = %(fin_year)s
              AND LEFT(LOWER(month), 3) = LEFT(LOWER(%(month_label)s), 3)
            ORDER BY NULLIF(anchorage_time,'')::timestamp
        """, {
            "fin_year": fin_year,
            "month_label": MONTH_LABELS[month_idx]
        })

        rows = [dict(r) for r in cur.fetchall()]

        print("TOTAL ROWS =", len(rows))
        if rows:
            print("FIRST ROW =", rows[0])

        return rows

    finally:
        cur.close()
def _fetch_new_schema(fin_year: str, month_idx: int) -> list[dict]:
    period_start, period_end = _period_bounds(fin_year, month_idx)

    conn = get_db()
    cur = get_cursor(conn)

    try:
        cur.execute(f"""
            WITH parcel_agg AS (
                SELECT
                    po.ldud_id,
                    MIN(po.start_dt) AS cargo_commenced,
                    CASE
                        WHEN BOOL_AND(po.end_dt IS NOT NULL)
                        THEN MAX(po.end_dt)
                        ELSE NULL
                    END AS cargo_completed,
                    COALESCE(SUM(lpl.quantity), 0) AS quantity,
                    MAX(po.cargo_name) AS cargo_name
                FROM ldud_parcel_ops po
                LEFT JOIN lueu_parcel_log lpl
                    ON lpl.parcel_op_id = po.id
                   AND COALESCE(lpl.is_deleted, FALSE) = FALSE
                GROUP BY po.ldud_id
            )

            SELECT
                to_char(NULLIF(lh.created_date, '')::date, '{DATE_ONLY_FMT}') AS month,
                vh.via_number AS via_no,
                COALESCE(v.imo_num, '') AS imo_no,
                lh.vessel_name AS vessel_name,
                vh.vessel_agent_name AS agent,
                COALESCE(v.gt, 0) AS grt,
                COALESCE(v.loa, 0) AS loa,
                'LIQUID' AS cargo_type,
                pa.cargo_name AS cargo,
                'JJLTPL' AS terminal,
                vh.berth_name AS berth,
                to_char(NULLIF(lh.anchored_datetime, '')::timestamp, '{DATETIME_FMT}') AS anchored_time,
                to_char(NULLIF(lh.pilot_pickup_time, '')::timestamp, '{DATETIME_FMT}') AS pilot_boarded,
                to_char(NULLIF(lh.alongside_datetime, '')::timestamp, '{DATETIME_FMT}') AS alongside_time,
                to_char(pa.cargo_commenced::timestamp, '{DATETIME_FMT}') AS cargo_commenced,
                to_char(pa.cargo_completed::timestamp, '{DATETIME_FMT}') AS cargo_completed,
                to_char(NULLIF(lh.cast_off_datetime, '')::timestamp, '{DATETIME_FMT}') AS cast_off_time,
                to_char(NULLIF(lh.pilot_disembarked, '')::timestamp, '{DATETIME_FMT}') AS pilot_disembarked,
                COALESCE(pa.quantity, 0) AS quantity

            FROM ldud_header lh

            LEFT JOIN vcn_header vh
                ON vh.id = lh.vcn_id

            LEFT JOIN vessels v
                ON TRIM(v.doc_num) =
                   TRIM(split_part(vh.vessel_master_doc, '/', 1))

            LEFT JOIN parcel_agg pa
                ON pa.ldud_id = lh.id

            WHERE NULLIF(lh.created_date, '') IS NOT NULL
              AND NULLIF(lh.created_date, '')::date >= %(period_start)s
              AND NULLIF(lh.created_date, '')::date < %(period_end)s
              AND COALESCE(lh.is_deleted, FALSE) = FALSE

            ORDER BY
                NULLIF(lh.created_date, '')::date,
                lh.vessel_name
        """, {
            "period_start": period_start,
            "period_end": period_end
        })

        rows = cur.fetchall()

        if rows is None:
            raise ReportDataError("No post-cutover vessel data returned.")

        return [dict(r) for r in rows]

    finally:
        cur.close()


# ---------------------------------------------------------------------------
# Export endpoint
# ---------------------------------------------------------------------------

# Reference-workbook headers, in exact left-to-right order
# (13_Vessel_wise_report_June-2026.xlsx, row 3). "SR. NO." is generated
# here at export time (1..N); everything else is either pulled from the
# existing row dict via EXPORT_FIELD_MAP or left blank.
EXPORT_HEADERS = [
    "SR. NO.", "Month", "VIA NO.", "IMO NUMBER", "NAME OF VESSEL",
    "VESESL AGENT", "GRT", "LOA", "CARGO TYPE", "CARGO", "TERMINAL", "BERTH",
    "WINDOW\nSCHEDULE\nDATE & TIME", "ANCHORED TIME", "READINESS TIME",
    "PILOT BOARDED TIME", "ALONGSIDE TIME",
    "CARGO COMMENCED 1", "CARGO COMPLETED 1",
    "CARGO COMMENCED 2", "CARGO COMPLETED 2",
    "CARGO COMMENCED 3", "CARGO COMPLETED 3",
    "CAST OFF TIME", "PILOT DISEMBARKED",
    "TOTAL CONTAINERS", "TOTAL TEUS", "TOTAL TONNES", "TOTAL MOVES",
    "TOTAL CRANE HOURS",
]

# Column widths lifted directly from the reference workbook (A..AD),
# so the exported sheet's proportions match it column-for-column.
EXPORT_COL_WIDTHS = [
    16.7, 15.2, 12.3, 12.9, 25.3, 19.6, 18.1, 8.7, 12.6, 41.4, 15.6, 12.5,
    36.3, 19.0, 22.7, 20.6, 17.3, 22.6, 21.1, 22.6, 21.1, 22.6, 21.1, 16.6,
    20.0, 19.3, 12.3, 14.7, 13.6, 20.7,
]

# Maps each EXPORT_HEADERS column (by position) to the key already
# present on the row dicts returned by _fetch_legacy / _fetch_new_schema.
# `None` means: this workbook column has no matching source field in the
# current queries, so it is written out blank rather than guessed at.
EXPORT_FIELD_MAP = [
    None,                # SR. NO.
    "month",
    "via_no",
    "imo_no",
    "vessel_name",
    "agent",
    "grt",
    "loa",
    "cargo_type",
    "cargo",
    "terminal",
    "berth",
    None,                # WINDOW SCHEDULE DATE & TIME
    "anchored_time",
    None,                # READINESS TIME
    "pilot_boarded",
    "alongside_time",
   "cargo_commenced",    # CARGO COMMENCED 1 -- blank
    None,                # CARGO COMPLETED 1 -- now blank
    None,                # CARGO COMMENCED 2
    None,                # CARGO COMPLETED 2
    None,               # -> CARGO COMMENCED 3
    "cargo_completed",   # -> CARGO COMPLETED 3
    "cast_off_time",
    "pilot_disembarked",
    None,
    None,
    "quantity",
    None,
    None,
]

@bp.route("/api/module/RP01/report13/export", methods=["GET"])
def report13_export():
    fin_year = request.args.get("fin_year")
    month_idx = request.args.get("month_idx", type=int)

    if not fin_year or month_idx is None:
        return jsonify({"error": "fin_year and month_idx are required"}), 400

    try:
        rows = (_fetch_legacy(fin_year, month_idx)
                if _uses_legacy_schema(fin_year, month_idx)
                else _fetch_new_schema(fin_year, month_idx))
    except ReportDataError as exc:
        return jsonify({"error": str(exc)}), 422
    except Exception as exc:
        return jsonify({"error": f"Failed to build export: {exc}"}), 500

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Report-13"

    n_cols = len(EXPORT_HEADERS)
    last_col_letter = get_column_letter(n_cols)
    month_label = f"{MONTH_LABELS[month_idx]}-{fin_year}"

    # -- Title rows, matching the reference workbook's top banner -------
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"VESSEL INFORMATION FOR {fin_year} ({month_label})"
    ws["A1"].font = Font(name="Aptos Narrow", size=18, bold=True)
    ws.row_dimensions[1].height = 22.8

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = "Terminal - JJLTPL"
    ws["A2"].font = Font(name="Aptos Narrow", size=16, bold=True)
    ws.row_dimensions[2].height = 21.0

    # -- Header row, styled to match the reference workbook -------------
    header_font = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
    )
    header_row_idx = 3
    ws.append(EXPORT_HEADERS)
    ws.row_dimensions[header_row_idx].height = 30
    for cell in ws[header_row_idx]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # -- Data rows --------------------------------------------------------
    # Blank (None) for any workbook column with no matching field in the
    # current row dicts (per EXPORT_FIELD_MAP), instead of guessing.
    for i, row in enumerate(rows, start=1):
        out_row = []
        for key in EXPORT_FIELD_MAP:
            if key is None:
                out_row.append(None)
            else:
                out_row.append(row.get(key))
        out_row[0] = i  # SR. NO.
        ws.append(out_row)

    # -- Column widths, matching the reference workbook ------------------
    for idx, width in enumerate(EXPORT_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = f"A{header_row_idx + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Report13_VesselMovement_{fin_year}_{MONTH_LABELS[month_idx]}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )