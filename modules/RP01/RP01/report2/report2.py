"""
Report-2 — Detailed Break-up of Traffic, Commodity-wise (Appendix-3 style)
Flask Blueprint version. Reads directly from mis_vessel_master (Postgres).

Design notes (updated):
- Row structure now follows the fixed Appendix-3 template: a CATEGORY
  (broad group, e.g. "POL") with one or more CARGO_TYPE sub-rows
  (e.g. "CRUDE", "PRODUCT", "LPG"), matching mis_vessel_master.category
  and mis_vessel_master.cargo respectively.
- The template order/labels are hardcoded (CATEGORY_STRUCTURE below) so the
  report always prints in the same standard Appendix-3 row order, exactly
  like the physical form, even for rows that currently have zero data.
- Matching between the DB text and the template is done on a normalised
  key (upper-case, punctuation/whitespace stripped) so things like
  "Fert. Raw Mat. - Dry" / "FERT RAW MAT DRY" still line up.

- *** CARGO CLASSIFICATION (CARGO_ALIAS) ***
  mis_vessel_master.category / .cargo are free-typed text and are NOT
  reliable on their own: the same real cargo shows up tagged under
  different `category` values depending on who entered it (e.g. "Acetic
  Acid" appears under category='Chemical' AND category='Other Liquid';
  "Phosphoric Acid" appears under category='Ph.Acid', a typo that never
  matches the template's "FERT. RAW MAT. - LIQUID" row; "CBFS"/"FO"
  appear under both 'POL' and 'POL Black').

  So classification here is driven primarily by the CARGO TEXT itself
  (normalised) via the CARGO_ALIAS table below, cross-referenced against
  the vessel_cargo lookup table + known trade abbreviations (CDSBO, IPA,
  VAM, SM, MEK, A.Acid, etc). This is what correctly:
    - rolls CBFS / FO into POL -> PRODUCT (instead of their own lines)
    - routes all Chemical-family cargo (Acetic Acid, Acetone, IPA,
      Methanol, MEK, Nitric Acid, Phenol, Styrene Monomer, Toluene,
      VAM, MDC, N-Butanol, ...) into a new OTHER LIQUID -> CHEMICAL
      sub-row, regardless of what `category` text they were entered
      under
    - routes Phosphoric Acid into FERT. RAW MAT. - LIQUID -> PH. ACID
    - routes edible-oil abbreviations (CDSBO, CPO, CSFO, CSBO, RBD Palm
      Olein, ...) into EDIBLE OIL, even when entered under
      category='Other Liquid'

  A handful of DB rows combine two different buckets in a single cargo
  string (e.g. "Phosphoric Acid/ A. Acid" = Farm-Liquid + Chemical,
  "Base Oil/A. Acid" = Other-Liquid + Chemical, "Edible Oil + Lube Oil"
  = Edible-Oil + Other-Liquid). Splitting the quantity between buckets
  would be a guess, so these are NOT silently merged into either side —
  they're routed to a dedicated "UNCLASSIFIED - NEEDS REVIEW" row (still
  counted in Grand Total) and listed individually in
  `debug.unclassified_ambiguous_rows` so they can be fixed at the
  source / manually re-coded.

  Cargo text that doesn't match CARGO_ALIAS at all falls back to the
  original category/cargo-template matching (and, failing that, gets
  auto-appended as a new row) — so brand-new cargo not yet catalogued
  here still surfaces instead of being dropped.

- Anything in the data that does NOT match the template (a brand-new
  cargo_type under a known category, or an entirely new category) is
  NOT dropped — it's appended automatically: as an extra sub-row under
  its matching category if the category is known, otherwise as a new
  top-level (flat) row at the end, just before Grand Total.
- Unloaded / Loaded is derived from mis_vessel_master.import_export:
    'Import' -> Unloaded
    'Export' -> Loaded
  Any other/unexpected value is excluded from the report and surfaced in
  the `debug.unrecognized_import_export` field of the /report response.
- Month selector is a plain calendar-month abbreviation (Apr..Mar), NOT
  tied to a single financial year. Selecting a month returns that month's
  figures (and the FY-to-date cumulative figures) for EVERY financial year
  present in mis_vessel_master that has data for that month — year columns
  are built dynamically from DISTINCT fin_year (not fixed to "this year
  vs last year"), but are always ordered MOST RECENT FIRST within each
  block, matching the physical Appendix-3 form (current FY, then prior
  FY, then the one before that, etc.)
- Units are Tonnes (not '000 Tonnes), matching the source Appendix-3 sheet.

DATA-SOURCE / FALLBACK NOTES (mirrors report3/report8/report9's pattern):
- Primary source is mis_vessel_master, exactly as before.
- For any (fin_year, fy_month_idx) period that has ZERO rows in
  mis_vessel_master, figures for that period only are pulled instead from
  the live LUEU01 pipeline (vcn_header / ldud_header / ldud_parcel_ops /
  lueu_parcel_log). mis_vessel_master always wins for periods where it has
  data; the live pipeline is purely a gap-filler for periods it hasn't
  reached yet.
- Checked directly against the live schema (information_schema.columns for
  vcn_header/ldud_header/ldud_parcel_ops/lueu_parcel_log): there is NO
  category-equivalent column anywhere in the live pipeline.
  vcn_header.cargo_type looked like a candidate but turned out to be
  cargo-name-level (and sometimes a combined string like
  "FO [E], FURNACE OIL"), not a broad category like "POL"/"Chemical" —
  so it is NOT used here.
- The cargo text for live rows comes from ldud_parcel_ops.cargo_name
  (confirmed clean single values like "BASE OIL", "FURNACE OIL"), joined
  through lueu_parcel_log for quantity/entry_date and up to vcn_header
  for operation_type (flow). This is the same join path as report3's
  _load_live_pipeline_data() / report9's fallback.
- Because live rows have no category field to fall back on, classification
  for live rows is driven ENTIRELY by CARGO_ALIAS. Any live cargo_name that
  CARGO_ALIAS doesn't recognize is routed to the same
  "UNCLASSIFIED - NEEDS REVIEW" row used for ambiguous/combined
  mis_vessel_master cargo strings (still counted in Grand Total, still
  logged in debug.unclassified_ambiguous_rows) — rather than inventing a
  brand-new bucket, since that's the report's existing convention for
  "don't guess, but don't drop it either".
- "FURNACE OIL" (the spelled-out live cargo_name) has been added to
  CARGO_ALIAS -> POL/PRODUCT, since "FO" (its abbreviation) was already
  aliased there.
"""

import io
import re
import datetime
import traceback
from functools import wraps

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, get_cursor

from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# FY month order: Apr .. Mar
MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
CAL_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

FLOW_MAP = {
    "import": "unloaded",
    "export": "loaded",
}

# ---------------------------------------------------------------------
# Fixed Appendix-3 row template: (category label, [sub-row / cargo_type
# labels] or None if the category itself is the row, with no breakdown).
# This defines the printed order — it mirrors the standard JNPA
# Appendix-3 "Detailed Break-up of Traffic - Commodity Wise" form.
# ---------------------------------------------------------------------
CATEGORY_STRUCTURE = [
    ("POL",                        ["CRUDE", "PRODUCT", "LPG"]),
    ("BIO DIESEL",                 None),
    ("EDIBLE OIL",                 None),
    ("OTHER LIQUID",               None),
    ("MOLASSES",                   None),
    ("IRON ORE",                   ["RAW", "PELLETS"]),
    ("OTHER ORE",                  None),
    ("FERTILIZERS FINISHED",       None),
    ("FERT. RAW MAT. - DRY",       ["ROCK PHOSPHATE", "SULPHUR"]),
    ("FERT. RAW MAT. - LIQUID",    ["PH. ACID" ]),
    ("FOOD GRAINS",                ["RICE", "WHEAT", "OTHER"]),
    ("COAL",                       ["THERMAL", "COKING", "OTHER"]),
    ("IRON & STEEL",               None),
    ("SUGAR",                      None),
    ("CEMENT",                     ["DRY BULK"]),
    ("WOODPULP",                   None),
    ("OTHER/MISC.",                ["DRY BULK", "BREAK BULK / PROJ. CARGO", "VEHICLES"]),
    ("PIG IRON & FINISHED STEEL",  None),
    ("AUTOMOBILES",                None),
    ("OTHERS",                     None),
]

UNCLASSIFIED_LABEL = "UNCLASSIFIED - NEEDS REVIEW (combined cargo)"

# ---------------------------------------------------------------------
# Cargo-text classification table. Keys are matched case/punctuation-
# insensitively (via _norm) against mis_vessel_master.cargo (and, for the
# live-pipeline fallback rows, ldud_parcel_ops.cargo_name). Values are
# (category_label, sub_label) where sub_label may be None for flat
# (no-breakdown) categories, or None (the whole value) for cargo strings
# that combine two different buckets and need manual review — see the
# module docstring above.
#
# Sourced from the vessel_cargo lookup table (cargo_category /
# cargo_sub_category_2) plus known trade abbreviations found in
# mis_vessel_master.cargo that don't literally appear in vessel_cargo.
# ---------------------------------------------------------------------
CARGO_ALIAS = {
    # ---- CHEMICAL, merged into the flat OTHER LIQUID total (no separate
    # sub-row) ----
    "Acetic Acid": ("OTHER LIQUID", None),
    "Chemical": ("OTHER LIQUID", None),
    "Chemicals": ("OTHER LIQUID", None),
    "A. acid": ("OTHER LIQUID", None),
    "A. acid/VAM": ("OTHER LIQUID", None),
    "A.Acid": ("OTHER LIQUID", None),
    "Aacid": ("OTHER LIQUID", None),
    "ACETONE/Phenol": ("OTHER LIQUID", None),
    "IPA": ("OTHER LIQUID", None),
    "IPA/A.Acid/SM": ("OTHER LIQUID", None),
    "ISOPROPYL ALCOHOL": ("OTHER LIQUID", None),
    "MDC": ("OTHER LIQUID", None),
    "MEK": ("OTHER LIQUID", None),
    "Methelene Choloride": ("OTHER LIQUID", None),
    "N Butonal/Tolune": ("OTHER LIQUID", None),
    "Nitric Acid": ("OTHER LIQUID", None),
    "Phenol/Aceton/VAM": ("OTHER LIQUID", None),
    "SM": ("OTHER LIQUID", None),
    "SM/IPA/Acetone": ("OTHER LIQUID", None),
    "SM/Meoh": ("OTHER LIQUID", None),
    "Strene Monomer": ("OTHER LIQUID", None),
    "VAM": ("OTHER LIQUID", None),
    "VAM/Aacid": ("OTHER LIQUID", None),

    # ---- OTHER LIQUID (base oil / lube oil - same flat total) ----
    "Base oil": ("OTHER LIQUID", None),
    "Base Oil": ("OTHER LIQUID", None),
    "Base Oil- 600/150": ("OTHER LIQUID", None),
    "Base Oil 150/600": ("OTHER LIQUID", None),
    "BASE OIL KIXX LUBO 150N / 600 N/ 6CST / 4CST": ("OTHER LIQUID", None),
    "BASE OIL KIXX LUBO 150N/BASE OIL KIXX LUBO 600N": ("OTHER LIQUID", None),
    "SHELL 500N/150N": ("OTHER LIQUID", None),
    "Lube oil": ("OTHER LIQUID", None),
    "Lube Oil": ("OTHER LIQUID", None),

    # ---- EDIBLE OIL (flat category, no sub-row) ----
    "CDSBO": ("EDIBLE OIL", None),
    "CPO": ("EDIBLE OIL", None),
    "CSFO": ("EDIBLE OIL", None),
    "CSFO/CSBO": ("EDIBLE OIL", None),
    "Edible oil": ("EDIBLE OIL", None),
    "Edible Oil": ("EDIBLE OIL", None),
    "EDIBLE OIL": ("EDIBLE OIL", None),
    "RBD PALM OLEIN": ("EDIBLE OIL", None),
    "SUNFLOWER OIL": ("EDIBLE OIL", None),
    "CPKO/CPO": ("EDIBLE OIL", None),
    "CPO/CPKO": ("EDIBLE OIL", None),
    "CPO/RBDPO": ("EDIBLE OIL", None),
    "CSBO": ("EDIBLE OIL", None),

    # ---- FERT. RAW MAT. - LIQUID -> PH. ACID ----
    "Phosphoric Acid": ("FERT. RAW MAT. - LIQUID", "PH. ACID"),

    # ---- POL -> PRODUCT (confirmed: CBFS and FO both roll up here) ----
    "CBFS": ("POL", "PRODUCT"),
    "FO": ("POL", "PRODUCT"),
    # Spelled-out form of "FO", confirmed live in ldud_parcel_ops.cargo_name
    # (query against the live DB: DISTINCT cargo_name -> 'FURNACE OIL').
    "FURNACE OIL": ("POL", "PRODUCT"),

    # Both halves of this one now land in the same flat OTHER LIQUID
    # total, so it's no longer ambiguous - merges cleanly.
    "Base Oil/A. Acid": ("OTHER LIQUID", None),

    # ---- Still genuinely ambiguous: a single row's cargo text combines
    # two DIFFERENT buckets, so the quantity can't be safely split
    # between them. Where a specific home was requested, routed there
    # instead of the generic UNCLASSIFIED_LABEL top-level row; still
    # logged in debug.unclassified_ambiguous_rows either way. ----
    "EDIBLE OIL + Lube Oil": ("OTHER LIQUID", None),
    "Phosphoric Acid/ A. Acid": ("FERT. RAW MAT. - LIQUID", "PH. ACID"),
}


MONTH_FULL_NAMES = {
    "Apr": "April", "May": "May", "Jun": "June", "Jul": "July",
    "Aug": "August", "Sep": "September", "Oct": "October", "Nov": "November",
    "Dec": "December", "Jan": "January", "Feb": "February", "Mar": "March",
}


def calendar_year_for_month(fin_year: str, month: str) -> int:
    """FY '2026-27' + month 'Jun' -> 2026 (Apr-Dec fall in the FY start
    year); FY '2026-27' + month 'Feb' -> 2027 (Jan-Mar fall in the FY end
    year)."""
    start_year = int(str(fin_year).split("-")[0])
    return start_year + 1 if month in ("Jan", "Feb", "Mar") else start_year


class ReportDataError(Exception):
    """Raised for any problem loading/validating the report's source data.
    Caught by the route handlers and turned into a clean JSON error response."""
    pass


def month_abbrev(month_str: str) -> str:
    return str(month_str).split("-")[0].strip()


def month_str_to_idx(month_str: str) -> int:
    """'Apr-26' -> 0, 'Dec-24' -> 8, etc. Matches MONTH_NAMES order (FY Apr..Mar)."""
    abbrev = month_abbrev(month_str)
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized value in mis_vessel_master.month: '{month_str}' "
            f"(expected something like 'Apr-26')"
        )


def _dt_to_fy_month(dt):
    """A real datetime/date -> (fin_year, fy_month_idx, month_abbrev),
    Apr-Mar FY convention. e.g. 2026-07-12 -> ('2026-27', 3, 'Jul')."""
    d = dt.date() if isinstance(dt, datetime.datetime) else dt
    fy_start = d.year if d.month >= 4 else d.year - 1
    fin_year = f"{fy_start}-{str(fy_start + 1)[-2:]}"
    mn = CAL_MONTH_ABBR[d.month - 1]
    return fin_year, MONTH_NAMES.index(mn), mn


def _parse_dt(v):
    """Parse the live pipeline's entry_date / free-text datetime values.
    Returns None if blank/unparseable."""
    if not v:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.datetime.strptime(s.replace("T", " ")[:16], "%Y-%m-%d %H:%M")
    except ValueError:
        try:
            return datetime.datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None


def _norm(s) -> str:
    """Normalise a label for matching: upper-case, strip everything that
    isn't a letter or digit. 'Fert. Raw Mat. - Dry' -> 'FERTRAWMATDRY'."""
    return re.sub(r"[^A-Z0-9]+", "", str(s or "").upper())


# Precompute normalised lookups for the fixed template
_TEMPLATE_CAT_NORM = {_norm(cat): cat for cat, _subs in CATEGORY_STRUCTURE}
_TEMPLATE_SUBS_NORM = {
    _norm(cat): {_norm(s): s for s in (subs or [])}
    for cat, subs in CATEGORY_STRUCTURE
}

# Normalised version of CARGO_ALIAS, built once at import time.
CARGO_ALIAS_NORM = {_norm(k): v for k, v in CARGO_ALIAS.items()}

# Cargo strings that combine two different buckets in one row. Logged in
# debug.unclassified_ambiguous_rows for visibility even when CARGO_ALIAS
# now routes them somewhere specific (rather than UNCLASSIFIED_LABEL) —
# the underlying data-entry problem (combined cargo, unsplit quantity)
# is still worth surfacing.
AMBIGUOUS_CARGO_KEYS_NORM = {
    _norm(k) for k, v in CARGO_ALIAS.items()
    if v is None
}

def _apply_cargo_alias(df: pd.DataFrame) -> pd.DataFrame:
    ambiguous_rows = []

    def resolve(row):
        cargo = str(row["cargo_type"]).strip()

        # Only Lube Oil goes to OTHER LIQUID
        if cargo.upper() == "LUBE OIL":
            return "OTHER LIQUID", _norm("OTHER LIQUID"), ""

        key = _norm(cargo)

        if key in CARGO_ALIAS_NORM:
            target = CARGO_ALIAS_NORM[key]

            if target is None:
                ambiguous_rows.append({
                    "fin_year": row["fin_year"],
                    "month": row["month_abbrev"],
                    "category": row["category"],
                    "cargo": row["cargo_type"],
                    "quantity": row["quantity"],
                    "routed_to": "UNCLASSIFIED_LABEL",
                })
                return UNCLASSIFIED_LABEL, _norm(UNCLASSIFIED_LABEL), ""

            cat_label, sub_label = target
            return (
                cat_label,
                _norm(cat_label),
                _norm(sub_label) if sub_label else ""
            )

        # Not in CARGO_ALIAS. mis_vessel_master rows carry their own
        # category text, so fall back to that (existing behaviour). Live
        # LUEU01-pipeline rows have no category field at all (row["category"]
        # is "" — see _load_live_pipeline_data) so there's nothing sensible
        # to fall back to; route those to the same UNCLASSIFIED_LABEL
        # "needs review" row used for ambiguous/combined cargo, rather than
        # guessing or silently dropping them.
        if row["category"]:
            return row["category"], row["category_norm"], row["cargo_norm"]

        ambiguous_rows.append({
            "fin_year": row["fin_year"],
            "month": row["month_abbrev"],
            "category": "(none — live pipeline)",
            "cargo": row["cargo_type"],
            "quantity": row["quantity"],
            "routed_to": "UNCLASSIFIED_LABEL",
        })
        return UNCLASSIFIED_LABEL, _norm(UNCLASSIFIED_LABEL), ""

    resolved = df.apply(resolve, axis=1, result_type="expand")

    df["category"] = resolved[0]
    df["category_norm"] = resolved[1]
    df["cargo_norm"] = resolved[2]

    df.attrs["unclassified_ambiguous_rows"] = ambiguous_rows

    return df


# ── ADD: fallback source -- live LUEU01 logging pipeline. Returns the same
# shape as the mis_vessel_master loader (fin_year, month_abbrev,
# fy_month_idx, category, cargo_type, import_export/flow, quantity) so it
# can be concatenated directly, BEFORE cargo-alias classification runs.
# Only used for (fin_year, fy_month_idx) periods that have ZERO rows in
# mis_vessel_master -- see load_data() below.
#
# Checked directly against the live schema: there is no category-
# equivalent column anywhere in the live pipeline (vcn_header.cargo_type
# is cargo-name-level, not broad-category, and sometimes a combined
# string like "FO [E], FURNACE OIL" — so it is deliberately NOT used
# here). category is left as "" for these rows; _apply_cargo_alias routes
# anything CARGO_ALIAS doesn't recognize to UNCLASSIFIED_LABEL instead of
# falling back to a (nonexistent) category. ──
def _load_live_pipeline_data() -> pd.DataFrame:
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT l.entry_date, po.cargo_name, v.operation_type, l.quantity
            FROM lueu_parcel_log l
            JOIN ldud_parcel_ops po ON po.id = l.parcel_op_id
            JOIN ldud_header ld ON ld.id = po.ldud_id
            JOIN vcn_header v ON v.id = ld.vcn_id
            WHERE l.is_deleted IS NOT TRUE
              AND l.entry_date IS NOT NULL
              AND l.quantity IS NOT NULL
        """)
        log_rows = cur.fetchall()
    finally:
        conn.close()

    empty = pd.DataFrame(columns=[
        "fin_year", "month_abbrev", "fy_month_idx",
        "category", "category_norm", "cargo_type", "cargo_norm",
        "flow", "quantity",
    ])
    if not log_rows:
        return empty

    ldf = pd.DataFrame(log_rows)
    ldf["quantity"] = pd.to_numeric(ldf["quantity"], errors="coerce").fillna(0.0)
    ldf["cargo_type"] = ldf["cargo_name"].fillna("").astype(str).str.strip()

    ldf["_flow_raw"] = ldf["operation_type"].fillna("").astype(str).str.strip().str.lower()
    ldf["flow"] = ldf["_flow_raw"].map(FLOW_MAP)
    ldf = ldf.dropna(subset=["flow"])
    if ldf.empty:
        return empty

    ldf["entry_dt"] = ldf["entry_date"].apply(_parse_dt)
    ldf = ldf.dropna(subset=["entry_dt"])
    if ldf.empty:
        return empty

    fy_list, idx_list, mn_list = [], [], []
    for dt in ldf["entry_dt"]:
        fy, idx, mn = _dt_to_fy_month(dt)
        fy_list.append(fy)
        idx_list.append(idx)
        mn_list.append(mn)
    ldf["fin_year"] = fy_list
    ldf["fy_month_idx"] = idx_list
    ldf["month_abbrev"] = mn_list

    # No category-equivalent column in the live schema (confirmed against
    # information_schema.columns) -- left blank; _apply_cargo_alias routes
    # anything CARGO_ALIAS doesn't recognize to UNCLASSIFIED_LABEL rather
    # than falling back to a nonexistent category.
    ldf["category"] = ""
    ldf["category_norm"] = ""
    ldf["cargo_norm"] = ldf["cargo_type"].apply(_norm)

    return ldf[[
        "fin_year", "month_abbrev", "fy_month_idx",
        "category", "category_norm", "cargo_type", "cargo_norm",
        "flow", "quantity",
    ]].copy()


def load_data() -> pd.DataFrame:
    """Primary source: mis_vessel_master, exactly as before.
    Fallback: for any (fin_year, fy_month_idx) period with ZERO rows in
    mis_vessel_master, figures are pulled instead from the live LUEU01
    pipeline for that period only. mis_vessel_master always wins for
    periods where it has data. Cargo-alias classification runs once, after
    the two sources are combined, so live rows get exactly the same
    CARGO_ALIAS treatment as mis_vessel_master rows."""
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT fin_year, month, category, cargo AS cargo_type, import_export, quantity
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month IS NOT NULL
              AND category IS NOT NULL
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        mv_df = pd.DataFrame(columns=[
            "fin_year", "month_abbrev", "fy_month_idx",
            "category", "category_norm", "cargo_type", "cargo_norm",
            "flow", "quantity",
        ])
        mv_unrecognized = []
    else:
        df = pd.DataFrame(rows)

        missing_cols = [c for c in ("fin_year", "month", "category", "cargo_type", "import_export", "quantity")
                        if c not in df.columns]
        if missing_cols:
            raise ReportDataError(f"Query result is missing column(s): {', '.join(missing_cols)}")

        df["fin_year"] = df["fin_year"].str.strip()
        df["category"] = df["category"].astype(str).str.strip()
        df["cargo_type"] = df["cargo_type"].fillna("").astype(str).str.strip()
        df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0.0)

        df["month_abbrev"] = df["month"].apply(month_abbrev)
        df["fy_month_idx"] = df["month"].apply(month_str_to_idx)

        df["_flow_raw"] = df["import_export"].fillna("").str.strip().str.lower()
        df["flow"] = df["_flow_raw"].map(FLOW_MAP)

        mv_unrecognized = sorted(
            df.loc[df["flow"].isna() & (df["_flow_raw"] != ""), "import_export"].unique().tolist()
        )

        df = df.dropna(subset=["flow"])

        # Raw normalised text (fallback path for cargo not covered by CARGO_ALIAS)
        df["category_norm"] = df["category"].apply(_norm)
        df["cargo_norm"] = df["cargo_type"].apply(_norm)

        mv_df = df[[
            "fin_year", "month_abbrev", "fy_month_idx",
            "category", "category_norm", "cargo_type", "cargo_norm",
            "flow", "quantity",
        ]].copy()

    # ---- which (fin_year, fy_month_idx) periods does mis_vessel_master
    # actually cover? Only periods with ZERO rows there fall back. ----
    # Load live data only from July 2026 onwards
    live_df = _load_live_pipeline_data()

    if not live_df.empty:
        live_df = live_df[
            (
                (live_df["fin_year"] > "2026-27")
            ) |
            (
                (live_df["fin_year"] == "2026-27") &
                (live_df["fy_month_idx"] >= 3)
            )
        ]

    df = pd.concat([mv_df, live_df], ignore_index=True)

    if df.empty:
        raise ReportDataError(
            "No usable rows found in mis_vessel_master or the live LUEU01 pipeline "
            "(or none had a recognized import_export/operation_type value)."
        )

    # Reclassify by cargo text where we have a known mapping. Runs once,
    # after mis_vessel_master + live rows are combined, so both sources go
    # through identical CARGO_ALIAS logic.
    df = _apply_cargo_alias(df)

    df.attrs["unrecognized_import_export"] = mv_unrecognized

    return df[[
        "fin_year", "month_abbrev", "fy_month_idx",
        "category", "category_norm", "cargo_type", "cargo_norm",
        "flow", "quantity",
    ]]


def available_months(df: pd.DataFrame):
    present = set(df["month_abbrev"].unique().tolist())
    return [m for m in MONTH_NAMES if m in present]


def _lookup_cat_only(pivot, cat_norm, yr, flow):
    try:
        return float(pivot.loc[(cat_norm, yr, flow)])
    except KeyError:
        return 0.0


def _lookup_cat_sub(pivot, cat_norm, sub_norm, yr, flow):
    try:
        return float(pivot.loc[(cat_norm, sub_norm, yr, flow)])
    except KeyError:
        return 0.0


def _display_category(df, cat_norm, fallback):
    match = df.loc[df["category_norm"] == cat_norm, "category"]
    return match.iloc[0] if not match.empty else fallback


def _display_cargo(df, cat_norm, sub_norm, fallback):
    match = df.loc[
        (df["category_norm"] == cat_norm) & (df["cargo_norm"] == sub_norm), "cargo_type"
    ]
    return match.iloc[0] if not match.empty else fallback


def build_row_plan(df: pd.DataFrame):
    """Returns an ordered list of row-plan dicts describing every printed
    row (headers + leaves), following the fixed template first, then
    auto-discovered extras that don't match the template."""

    observed = df[["category_norm", "cargo_norm"]].drop_duplicates()
    observed_by_cat = {
        cat: set(sub.tolist())
        for cat, sub in observed.groupby("category_norm")["cargo_norm"]
    }

    plan = []

    for cat_display, subs in CATEGORY_STRUCTURE:
        cat_norm = _norm(cat_display)
        known_subs = subs or []
        known_subs_norm = {_norm(s) for s in known_subs}

        if subs:
            plan.append({"kind": "header", "category": cat_display, "label": None})
            for s in known_subs:
                plan.append({
                    "kind": "leaf", "category": None, "label": f"- {s}",
                    "cat_norm": cat_norm, "sub_norm": _norm(s), "agg_whole_category": False,
                })
            # auto-discovered extra sub-rows under this known category
            extra_subs = sorted(observed_by_cat.get(cat_norm, set()) - known_subs_norm)
            for sub_norm in extra_subs:
                if sub_norm == "":
                    continue
                label = _display_cargo(df, cat_norm, sub_norm, sub_norm)
                plan.append({
                    "kind": "leaf", "category": None, "label": f"- {label}",
                    "cat_norm": cat_norm, "sub_norm": sub_norm, "agg_whole_category": False,
                })
        else:
            plan.append({
                "kind": "leaf", "category": cat_display, "label": None,
                "cat_norm": cat_norm, "sub_norm": None, "agg_whole_category": True,
            })

    # brand-new categories not present in the template at all
    # (this is also where UNCLASSIFIED_LABEL rows land, since it's
    # intentionally not part of CATEGORY_STRUCTURE)
    known_cat_norms = set(_TEMPLATE_CAT_NORM.keys())
    extra_cats = sorted(set(observed_by_cat.keys()) - known_cat_norms)
    for cat_norm in extra_cats:
        if cat_norm == "":
            continue
        label = _display_category(df, cat_norm, cat_norm)
        plan.append({
            "kind": "leaf", "category": label, "label": None,
            "cat_norm": cat_norm, "sub_norm": None, "agg_whole_category": True,
        })

    return plan


def compute_report(df: pd.DataFrame, month: str):
    if month not in MONTH_NAMES:
        raise ReportDataError(f"Unrecognized month '{month}'. Expected one of: {', '.join(MONTH_NAMES)}")

    month_idx = MONTH_NAMES.index(month)

    all_years = sorted(df["fin_year"].unique().tolist())

    for_month_df = df[df["fy_month_idx"] == month_idx]
    upto_month_df = df[df["fy_month_idx"] <= month_idx]

    # most-recent-FY-first, matching the physical Appendix-3 form
    years_with_data = sorted(
        [
            yr for yr in all_years
            if not for_month_df[for_month_df["fin_year"] == yr].empty
            or not upto_month_df[upto_month_df["fin_year"] == yr].empty
        ],
        reverse=True,
    )

    pivot_for_cat = for_month_df.groupby(["category_norm", "fin_year", "flow"])["quantity"].sum()
    pivot_for_sub = for_month_df.groupby(["category_norm", "cargo_norm", "fin_year", "flow"])["quantity"].sum()
    pivot_upto_cat = upto_month_df.groupby(["category_norm", "fin_year", "flow"])["quantity"].sum()
    pivot_upto_sub = upto_month_df.groupby(["category_norm", "cargo_norm", "fin_year", "flow"])["quantity"].sum()

    plan = build_row_plan(df)

    rows = []
    for item in plan:
        if item["kind"] == "header":
            rows.append({
                "type": "header",
                "commodity": item["category"],
                "sub_label": None,
                "for_month": {},
                "upto_month": {},
                "has_data": False,
            })
            continue

        cat_norm = item["cat_norm"]
        sub_norm = item["sub_norm"]
        agg_whole = item["agg_whole_category"]

        row = {
            "type": "leaf",
            "commodity": item["category"],
            "sub_label": item["label"],
            "for_month": {},
            "upto_month": {},
            "has_data": False,
        }

        for yr in years_with_data:
            if agg_whole:
                un_f = round(_lookup_cat_only(pivot_for_cat, cat_norm, yr, "unloaded"), 3)
                ld_f = round(_lookup_cat_only(pivot_for_cat, cat_norm, yr, "loaded"), 3)
                un_u = round(_lookup_cat_only(pivot_upto_cat, cat_norm, yr, "unloaded"), 3)
                ld_u = round(_lookup_cat_only(pivot_upto_cat, cat_norm, yr, "loaded"), 3)
            else:
                un_f = round(_lookup_cat_sub(pivot_for_sub, cat_norm, sub_norm, yr, "unloaded"), 3)
                ld_f = round(_lookup_cat_sub(pivot_for_sub, cat_norm, sub_norm, yr, "loaded"), 3)
                un_u = round(_lookup_cat_sub(pivot_upto_sub, cat_norm, sub_norm, yr, "unloaded"), 3)
                ld_u = round(_lookup_cat_sub(pivot_upto_sub, cat_norm, sub_norm, yr, "loaded"), 3)

            row["for_month"][yr] = {"unloaded": un_f, "loaded": ld_f, "total": round(un_f + ld_f, 3)}
            row["upto_month"][yr] = {"unloaded": un_u, "loaded": ld_u, "total": round(un_u + ld_u, 3)}
            if un_f or ld_f or un_u or ld_u:
                row["has_data"] = True

        rows.append(row)

    totals = {"for_month": {}, "upto_month": {}}
    leaf_rows = [r for r in rows if r["type"] == "leaf"]
    for yr in years_with_data:
        totals["for_month"][yr] = {
            "unloaded": round(sum(r["for_month"][yr]["unloaded"] for r in leaf_rows), 3),
            "loaded": round(sum(r["for_month"][yr]["loaded"] for r in leaf_rows), 3),
            "total": round(sum(r["for_month"][yr]["total"] for r in leaf_rows), 3),
        }
        totals["upto_month"][yr] = {
            "unloaded": round(sum(r["upto_month"][yr]["unloaded"] for r in leaf_rows), 3),
            "loaded": round(sum(r["upto_month"][yr]["loaded"] for r in leaf_rows), 3),
            "total": round(sum(r["upto_month"][yr]["total"] for r in leaf_rows), 3),
        }

    return {
        "month": month,
        "years": years_with_data,
        "rows": rows,
        "totals": totals,
        "debug": {
            "unrecognized_import_export": df.attrs.get("unrecognized_import_export", []),
            "unclassified_ambiguous_rows": df.attrs.get("unclassified_ambiguous_rows", []),
            "years_all": all_years,
            "years_with_data_for_month": years_with_data,
        },
    }


@bp.route("/module/RP01/report2/")
@login_required
def report2_index():
    return render_template("report2/report2.html")


@bp.route("/api/module/RP01/report2/meta")
@login_required
def report2_meta():
    try:
        df = load_data()
        months = available_months(df)
        return jsonify({
            "months": [{"abbrev": m, "label": m} for m in months],
            "years": sorted(df["fin_year"].unique().tolist(), reverse=True),
        })
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report2/report")
@login_required
def report2_report():
    try:
        df = load_data()
        months = available_months(df)
        if not months:
            return jsonify({"error": "No months with data found."}), 400

        month = request.args.get("month", months[-1])
        if month not in months:
            return jsonify({
                "error": f"Unknown/unavailable month '{month}'. Available: {', '.join(months)}"
            }), 400

        result = compute_report(df, month)
        return jsonify(result)
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report2/export")
@login_required
def report2_export():
    try:
        df = load_data()
        months = available_months(df)
        if not months:
            return jsonify({"error": "No months with data found."}), 400

        month = request.args.get("month", months[-1])
        if month not in months:
            return jsonify({
                "error": f"Unknown/unavailable month '{month}'. Available: {', '.join(months)}"
            }), 400

        port_name = request.args.get("port_name", "JAWAHARLAL NEHRU PORT AUTHORITY")

        result = compute_report(df, month)
        years = result["years"]
        rows = result["rows"]
        totals = result["totals"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Report-2"

        # ---- styles -------------------------------------------------
        bold = Font(bold=True)
        title_font = Font(bold=True, underline="single", color="1F4E78", size=12)
        label_font = Font(bold=True, color="1F4E78")
        commodity_font = Font(color="7B241C")
        header_cat_font = Font(bold=True)
        header_font = Font(bold=True)
        unit_font = Font(italic=True, size=10)

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        thick_green = Side(style="medium", color="1E7145")
        title_border = Border(left=thick_green, right=thick_green, top=thick_green, bottom=thick_green)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        n_years = len(years)
        # columns: A margin, B Category, C Sub-commodity, then 3 cols per
        # year for "for month" block, then 3 cols per year for "upto month"
        # block. Years within each block are already ordered current-FY-first.
        FIRST_DATA_COL = 4  # D
        for_month_start = FIRST_DATA_COL
        for_month_end = for_month_start + (n_years * 3) - 1
        upto_month_start = for_month_end + 1
        upto_month_end = upto_month_start + (n_years * 3) - 1
        last_col = upto_month_end

        # ---- title (dynamic, Appendix-3 style) ---------------------------
        full_month = MONTH_FULL_NAMES.get(month, month)

        cur_yr = years[0] if len(years) >= 1 else None
        prev_yr = years[1] if len(years) >= 2 else None

        cur_cal_year = calendar_year_for_month(cur_yr, month) if cur_yr else None
        prev_cal_year = calendar_year_for_month(prev_yr, month) if prev_yr else None

        if cur_cal_year and prev_cal_year:
            month_line = (
                f"FOR THE MONTH {full_month}-{cur_cal_year} VIS-\u00c0-VIS {full_month} {prev_cal_year} "
                f"AND Apr {cur_cal_year} To {full_month}-{cur_cal_year} "
                f"VIA-A-VIS APR {prev_cal_year} To {full_month} {prev_cal_year}"
            )
        elif cur_cal_year:
            month_line = (
                f"FOR THE MONTH {full_month}-{cur_cal_year} "
                f"(Apr {cur_cal_year} To {full_month}-{cur_cal_year})"
            )
        else:
            month_line = f"FOR THE MONTH {full_month}"

        report_date = pd.Timestamp.now().strftime("%d-%m-%Y")

        # Row 2: DATE (left) ................ Appendix-3 (right)
        ws.cell(row=2, column=2, value=f"DATE: {report_date}").font = label_font
        ws.cell(row=2, column=2).alignment = left
        ws.cell(row=2, column=last_col, value="Appendix-3").font = label_font
        ws.cell(row=2, column=last_col).alignment = right

        # Row 3: Port authority name
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
        ws.cell(row=3, column=2, value=port_name).font = title_font
        ws.cell(row=3, column=2).alignment = center
        for col in range(2, last_col + 1):
            ws.cell(row=3, column=col).border = title_border

        # Row 4: report subtitle
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=last_col)
        ws.cell(row=4, column=2, value="DETAILED BREAK-UP OF TRAFFIC - COMMODITY WISE").font = label_font
        ws.cell(row=4, column=2).alignment = center

        # Row 5: dynamic month/FY comparison line
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=last_col)
        ws.cell(row=5, column=2, value=month_line).font = label_font
        ws.cell(row=5, column=2).alignment = center

        ws.cell(row=6, column=last_col, value="(In Tonnes)").font = unit_font
        ws.cell(row=6, column=last_col).alignment = right

        # ---- header rows --------------------------------------------------
        header_row1 = 8
        header_row2 = 9
        header_row3 = 10

        ws.row_dimensions[header_row2].height = 30

        ws.merge_cells(start_row=header_row1, start_column=2, end_row=header_row3, end_column=3)
        ws.cell(row=header_row1, column=2, value="COMMODITY").font = header_font

        col = for_month_start
        for yr in years:
            cal_year = calendar_year_for_month(yr, month)
            yy = str(cal_year)[-2:]

            ws.merge_cells(start_row=header_row1, start_column=col, end_row=header_row1, end_column=col + 2)
            ws.cell(row=header_row1, column=col, value="TRAFFIC FOR THE MONTH OF").font = header_font

            ws.merge_cells(start_row=header_row2, start_column=col, end_row=header_row2, end_column=col + 2)
            ws.cell(row=header_row2, column=col, value=f"{month}-{yy}").font = header_font

            for i, lbl in enumerate(("UNLOADED", "LOADED", "TOTAL")):
                ws.cell(row=header_row3, column=col + i, value=lbl).font = header_font
            col += 3

        for yr in years:
            start_year = yr.split("-")[0]
            end_year = str(calendar_year_for_month(yr, month))

            ws.merge_cells(start_row=header_row1, start_column=col, end_row=header_row1, end_column=col + 2)
            ws.cell(row=header_row1, column=col, value="CUMMULATIVE TRAFFIC UPTO THE").font = header_font

            ws.merge_cells(start_row=header_row2, start_column=col, end_row=header_row2, end_column=col + 2)
            ws.cell(
                row=header_row2, column=col,
                value=f"FY {yr}\n(MONTH APR-{start_year} to {month}-{end_year})",
            ).font = header_font

            for i, lbl in enumerate(("UNLOADED", "LOADED", "TOTAL")):
                ws.cell(row=header_row3, column=col + i, value=lbl).font = header_font
            col += 3

        for r in (header_row1, header_row2, header_row3):
            for c in range(2, last_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center
                cell.border = thin_border

        # ---- data rows ------------------------------------------------
        row_i = header_row3 + 1
        for row in rows:
            b_cell = ws.cell(row=row_i, column=2)
            c_cell = ws.cell(row=row_i, column=3)
            b_cell.border = thin_border
            c_cell.border = thin_border

            if row["type"] == "header":
                b_cell.value = row["commodity"]
                b_cell.font = header_cat_font
                b_cell.alignment = left
                for c in range(FIRST_DATA_COL, last_col + 1):
                    cell = ws.cell(row=row_i, column=c, value=None)
                    cell.border = thin_border
                    cell.alignment = right
                row_i += 1
                continue

            if row["sub_label"]:
                c_cell.value = row["sub_label"]
                c_cell.alignment = left
                c_cell.font = commodity_font
            else:
                b_cell.value = row["commodity"]
                b_cell.alignment = left
                b_cell.font = commodity_font

            col = for_month_start
            for yr in years:
                vals = row["for_month"].get(yr, {"unloaded": 0, "loaded": 0, "total": 0})
                for key in ("unloaded", "loaded", "total"):
                    cell = ws.cell(row=row_i, column=col, value=vals[key])
                    cell.number_format = "0.000"
                    cell.alignment = right
                    cell.border = thin_border
                    cell.font = commodity_font
                    col += 1

            for yr in years:
                vals = row["upto_month"].get(yr, {"unloaded": 0, "loaded": 0, "total": 0})
                for key in ("unloaded", "loaded", "total"):
                    cell = ws.cell(row=row_i, column=col, value=vals[key])
                    cell.number_format = "0.000"
                    cell.alignment = right
                    cell.border = thin_border
                    cell.font = commodity_font
                    col += 1

            if row["has_data"]:
                for c in range(2, last_col + 1):
                    ws.cell(row=row_i, column=c).fill = yellow_fill

            row_i += 1

        # ---- grand total row --------------------------------------------
        total_row = row_i
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
        ws.cell(row=total_row, column=2, value="GRAND TOTAL").font = bold
        ws.cell(row=total_row, column=2).alignment = center
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=3).border = thin_border

        col = for_month_start
        for yr in years:
            vals = totals["for_month"].get(yr, {"unloaded": 0, "loaded": 0, "total": 0})
            for key in ("unloaded", "loaded", "total"):
                cell = ws.cell(row=total_row, column=col, value=vals[key])
                cell.number_format = "0.000"
                cell.font = bold
                cell.alignment = right
                cell.border = thin_border
                col += 1

        for yr in years:
            vals = totals["upto_month"].get(yr, {"unloaded": 0, "loaded": 0, "total": 0})
            for key in ("unloaded", "loaded", "total"):
                cell = ws.cell(row=total_row, column=col, value=vals[key])
                cell.number_format = "0.000"
                cell.font = bold
                cell.alignment = right
                cell.border = thin_border
                col += 1

        # ---- note ---------------------------------------------------------
        note_row = total_row + 2
        ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=last_col)
        ws.cell(row=note_row, column=2,
                value="(**) TARE WEIGHT IS INCLUDED")
        ws.cell(row=note_row, column=2).alignment = left
        ws.cell(row=note_row, column=2).font = unit_font

        sign_row = note_row + 3
        ws.cell(row=sign_row, column=last_col, value="Sr. Manager (Traffic)").font = bold
        ws.cell(row=sign_row, column=last_col).alignment = right

        # ---- column widths -------------------------------------------------
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        for c in range(FIRST_DATA_COL, last_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 13

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-2_{month}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500