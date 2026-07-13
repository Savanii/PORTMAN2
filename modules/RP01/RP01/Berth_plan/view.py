"""
RP02 — Berth Plan (page + daily-report data), living on the shared

"""

from datetime import datetime, timedelta
from functools import wraps

from flask import request, jsonify, render_template, session, redirect, url_for

from .. import bp          # shared RP01 blueprint
from database import get_db, get_cursor, get_user_permissions
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file
# from modules.LUEU01.model import get_started_parcels

MODULE_CODE = 'RP01'

# ---- adjust if your real column name differs ----------
SAIL_COLUMN = 'cast_off_datetime'   # ldud_header column: actual sail time
# BERTHS constant removed — berths are now fetched live from port_berth_master
# -------------------------------------------------------------------------


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_perms():
    if session.get('is_admin'):
        return {'can_read': 1, 'can_add': 1, 'can_edit': 1, 'can_delete': 1}
    return get_user_permissions(session.get('user_id'), MODULE_CODE)


def get_berths(cur=None):
    """Live list of berth names from port_berth_master (no hardcoding)."""
    own_conn = cur is None
    if own_conn:
        conn = get_db()
        cur = get_cursor(conn)
    cur.execute('SELECT berth_name FROM port_berth_master ORDER BY berth_name')
    berths = [r['berth_name'] for r in cur.fetchall()]
    if own_conn:
        conn.close()
    return berths


def get_expected_waiting_vessels():
    """Section C data — queried directly from expected_vessels (EV01 table).
    Excludes vessels already moved to VCN or closed to another terminal,
    and excludes vessels that already have a berth assigned (those show
    up in Section A/B instead)."""
    conn = get_db()
    cur = get_cursor(conn)
    try:
        cur.execute("""
            SELECT terminal_name, vessel_name, via_number, loa, draft,
                agents, tanks, consignees, cargo_name, mla, quantity,
                eta, ata, lpc, doc, nor, berth_name
            FROM expected_vessels
            WHERE (doc_status IS NULL OR doc_status NOT IN ('Moved to VCN', 'Closed - Other Terminal'))
              AND (berth_name IS NULL OR TRIM(berth_name) = '')
            ORDER BY eta ASC NULLS LAST, id DESC
        """)
        rows = [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()

    def _combine(r):
        parts = [r.get('agents'), r.get('tanks'), r.get('consignees')]
        return ' / '.join(p for p in parts if p)

    out = []
    for r in rows:
        out.append({
            'terminal':     r.get('terminal_name'),
            'vessel_name':  r.get('vessel_name'),
            'via_no':       r.get('via_number'),
            'loa':          r.get('loa'),
            'dft':          r.get('draft'),
            'agt_tnk_cons': _combine(r),
            'cargo':        r.get('cargo_name'),
            'mla':          r.get('mla'),
            'quantity':     r.get('quantity'),
            'eta':          _fmt_dt(r.get('eta')),
            'ata':          _fmt_dt(r.get('ata')),
            'lpc':          _fmt_dt(r.get('lpc')),
            'doc':          _fmt_dt(r.get('doc')),
            'nor':          _fmt_dt(r.get('nor')),
            'berth':        r.get('berth_name'),
        })
    return out

# ══════════════════════════════════════════════════════════════════
#  Page route
# ══════════════════════════════════════════════════════════════════

@bp.route('/module/RP01/berth-plan/')
@login_required
def berth_plan_page():
    perms = get_perms()
    if not perms.get('can_read'):
        return render_template('no_access.html'), 403
    return render_template('berth_plan.html', permissions=perms, berths=get_berths())


def get_report_window(plan_date_str):
    """'As on <date> @ 07:00' covers the NEXT 24 hours: date 07:00 -> (date+1) 07:00."""
    plan_date = datetime.strptime(plan_date_str, '%Y-%m-%d')
    window_start = plan_date.replace(hour=7, minute=0, second=0, microsecond=0)
    window_end = window_start + timedelta(days=1)
    return window_start, window_end


def _fmt_dt(v, fmt='%d-%m-%Y %H:%M'):
    """Format a value that may be a real datetime OR a raw text timestamp
    (e.g. po.start_dt/end_dt stored as 'YYYY-MM-DDTHH:MM' strings)."""
    if not v:
        return ''
    if isinstance(v, str):
        try:
            return datetime.strptime(v.replace('T', ' ')[:16], '%Y-%m-%d %H:%M').strftime(fmt)
        except ValueError:
            return v
    return v.strftime(fmt)


def _cumulative_as_of(cur, ldud_id, window_end):
    """Everything logged strictly BEFORE window_end (i.e. as of the report's
    cut-off date/time) — so 'Till now', 'Balance', and 'Present Flow Rate' for
    a past-dated report reflect only the data that existed up to that day,
    not entries operators logged later for subsequent days.

    Returns (logged_qty, total_hours) — real (non-shortclose) rows only, so
    the flow-rate math matches get_started_parcels' avg_rate definition.
    Shortclosed quantity is intentionally excluded here too, matching the
    'real_qty' used for avg_rate in LUEU01; balance still nets against target
    the same way callers already expect (target - logged), consistent with
    how a report for an in-progress vessel should read as of that date.
    """
    cur.execute('''
        SELECT po.id AS parcel_op_id
        FROM ldud_parcel_ops po WHERE po.ldud_id = %s
    ''', [ldud_id])
    pop_ids = [r['parcel_op_id'] for r in cur.fetchall()]
    if not pop_ids:
        return 0.0, 0.0

    cur.execute('''
        SELECT COALESCE(SUM(quantity), 0) AS qty,
               COALESCE(SUM(
                 EXTRACT(EPOCH FROM (
                   COALESCE(NULLIF(to_time,'')::time, '00:00'::time)
                   - COALESCE(NULLIF(from_time,'')::time, '00:00'::time)
                   + CASE WHEN NULLIF(to_time,'')::time < NULLIF(from_time,'')::time
                          THEN INTERVAL '24 hours' ELSE INTERVAL '0' END
                 )) / 3600.0
               ), 0) AS hrs
        FROM lueu_parcel_log
        WHERE parcel_op_id = ANY(%s)
          AND is_deleted IS NOT TRUE
          AND is_shortclose IS NOT TRUE
          AND (NULLIF(entry_date, '')::timestamp
               + COALESCE(NULLIF(from_time, '')::time, '00:00'::time)) < %s
    ''', [pop_ids, window_end])
    row = cur.fetchone()
    return float(row['qty'] or 0), float(row['hrs'] or 0)


MIN_HOURS_FOR_ACTUAL = 4      # must match MIN_HOURS_FOR_ACTUAL in lueu01.html
MAX_REASONABLE_DAYS = 90      # if the actual-rate ETC projects beyond this, treat it as unreliable


def _enrich_vessel(cur, vcn_id, ldud_id, window_start, window_end):
    from modules.LUEU01.model import get_started_parcels

    cur.execute('SELECT operation_type FROM vcn_header WHERE id=%s', [vcn_id])
    row = cur.fetchone()
    is_export = (row or {}).get('operation_type') == 'Export'
    tbl = 'vcn_export_cargo_declaration' if is_export else 'vcn_consigners'

    cur.execute(f'SELECT DISTINCT consigner_name FROM {tbl} '
                f'WHERE vcn_id=%s AND consigner_name IS NOT NULL', [vcn_id])
    consigner = ', '.join(r['consigner_name'] for r in cur.fetchall() if r['consigner_name'])

    cur.execute('''SELECT MIN(po.start_dt::timestamp) AS started
                   FROM ldud_parcel_ops po WHERE po.ldud_id=%s''', [ldud_id])
    ops_commenced = cur.fetchone()['started']

    # target_qty comes from the live VCN parcel quantities (via get_started_parcels) —
    # this is a "current truth" number and is NOT date-dependent.
    parcels = get_started_parcels(vcn_id)
    target_qty = float(sum(p['target_qty'] for p in parcels))

    # ── logged_qty / total_hours are date-dependent: they must reflect ONLY
    #    what had actually been logged as of this report's date, not the
    #    all-time total (which may already include days after the report date
    #    if operators have logged ahead). This makes 'Till now', 'Balance',
    #    and 'Present Flow Rate' change day-by-day correctly for a vessel that
    #    spans multiple days, instead of always showing the final/latest state. ──
    logged_qty, total_hours = _cumulative_as_of(cur, ldud_id, window_end)
    balance_qty = round(target_qty - logged_qty, 3)
    present_flow_rate = round(logged_qty / total_hours, 2) if total_hours > 0 else 0

    # Last-24-hours figure — already correctly window-bound (unchanged).
    cur.execute('''
        SELECT COALESCE(SUM(l.quantity), 0) AS q
        FROM lueu_parcel_log l
        JOIN ldud_parcel_ops po ON po.id = l.parcel_op_id
        WHERE po.ldud_id = %s
          AND l.is_deleted IS NOT TRUE
          AND (NULLIF(l.entry_date, '')::timestamp
               + COALESCE(NULLIF(l.from_time, '')::time, '00:00'::time))
              BETWEEN %s AND %s
    ''', [ldud_id, window_start, window_end])
    last_24hr_qty = round(float(cur.fetchone()['q'] or 0), 3)

    expected_completion = ''
    is_planned = False
    display_rate = present_flow_rate   # default: show the real actual rate (as-of this date)

    def _planned_etc():
        """Latest projected ETC + its rate, across all parcels (Expected Start/Rate)."""
        best_etc, best_rate = None, 0
        for p in parcels:
            rate = float(p.get('expected_flow_rate') or 0)
            start = p.get('expected_start')
            tgt = float(p.get('target_qty') or 0)
            if not (rate and start and tgt):
                continue
            try:
                start_dt = datetime.strptime(str(start).replace('T', ' ')[:16], '%Y-%m-%d %H:%M')
            except ValueError:
                continue
            etc_dt = start_dt + timedelta(hours=tgt / rate)
            if best_etc is None or etc_dt > best_etc:
                best_etc, best_rate = etc_dt, rate
        return best_etc, best_rate

    # whether the actual logged rate (as of this report's date) is trustworthy
    # enough to use as-is
    has_actual = present_flow_rate > 0 and total_hours >= MIN_HOURS_FOR_ACTUAL

    if balance_qty > 0:
        actual_etc = None
        if has_actual:
            hrs_left = balance_qty / present_flow_rate
            actual_etc = window_end + timedelta(hours=hrs_left)

        if actual_etc and (actual_etc - window_end).days <= MAX_REASONABLE_DAYS:
            expected_completion = actual_etc.strftime('%d-%m-%Y %H:%M')
        else:
            best_etc, _ = _planned_etc()
            if best_etc:
                expected_completion = best_etc.strftime('%d-%m-%Y %H:%M')
                is_planned = True

    # ── Flow-rate display: SAME rule for BOTH Section A (berthed) and
    #    Section B (sailed) rows — no longer gated on balance_qty>0.
    #    Use the real actual rate whenever there's enough logged data (as of
    #    this report's date); otherwise fall back to the planned/expected rate. ──
    if not has_actual:
        _, best_rate = _planned_etc()
        if best_rate:
            display_rate = best_rate
            is_planned = True   # amber highlight, same as before

    return {
        'consigner': consigner,
        'quantity': target_qty,
        'ops_commenced': _fmt_dt(ops_commenced),
        'last_24hr_qty': last_24hr_qty,
        'till_now_qty': round(logged_qty, 3),
        'balance': balance_qty,
        'expected_completion': expected_completion,
        'present_flow_rate': display_rate,
        'is_planned': is_planned,
    }


def _base_row(h):
    return {
        'via_no': h.get('via_number') or '',
        'vessel_name': h.get('vessel_name') or '',
        'loa': h.get('loa'),
        'draft': h.get('draft'),
        'agent': h.get('vessel_agent_name') or '',
        'cargo': h.get('cargo_type') or '',
        'berth_name': h.get('berth_name') or '',
        'imo_num': h.get('imo_num') or '',
        'nationality': h.get('nationality') or '',
        'remarks': '',
    }



def get_berthed_vessels(window_start, window_end, berths):
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute('''
        SELECT DISTINCT h.id AS vcn_id, l.id AS ldud_id, h.via_number, h.vessel_name,
               h.loa, h.draft, h.vessel_agent_name, h.cargo_type, h.berth_name,
               h.operation_type,
               v.imo_num, v.nationality,
               l.alongside_datetime,
               (SELECT ec.unload_terminal FROM vcn_export_cargo_declaration ec
                 WHERE ec.vcn_id = h.id LIMIT 1) AS exp_terminal,
               (SELECT ec.pipeline_name FROM vcn_export_cargo_declaration ec
                 WHERE ec.vcn_id = h.id LIMIT 1) AS exp_pipeline,
               (SELECT cn.unload_terminal FROM vcn_consigners cn
                 WHERE cn.vcn_id = h.id LIMIT 1) AS imp_terminal,
               (SELECT cn.pipeline_name FROM vcn_consigners cn
                 WHERE cn.vcn_id = h.id LIMIT 1) AS imp_pipeline
        FROM ldud_parcel_ops po
        JOIN ldud_header l ON l.id = po.ldud_id
        JOIN vcn_header h ON h.id = l.vcn_id
        LEFT JOIN vessels v ON v.vessel_name = h.vessel_name
        WHERE h.berth_name = ANY(%s)
          AND po.start_dt IS NOT NULL
          AND po.end_dt IS NULL
        ORDER BY h.berth_name
    ''', [berths])
    headers = [dict(r) for r in cur.fetchall()]

    out = []
    for h in headers:
        row = _base_row(h)
        row['alongside'] = _fmt_dt(h['alongside_datetime'])
        row['vessel_agent'] = h['vessel_agent_name']
        row['terminal'] = h['exp_terminal'] if h['operation_type'] == 'Export' else h['imp_terminal']
        row['pipeline'] = h['exp_pipeline'] if h['operation_type'] == 'Export' else h['imp_pipeline']
        row.update(_enrich_vessel(cur, h['vcn_id'], h['ldud_id'], window_start, window_end))
        out.append(row)
    conn.close()
    return out

def get_sailed_vessels(window_start, window_end, berths):
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute(f'''
        SELECT DISTINCT h.id AS vcn_id, l.id AS ldud_id, h.via_number, h.vessel_name,
               h.loa, h.draft, h.vessel_agent_name, h.cargo_type, h.berth_name,
               h.operation_type,
               v.imo_num, v.nationality,
               l.alongside_datetime, l.{SAIL_COLUMN}::timestamp AS sail_dt,
               (SELECT ec.unload_terminal FROM vcn_export_cargo_declaration ec
                 WHERE ec.vcn_id = h.id LIMIT 1) AS exp_terminal,
               (SELECT ec.pipeline_name FROM vcn_export_cargo_declaration ec
                 WHERE ec.vcn_id = h.id LIMIT 1) AS exp_pipeline,
               (SELECT cn.unload_terminal FROM vcn_consigners cn
                 WHERE cn.vcn_id = h.id LIMIT 1) AS imp_terminal,
               (SELECT cn.pipeline_name FROM vcn_consigners cn
                 WHERE cn.vcn_id = h.id LIMIT 1) AS imp_pipeline,
               (SELECT MAX(po.end_dt::timestamp)
                  FROM ldud_parcel_ops po
                 WHERE po.ldud_id = l.id) AS cargo_completion_dt
        FROM ldud_header l
        JOIN vcn_header h ON h.id = l.vcn_id
        LEFT JOIN vessels v ON v.vessel_name = h.vessel_name
        ORDER BY h.berth_name
    ''')
    headers = [dict(r) for r in cur.fetchall()]

    def _in_window(dt):
        # window_start (inclusive) -> window_end (exclusive), so a vessel
        # landing exactly on the boundary belongs to ONE day only, never both
        return dt is not None and window_start <= dt < window_end

    out = []
    for h in headers:
        sail_dt = h['sail_dt']
        completion_dt = h['cargo_completion_dt']

        if sail_dt:
            # ---- PRIORITY 1: cast off exists -> show ONLY on the day it happened ----
            if not _in_window(sail_dt):
                continue   # previous day or next day -> skip entirely
        else:
            # ---- PRIORITY 2: no cast off yet -> fall back to cargo completion date ----
            if not _in_window(completion_dt):
                continue

        row = _base_row(h)
        row['alongside'] = _fmt_dt(h['alongside_datetime'])
        row['cast_off'] = _fmt_dt(sail_dt)
        row['cargo_completion'] = _fmt_dt(completion_dt)
        row['vessel_agent'] = h['vessel_agent_name']
        row['terminal'] = h['exp_terminal'] if h['operation_type'] == 'Export' else h['imp_terminal']
        row['pipeline'] = h['exp_pipeline'] if h['operation_type'] == 'Export' else h['imp_pipeline']
        row.update(_enrich_vessel(cur, h['vcn_id'], h['ldud_id'], window_start, window_end))

        balance = row.get('balance')
        if not (balance is not None and balance <= 0):
            continue   # cargo अजून पूर्ण झालेला नाही -> sailed मध्ये नको

        out.append(row)
    conn.close()
    return out

def get_daily_report(plan_date_str):
    window_start, window_end = get_report_window(plan_date_str)
    conn = get_db()
    cur = get_cursor(conn)
    berths = get_berths(cur)
    conn.close()

    return {
        'as_on_date': window_start.strftime('%d-%m-%Y'),
        'as_on_time': window_start.strftime('%H:%M'),
        'window_start': window_start.isoformat(),
        'window_end': window_end.isoformat(),
        'berthed': get_berthed_vessels(window_start, window_end, berths),
        'sailed': get_sailed_vessels(window_start, window_end, berths),
        'berths': berths,
        'expected': get_expected_waiting_vessels(),
    }

@bp.route('/api/module/RP02/berthplan/data')
@login_required
def rp02_data():
    plan_date = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    try:
        report = get_daily_report(plan_date)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify(report)

FIELDS_A = [
    ('VIA/F/OVERSEAS/IMO', '__via_f_ovs_imo'), ('VESSEL NAME', 'vessel_name'),
    ('LOA/BHC/DFT', '__loa_dft'), ('AGT / TF/PL', '__agt_tf_pl'), ('Consigner', 'consigner'),
    ('CARGO', 'cargo'), ('QUANTITY', 'quantity'), ('Alongside (Date/Time)', 'alongside'),
    ('OPS COMMENECED', 'ops_commenced'), ('Last 24 hrs Load/Discharge', 'last_24hr_qty'),
    ('Till now discharged / Load', 'till_now_qty'), ('BALANCE', 'balance'),
    ('Expected completion (Date/Time)', 'expected_completion'),
    ('Present Flow Rate(MT/hr)', 'present_flow_rate'), ('Remarks', 'remarks'),
]
FIELDS_B = [
    ('VIA/F/OVERSEAS/IMO', '__via_f_ovs_imo'), ('VESSEL NAME', 'vessel_name'),
    ('LOA/BHC/DFT', '__loa_dft'), ('AGT / TF/PL', '__agt_tf_pl'), ('Consigner', 'consigner'),
    ('CARGO', 'cargo'), ('QUANTITY', 'quantity'), ('Alongside (Date/Time)', 'alongside'),
    ('OPS COMMENECED', 'ops_commenced'), ('Cargo Completion Time', 'cargo_completion'),
    ('Sail Cast off time', 'cast_off'), ('Flow Rate (MT/hr)', 'present_flow_rate'),
]

def _xl_field_value(row, field):
    if row is None:
        return ''
    if field == '__loa_dft':
        loa, dft = row.get('loa'), row.get('draft')
        return f"{loa or ''} / {dft or ''}" if (loa or dft) else ''
    if field == '__via_f_ovs_imo':
        via = row.get('via_no') or ''
        flag = (row.get('nationality') or '').strip().lower()
        ovs = 'Domestic' if flag == 'india' else 'Overseas'
        imo = row.get('imo_num') or ''
        return f"{via} / F / {ovs} / {imo}"
    if field == '__agt_tf_pl':
        agt = row.get('vessel_agent') or row.get('agent') or ''
        tf = row.get('terminal') or ''
        pl = row.get('pipeline') or ''
        return f"{agt} / {tf} / {pl}"
    val = row.get(field)
    if val is None:
        return ''
    if field == 'expected_completion' and row.get('is_planned') and val:
        return f"{val} (Exp)"
    return val


@bp.route('/api/module/RP02/berthplan/export/excel')
@login_required
def rp02_export_excel():
    plan_date = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    report = get_daily_report(plan_date)
    berths = report['berths']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Berth Plan'

    bold = Font(bold=True)
    bold_underline = Font(bold=True, underline='single')   # title lines 1 & 2
    section_font = Font(bold=True, underline='single', color='0000C8')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill('solid', fgColor='DCE6F1')
    label_fill = PatternFill('solid', fgColor='F4F6FA')
    thin = Side(style='thin', color='444444')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # value-cell fonts, matching the webpage's CSS color rules
    FONT_VALUE   = Font(color='0000C8', bold=True)   # normal logged value      -> blue
    FONT_PLANNED = Font(color='B8860B', bold=True)   # planned/no data yet      -> amber
    FONT_EMPTY   = Font(color='CBD5E1', bold=False)  # no vessel at this berth  -> light grey

    r = 1
    title_rows = []
    title_rows.append((r, 'JSW JNPT LIQUID TERMINAL PRIVATE LIMITED', bold_underline)); r += 1
    title_rows.append((r, 'JAWAHARLAL NEHRU PORT AUTHORITY (BULK TERMINAL)', bold_underline)); r += 1
    title_rows.append((r, 'DAILY PERFORMANCE REPORT', bold)); r += 2

    ws.cell(r, 1, 'As on').font = bold
    date_cell = ws.cell(r, 2, report['as_on_date'])
    date_cell.font = Font(bold=True, color='C00000')
    date_cell.alignment = Alignment(horizontal='center')
    ws.cell(r, 3, '@').font = bold
    time_cell = ws.cell(r, 4, report['as_on_time'])
    time_cell.font = Font(bold=True, color='0000C8')
    time_cell.alignment = Alignment(horizontal='center')
    r += 2

    c_headers = ['TERMINAL', 'VESSEL NAME', 'VIA NO.', 'LOA', 'DFT', 'AGT/TNK/CONS', 'CARGO',
                 'MLA', 'QTY.', 'ETA', 'ATA', 'LPC', 'DOC', 'NOR', 'BERTH']

    # ------------------------------------------------------------------
    # Column-width matching logic
    # ------------------------------------------------------------------
    # Section C always has len(c_headers) = 15 physical columns (cols 1-15).
    # Section A/B use column 1 as the "DETAILS" label, so their data needs
    # to fill columns 2..15 (14 columns) for the two tables' right edges
    # to line up exactly - whether those columns hold real vessel data or
    # are empty, they get merged/stretched to fill the same total width
    # as Section C (same as how the webpage stretches columns to 100%).
    #
    # If there are genuinely more concurrent vessels than 14 slots, we
    # expand beyond 14 (never hide real data) - in that rare case Section
    # C will end up narrower than A/B, since C's column count is fixed.
    # ------------------------------------------------------------------

    def group_by_berth(vessels):
        by_berth = {b: [] for b in berths}
        for v in vessels:
            b = v.get('berth_name') or '(no berth)'
            by_berth.setdefault(b, []).append(v)
        return [(b, vs if vs else [None]) for b, vs in by_berth.items()]

    cols_A = group_by_berth(report['berthed'])
    cols_B = group_by_berth(report['sailed'])

    # per-berth minimum columns needed = max vessels at that berth across A & B,
    # so berth columns line up vertically between Section A and Section B too
    mins = [max(len(cols_A[i][1]), len(cols_B[i][1])) for i in range(len(cols_A))]
    total_min = sum(mins) if mins else 0

    target_data_cols = max(len(c_headers) - 1, total_min)   # normally 14
    extra = target_data_cols - total_min
    spans = mins[:]
    if spans:
        i = 0
        while extra > 0:                       # spread leftover width evenly, round-robin
            spans[i % len(spans)] += 1
            extra -= 1
            i += 1

    max_col = max(1 + target_data_cols, len(c_headers))

    def distribute(total, parts):
        """Split `total` columns into `parts` groups, as evenly as possible."""
        if parts <= 0:
            return []
        base, rem = divmod(total, parts)
        return [base + (1 if i < rem else 0) for i in range(parts)]

    fields_map = {
        'A] VESSEL OPERATION :- BERTHED VESSELS': FIELDS_A,
        'B] VESSEL OPERATION :- SAILED VESSELS': FIELDS_B,
    }

    def write_vertical_section(title, cols):
        nonlocal r
        ws.cell(r, 1, title).font = section_font
        r += 1

        ws.cell(r, 1, 'DETAILS').font = bold
        ws.cell(r, 1).fill = header_fill
        ws.cell(r, 1).border = border

        c = 2
        vessel_ranges = []   # (start_col, span) per vessel, in cols order
        for (b, vs), span in zip(cols, spans):
            start_c = c
            for cc in range(start_c, start_c + span):
                cell = ws.cell(r, cc)
                cell.fill = header_fill
                cell.border = border
            ws.cell(r, start_c, b).font = bold
            ws.cell(r, start_c).alignment = center
            if span > 1:
                ws.merge_cells(start_row=r, start_column=start_c,
                                end_row=r, end_column=start_c + span - 1)
            sub_spans = distribute(span, len(vs))
            cc = start_c
            for vs_span in sub_spans:
                vessel_ranges.append((cc, vs_span))
                cc += vs_span
            c = start_c + span
        r += 1

        for label, field in fields_map[title]:
            ws.cell(r, 1, label).font = bold
            ws.cell(r, 1).fill = label_fill
            ws.cell(r, 1).border = border
            idx = 0
            for b, vs in cols:
                for v in vs:
                    start_c, span = vessel_ranges[idx]
                    idx += 1
                    val = _xl_field_value(v, field)
                    cell = ws.cell(r, start_c, val)
                    for cc in range(start_c, start_c + span):
                        ws.cell(r, cc).border = border
                    cell.alignment = center
                    if span > 1:
                        ws.merge_cells(start_row=r, start_column=start_c,
                                        end_row=r, end_column=start_c + span - 1)
                    if v is None:
                        cell.font = FONT_EMPTY
                    elif v.get('is_planned') and field in ('present_flow_rate', 'expected_completion'):
                        cell.font = FONT_PLANNED
                    else:
                        cell.font = FONT_VALUE
            r += 1
        r += 1

    write_vertical_section('A] VESSEL OPERATION :- BERTHED VESSELS', cols_A)
    write_vertical_section('B] VESSEL OPERATION :- SAILED VESSELS', cols_B)

    ws.cell(r, 1, 'C] Expected /Waiting Tank Vessels at JJLTPL Berth').font = section_font
    r += 1
    for col_i, h in enumerate(c_headers, start=1):
        cell = ws.cell(r, col_i, h)
        cell.font = bold; cell.fill = header_fill; cell.alignment = center; cell.border = border
    r += 1

    for row in report['expected']:
        vals = [row.get('terminal'), row.get('vessel_name'), row.get('via_no'), row.get('loa'),
                row.get('dft'), row.get('agt_tnk_cons'), row.get('cargo'), row.get('mla'),
                row.get('quantity'), row.get('eta'), row.get('ata'), row.get('lpc'),
                row.get('doc'), row.get('nor'), row.get('berth')]
        for col_i, v in enumerate(vals, start=1):
            cell = ws.cell(r, col_i, v if v is not None else '')
            cell.border = border
            cell.alignment = center
        r += 1

    # merge + center the three title rows now that max_col is final
    for row_idx, text, font in title_rows:
        ws.cell(row_idx, 1, text).font = font
        ws.cell(row_idx, 1).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)

    ws.column_dimensions['A'].width = 26
    for i in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Berth_Plan_{plan_date}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )