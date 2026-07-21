"""
Report-4 — Commodity-wise Import Cargo Despatched by Different Modes of Transport
              + Commodity-wise Export Cargo Received by Different Modes of Transport
Flask Blueprint version.

DATA FLOW (current, corrected version):
  1. SOURCE SWITCH BY PERIOD:
       - For any selected fin_year/month_idx BEFORE Jul-2026: source both
         the overall commodity totals and the Import/Export-specific
         totals from mis_vessel_master, grouped by commodity bucket, split
         using mis_vessel_master's OWN `import_export` column.
       - For Jul-2026 and any month AFTER: source both the overall
         commodity totals and the Import/Export-specific totals from
         lueu_parcel_log, reached via:
             ldud_header (vessel_name, cast_off_datetime, operation_type)
               -> ldud_parcel_ops (ldud_id, cargo_name)
                 -> lueu_parcel_log (parcel_op_id, quantity, is_deleted)
         filtered to the selected fin_year/month_idx (derived from
         ldud_header.cast_off_datetime), excluding is_deleted rows,
         grouped into the same commodity buckets via ldud_parcel_ops.cargo_name,
         and split into Import/Export using ldud_header's OWN
         `operation_type` column.
  2. Pipe Line for a bucket = month_total(bucket) - op_total(bucket), where
     month_total (overall, both directions) and op_total (Import- or
     Export-specific) both come from WHICHEVER single source applies to
     the selected period (never mixed/combined across mis_vessel_master
     and lueu_parcel_log for the same month).

  *** HISTORY OF FIXES (kept for context) ***
  - Originally, lueu_parcel_log rows were matched to Import/Export by
    joining vessel_name against vcn_header.operation_type (since
    ldud_header's own operation_type column wasn't known about yet). This
    produced crossed/incorrect Despatched vs Received numbers.
  - As an interim fix, the Despatched(Import) table was fed with "export"
    data and the Received(Export) table was fed with "import" data (a
    "swap") to compensate for the crossed vcn_header-vessel-matching
    results.
  - ldud_header.operation_type was then confirmed to exist directly
    (verified against working SQL provided), which is the real fix:
    lueu_parcel_log rows are now split into Import/Export using
    ldud_header.operation_type directly — the exact same pattern as
    mis_vessel_master.import_export. Vessel-name matching against
    vcn_header, and the swap workaround, have both been removed.
    get_operation_vessel_names() is kept in the file but is now unused
    dead code (see its docstring).
  - CONFIRMED (via a Jun-26 screenshot, sourced from mis_vessel_master —
    BEFORE the Jul-2026 cutoff): the same crossing was happening even on
    the direct-filter path. A swap was applied at that point (Despatched
    fed by "export"-labeled rows, Received fed by "import"-labeled rows),
    uniformly across both sources.
  - THEN REVERSED AGAIN per explicit follow-up instruction ("Despatch
    value should come to Receive and Receive value should come in
    Despatch" — i.e. swap the two tables' contents back from what the
    previous fix produced). The CURRENT, final mapping — applied
    uniformly in both report4_api_report() and report4_api_export() — is
    the DIRECT, unswapped one:
        Despatched (Import) table  <- rows where import_export == "import"
        Received   (Export) table  <- rows where import_export == "export"
    Given how many times this mapping has flipped based on screenshots
    and verbal descriptions alone, if it's reported wrong again, the next
    fix should be based on a SIDE-BY-SIDE comparison: the actual SQL
    query result for a specific bucket/month run directly against the
    database, placed next to the app's displayed value for that same
    bucket/month, rather than another visual/verbal description — to
    avoid continuing to flip this back and forth without settling it.

ASSUMPTIONS MADE (please confirm / correct these):
  1. Rail / Road / Inland-Water-Transport-or-Coastal-Movement figures have
     no data source identified yet, so they are hardcoded to 0 for every
     commodity (matches the sample screenshot, where these are all 0.00).
  2. mis_vessel_master.import_export AND ldud_header.operation_type values
     are both matched case-insensitively against "import" / "export" (so
     "Import", "IMPORT", "import " etc. all match), trimmed of whitespace.
  3. mis_vessel_master.category currently only contains liquid-type values
     (POL, POL Black, Other Liquid, Edible Oil, Chemical, Ph.Acid) — all
     mapped into the single "Liquid" bucket. "Cement" / "Break Bulk" /
     "Containers" category values are mapped too, in case they appear in
     the data later, but as of now there's no data for those buckets.
  4. SOURCE-SWITCH CUTOFF: hardcoded to calendar Jul-2026 (see
     CUTOFF_CALENDAR below). Before that calendar month -> mis_vessel_master.
     That month and after -> lueu_parcel_log. CONFIRM this is the right
     cutover point (vs. e.g. detecting automatically whenever
     mis_vessel_master has no rows for the selected period) — flag if this
     needs to change.
  5. lueu_parcel_log-SPECIFIC: ldud_header has no fin_year/month columns, so
     its `cast_off_datetime` (text) is parsed into a real date and converted
     to fin_year/fy_month_idx the same way mis_vessel_master's fin_year/
     month work (Apr start of financial year).
  6. lueu_parcel_log-SPECIFIC: ldud_parcel_ops.cargo_name is mapped to the
     same Liquid/Cement/Break-Bulk buckets using the SAME CATEGORY_MAP
     guesses used for mis_vessel_master.category. I don't actually know
     cargo_name's real distinct values — if this mapping is wrong, tell me
     the real values and I'll correct CATEGORY_MAP (or add a separate map
     for cargo_name specifically).
  7. lueu_parcel_log-SPECIFIC: rows where is_deleted is true are excluded
     from the sum. Rows with is_shortclose = true are currently still
     INCLUDED — flag if those should be excluded too.
  8. Sr. No. values (2, 3, 4) are kept exactly as shown in your screenshot.
  9. The Export table is titled "COMMODITY-WISE EXPORT CARGO RECEIVED BY
     DIFFERENT MODES OF TRANSPORT" with "Received by ..." column headers.
  10. The JSON API returns the Export table under new keys
      (export_rows / export_grand / export_total) alongside the existing
      top-level keys (rows / grand / import_total).
  11. report4_api_meta's year/month picklist is still derived only from
      mis_vessel_master (via _get_df_and_years). If fin_years/months exist
      ONLY in lueu_parcel_log (e.g. far in the future with no
      mis_vessel_master rows at all), they won't show up in the picker yet.
      Flag if the picklist needs to also incorporate ldud_header dates.
  12. Excel export contains only the Import and Export table sheets — no
      Summary sheet (removed per request).
  13. Tonnage figures (Rail/Road/Inland/Pipe Line/Total columns) are shown
      to 6 decimal places, both in the JSON API and the Excel export.
      Percentages remain at 2 decimal places.
"""

import io
import traceback
from functools import wraps

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, get_cursor

from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# Commodity buckets shown in report4, in display order, with the Sr. No.
# exactly as shown in your screenshot.
CATEGORY_ORDER = [
    {"sr": 2, "key": "DRY BULK (CEMENT)"},
    {"sr": 3, "key": "Liquid"},
    {"sr": 4, "key": "Break Bulk / Containers"},
]

# mis_vessel_master.category (and, as a fallback guess, ldud_parcel_ops.cargo_name)
# -> report4 bucket. Only the "Liquid" entries have confirmed data today (per
# `SELECT DISTINCT category FROM mis_vessel_master`). Cement / Break Bulk /
# Containers mappings are included pre-emptively — update the left-hand
# values if your actual category/cargo_name text differs from these guesses.
CATEGORY_MAP = {
    # Liquid
    "POL": "Liquid",
    "POL Black": "Liquid",
    "Other Liquid": "Liquid",
    "Edible Oil": "Liquid",
    "Chemical": "Liquid",
    "Ph.Acid": "Liquid",

    # Add these two
    "ACETONE": "Liquid",
    "FURNACE OIL": "Liquid",

    # Dry Bulk
    "Cement": "DRY BULK (CEMENT)",

    # Break Bulk
    "Break Bulk": "Break Bulk / Containers",
    "Containers": "Break Bulk / Containers",
    "General Cargo": "Break Bulk / Containers",
}

# Calendar (year, month) cutover: any selected period STRICTLY BEFORE this
# is sourced from mis_vessel_master; this month and any AFTER it is sourced
# from lueu_parcel_log. See ASSUMPTION 5 above.
CUTOFF_CALENDAR = (2026, 7)  # Jul-2026


class ReportDataError(Exception):
    """Raised for any problem loading/validating the report's source data.
    Caught by the route handlers and turned into a clean JSON error response."""
    pass


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


def month_str_to_idx(month_str: str) -> int:
    """'Apr-26' -> 0, 'Dec-24' -> 8, etc. Matches MONTH_NAMES order (FY Apr..Mar)."""
    abbrev = str(month_str).split("-")[0].strip()
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized value in mis_vessel_master.month: '{month_str}' "
            f"(expected something like 'Apr-26')"
        )


def idx_to_month_label(fin_year: str, month_idx: int) -> str:
    """Reverse of month_str_to_idx — builds e.g. 'Jun-26' for a given fin_year/idx.
    Assumes mis_vessel_master.month is stored in this exact 'Mon-YY' format."""
    opts = month_options_for(fin_year)
    match = next((o for o in opts if o["idx"] == month_idx), None)
    if not match:
        raise ReportDataError(f"Invalid month_idx {month_idx} for fin_year {fin_year}")
    return match["label"]


def calendar_date_to_fy(dt: pd.Timestamp):
    """Converts a real calendar date into (fin_year_str, fy_month_idx),
    using an Apr-start financial year — same convention as
    mis_vessel_master.fin_year/month. Returns (None, None) if dt is NaT."""
    if pd.isna(dt):
        return None, None
    y, m = dt.year, dt.month
    if m >= 4:
        fin_year = f"{y}-{str((y + 1) % 100).zfill(2)}"
        fy_month_idx = m - 4
    else:
        fin_year = f"{y - 1}-{str(y % 100).zfill(2)}"
        fy_month_idx = m + 8
    return fin_year, fy_month_idx


def month_idx_to_calendar(fin_year: str, month_idx: int):
    """Reverse of calendar_date_to_fy: (fin_year, fy_month_idx) -> (calendar_year,
    calendar_month). E.g. fin_year='2026-27', month_idx=3 (Jul) -> (2026, 7).
    fin_year='2026-27', month_idx=9 (Jan) -> (2027, 1)."""
    start_y = fy_start_year(fin_year)
    calendar_month = ((month_idx + 3) % 12) + 1  # idx0(Apr)->4 ... idx8(Dec)->12, idx9(Jan)->1 ...
    calendar_year = start_y if month_idx < 9 else start_y + 1
    return calendar_year, calendar_month


def use_lueu_source(fin_year: str, month_idx: int) -> bool:
    """True for the selected period at/after CUTOFF_CALENDAR (Jul-2026) ->
    source from lueu_parcel_log. False for anything before that -> source
    from mis_vessel_master."""
    return month_idx_to_calendar(fin_year, month_idx) >= CUTOFF_CALENDAR


def _normalize_name(val) -> str:
    """Trims and casefolds a vessel_name for cross-table matching."""
    return str(val).strip().casefold()


def load_data() -> pd.DataFrame:
    """Loads mis_vessel_master rows (fin_year/month/category/quantity/
    vessel_name/import_export)."""

    conn = get_db()

    try:
        cur = get_cursor(conn)

        cur.execute("""
            SELECT
                fin_year,
                month,
                category,
                quantity,
                vessel_name,
                import_export
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month IS NOT NULL
        """)

        rows = cur.fetchall()

    finally:
        conn.close()

    if not rows:
        raise ReportDataError("No rows found in mis_vessel_master.")

    df = pd.DataFrame(rows)

    df["fin_year"] = df["fin_year"].astype(str).str.strip()
    df["month"] = df["month"].astype(str).str.strip()
    df["category"] = df["category"].astype(str).str.strip()
    df["vessel_name"] = df["vessel_name"].astype(str).str.strip()
    df["import_export"] = df["import_export"].astype(str).str.strip().str.lower()

    df["quantity"] = (
        pd.to_numeric(df["quantity"], errors="coerce")
        .fillna(0.0)
    )

    df["fy_month_idx"] = df["month"].apply(month_str_to_idx)

    df["bucket"] = df["category"].map(CATEGORY_MAP)

    df = df.dropna(subset=["bucket"])

    df["vessel_name_norm"] = df["vessel_name"].apply(_normalize_name)

    df["quantity_000t"] = df["quantity"] / 1000.0

    print("=" * 80)
    print("MIS DATA")
    print(df[[
        "fin_year",
        "month",
        "bucket",
        "vessel_name",
        "import_export",
        "quantity_000t"
    ]])
    print("=" * 80)

    return df[[
        "fin_year",
        "fy_month_idx",
        "bucket",
        "quantity_000t",
        "vessel_name_norm",
        "import_export"
    ]]


def load_lueu_data(fin_year: str, month_idx: int) -> pd.DataFrame:
    """Loads lueu_parcel_log quantities for the GIVEN fin_year/month_idx only,
    via ldud_header -> ldud_parcel_ops -> lueu_parcel_log, excluding
    is_deleted rows. Period is derived from ldud_header.cast_off_datetime
    (parsed into a real date, then converted to fin_year/fy_month_idx using
    the same Apr-start financial-year convention as mis_vessel_master).

    UPDATED: ldud_header has its own `operation_type` column ('Import' /
    'Export'), confirmed directly against your own working queries:
        SELECT COALESCE(SUM(lul.quantity), 0)
        FROM ldud_header lh
        JOIN ldud_parcel_ops lpo ON lpo.ldud_id = lh.id
        JOIN lueu_parcel_log lul ON lul.parcel_op_id = lpo.id
        WHERE (lul.is_deleted IS NULL OR lul.is_deleted = FALSE)
          AND LOWER(TRIM(lh.operation_type)) = 'import'/'export'
          AND TO_DATE(SPLIT_PART(lh.cast_off_datetime, 'T', 1), 'YYYY-MM-DD')
              BETWEEN ... AND ...
    This means lueu_parcel_log/ldud_header does NOT need vessel-name
    matching against vcn_header at all — exactly like mis_vessel_master's
    own `import_export` column, `lh.operation_type` (normalized the same
    way: trimmed + lowercased) is now used directly to split Import vs
    Export for this source. The old vcn_header vessel-name-matching
    approach for this path (and the "swap" workaround that compensated for
    its incorrect crossing) has been removed / reverted."""

    conn = get_db()

    try:
        cur = get_cursor(conn)

        cur.execute("""
        SELECT
            lh.vessel_name,
            lh.cast_off_datetime,
            LOWER(TRIM(lh.operation_type)) AS operation_type,
            MAX(lpo.cargo_name) AS cargo_name,
            SUM(COALESCE(lul.quantity,0)) AS quantity
        FROM ldud_header lh
        JOIN ldud_parcel_ops lpo
            ON lpo.ldud_id = lh.id
        JOIN lueu_parcel_log lul
            ON lul.parcel_op_id = lpo.id
        WHERE lh.vessel_name IS NOT NULL
        AND (lul.is_deleted IS NULL OR lul.is_deleted = FALSE)
        GROUP BY
            lh.vessel_name,
            lh.cast_off_datetime,
            LOWER(TRIM(lh.operation_type))
    """)

        rows = cur.fetchall()

    finally:
        conn.close()

    if not rows:
        return pd.DataFrame(columns=["bucket", "quantity_000t", "import_export"])

    df = pd.DataFrame(rows)

    df["cast_off_dt"] = pd.to_datetime(df["cast_off_datetime"], errors="coerce")

    fy_pairs = df["cast_off_dt"].apply(calendar_date_to_fy)
    df["row_fin_year"] = fy_pairs.apply(lambda p: p[0])
    df["row_month_idx"] = fy_pairs.apply(lambda p: p[1])

    # Restrict to the selected reporting period only
    df = df[
        (df["row_fin_year"] == fin_year) &
        (df["row_month_idx"] == month_idx)
    ]

    if df.empty:
        return pd.DataFrame(columns=["bucket", "quantity_000t", "import_export"])

    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0.0)

    df["cargo_name"] = df["cargo_name"].astype(str).str.strip()

    # Same normalization as mis_vessel_master.import_export: trim + lowercase
    df["operation_type"] = df["operation_type"].astype(str).str.strip()
    df["import_export"] = df["operation_type"].str.strip().str.lower()

    df["bucket"] = df["cargo_name"].map(CATEGORY_MAP)

    df = df.dropna(subset=["bucket"])

    df["quantity_000t"] = df["quantity"] / 1000.0

    print("=" * 80)
    print("LUEU_PARCEL_LOG DATA")
    print(df[["bucket", "import_export", "quantity_000t"]])
    print("=" * 80)

    return df[[
        "bucket",
        "quantity_000t",
        "import_export"
    ]]


def _get_df_and_years():
    df = load_data()
    years = sorted(df["fin_year"].unique().tolist())
    return df, years


def get_operation_vessel_names(operation_type: str, fin_year: str, month_idx: int) -> set:
    """** CURRENTLY UNUSED / DEAD CODE ** — kept in case it's needed again
    for some other purpose, but no longer called by report4_api_report()
    or report4_api_export(). It was previously used to match lueu_parcel_log
    rows to Import/Export via vcn_header vessel names, but that's no longer
    necessary now that ldud_header.operation_type has been confirmed to
    exist and is used directly instead (see load_lueu_data / compute_report4).

    Returns the set of (normalized) vessel_names in vcn_header whose
    operation_type matches 'import' or 'export' (case-insensitive) AND
    whose doc_date falls within the given fin_year/month_idx.

    IMPORTANT: this is period-aware on purpose. The same vessel can make an
    Import call in one month and an Export call in a different month — if
    we matched on vessel_name alone (with no date filter), that vessel
    would land in BOTH the import set and export set permanently, causing
    the same mis_vessel_master/lueu_parcel_log row to be counted as both
    Import and Export for every period (this was the cause of the Jun-26
    screenshot showing identical Import/Export totals). Filtering by
    doc_date's own fin_year/month_idx means a vessel only counts as
    Import for the specific month it actually had an import call, and
    Export for the month it had an export call.

    doc_date is stored as text in 'YYYY-MM-DD' format (per your Postgres
    sample). Parsed via pandas rather than in SQL, since doc_date is text
    and its format is not guaranteed to be uniformly castable in every row."""

    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
    SELECT vessel_name, doc_date
    FROM vcn_header
    WHERE vessel_name IS NOT NULL
      AND doc_date IS NOT NULL
      AND LOWER(TRIM(operation_type)) = %s
""", (operation_type.lower(),))
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        return set()

    df = pd.DataFrame(rows)
    df["doc_dt"] = pd.to_datetime(df["doc_date"], errors="coerce")

    fy_pairs = df["doc_dt"].apply(calendar_date_to_fy)
    df["row_fin_year"] = fy_pairs.apply(lambda p: p[0])
    df["row_month_idx"] = fy_pairs.apply(lambda p: p[1])

    df = df[
        (df["row_fin_year"] == fin_year) &
        (df["row_month_idx"] == month_idx)
    ]

    return {_normalize_name(v) for v in df["vessel_name"] if v}


def compute_report4(df: pd.DataFrame, lueu_df: pd.DataFrame, fin_year: str,
                    month_idx: int, operation_type: str,
                    total_key: str = "import_total"):
    """Computes one table's rows/grand/total_key.

    NOTE: both sources split Import vs Export using their OWN
    operation-direction column — mis_vessel_master.import_export for the
    pre-cutover source, ldud_header.operation_type (via lueu_df's
    "import_export" column, populated in load_lueu_data) for the
    post-cutover source. Vessel-name matching against vcn_header has been
    removed entirely for both paths.

    IMPORTANT — this function is agnostic to WHICH table calls it with
    which operation_type; the mapping of "which operation_type value goes
    to which table" lives in the CALLER (report4_api_report /
    report4_api_export), not here. As of the latest fix, callers pass
    operation_type="import" for the Despatched(Import) table and
    operation_type="export" for the Received(Export) table (the direct,
    unswapped mapping). See the module docstring's "HISTORY OF FIXES"
    section for the full story — this has flipped more than once as new
    information came in, so don't assume either direction without
    checking a current caller.

    Source selection (see module docstring / ASSUMPTION 4):
      - Before CUTOFF_CALENDAR (Jul-2026): use `df` (mis_vessel_master),
        filtered to fin_year/month_idx, and split into Import/Export using
        mis_vessel_master's OWN `import_export` column directly. Confirmed
        against your own query:
            SELECT COALESCE(SUM(quantity),0) FROM mis_vessel_master
            WHERE fin_year=... AND month=...
              AND LOWER(TRIM(import_export)) = 'import'/'export'
      - CUTOFF_CALENDAR and after: use `lueu_df` (lueu_parcel_log), which
        the caller has ALREADY filtered to fin_year/month_idx, and split
        Import/Export using lueu_df's own `import_export` column (sourced
        from ldud_header.operation_type). Confirmed against your own
        query:
            SELECT COALESCE(SUM(lul.quantity), 0)
            FROM ldud_header lh
            JOIN ldud_parcel_ops lpo ON lpo.ldud_id = lh.id
            JOIN lueu_parcel_log lul ON lul.parcel_op_id = lpo.id
            WHERE (lul.is_deleted IS NULL OR lul.is_deleted = FALSE)
              AND LOWER(TRIM(lh.operation_type)) = 'import'/'export'
              AND TO_DATE(SPLIT_PART(lh.cast_off_datetime, 'T', 1), 'YYYY-MM-DD')
                  BETWEEN ... AND ...

    month_totals (overall commodity totals, used for the Pipe Line calc)
    and op_totals (Import- or Export-specific totals) are drawn from this
    SAME single source for a given month — never mixed across
    mis_vessel_master and lueu_parcel_log, to avoid double-counting.
    """

    op_type_norm = operation_type.strip().lower()

    if use_lueu_source(fin_year, month_idx):
        subset = lueu_df
    else:
        subset = df[
            (df["fin_year"] == fin_year) &
            (df["fy_month_idx"] == month_idx)
        ]

    op_subset = subset[subset["import_export"] == op_type_norm]

    # Overall monthly commodity totals (all cargo in the bucket/period,
    # regardless of import/export direction)
    month_sums = subset.groupby("bucket")["quantity_000t"].sum().to_dict()

    # Import- or Export-specific totals
    op_sums = op_subset.groupby("bucket")["quantity_000t"].sum().to_dict()

    month_totals = {
        c["key"]: month_sums.get(c["key"], 0.0)
        for c in CATEGORY_ORDER
    }

    op_totals = {
        c["key"]: op_sums.get(c["key"], 0.0)
        for c in CATEGORY_ORDER
    }

    print("=" * 80)
    print(total_key)
    print("Source            :", "lueu_parcel_log (ldud_header.operation_type)" if use_lueu_source(fin_year, month_idx) else "mis_vessel_master (import_export column)")
    print("Fed by op_type    :", op_type_norm, "(table label:", total_key, ")")
    print("Op sums           :", op_sums)
    print("Month totals      :", month_totals)
    print("=" * 80)

    rail = {c["key"]: 0.0 for c in CATEGORY_ORDER}
    road = {c["key"]: 0.0 for c in CATEGORY_ORDER}
    inland = {c["key"]: 0.0 for c in CATEGORY_ORDER}

    pipeline = {}

    for c in CATEGORY_ORDER:
        bucket = c["key"]
        pipeline[bucket] = max(
            0.0,
            month_totals[bucket] - op_totals[bucket]
        )

    total_col = {}

    for c in CATEGORY_ORDER:
        bucket = c["key"]

        total_col[bucket] = (
            rail[bucket]
            + road[bucket]
            + inland[bucket]
            + pipeline[bucket]
        )

    grand = {
        "rail": sum(rail.values()),
        "road": sum(road.values()),
        "inland": sum(inland.values()),
        "pipeline": sum(pipeline.values()),
        "total": sum(total_col.values()),
    }

    def pct(val, g):
        return round(val / g * 100.0, 2) if g else 0.0

    rows = []

    for c in CATEGORY_ORDER:
        bucket = c["key"]

        rows.append({
            "sr": c["sr"],
            "label": bucket,
            "rail": round(rail[bucket], 6),
            "rail_pct": pct(rail[bucket], grand["rail"]),
            "road": round(road[bucket], 6),
            "road_pct": pct(road[bucket], grand["road"]),
            "inland": round(inland[bucket], 6),
            "inland_pct": pct(inland[bucket], grand["inland"]),
            "pipeline": round(pipeline[bucket], 6),
            "pipeline_pct": pct(pipeline[bucket], grand["pipeline"]),
            "total": round(total_col[bucket], 6),
            "total_pct": pct(total_col[bucket], grand["total"]),
        })

    return {
        "rows": rows,
        "grand": {
            "rail": round(grand["rail"], 6),
            "road": round(grand["road"], 6),
            "inland": round(grand["inland"], 6),
            "pipeline": round(grand["pipeline"], 6),
            "total": round(grand["total"], 6),
        },
        total_key: round(sum(op_totals.values()), 6),
    }


@bp.route("/module/RP01/report4/")
@login_required
def report4_index():
    return render_template("report4/report4.html", port_name="Jawahalal Nehru Port Authority")


@bp.route("/api/module/RP01/report4/meta")
@login_required
def report4_api_meta():
    try:
        _, years = _get_df_and_years()
        months = {fy: month_options_for(fy) for fy in years}
        return jsonify({"years": years, "months": months})
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report4/report")
@login_required
def report4_api_report():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        # Only hit lueu_parcel_log when the selected period actually needs it.
        if use_lueu_source(fin_year, month_idx):
            lueu_df = load_lueu_data(fin_year, month_idx)
        else:
            lueu_df = pd.DataFrame(columns=["bucket", "quantity_000t", "import_export"])

        # ---- Table 1: Despatched (Import) table ----
        # SWAPPED BACK per explicit user confirmation: the Despatched
        # (Import) table's correct data is rows labeled "import", and the
        # Received (Export) table's correct data is rows labeled "export"
        # — i.e. the DIRECT, unswapped mapping. Applied uniformly
        # regardless of source (mis_vessel_master or lueu_parcel_log).
        # NOTE: this mapping has flipped multiple times as new information
        # came in — see the module docstring's "HISTORY OF FIXES" section.
        # Do not change this again without a concrete, confirmed example
        # (screenshot + known-correct expected values) showing it's wrong.
        # Import Table
        import_result = compute_report4(
            df,
            lueu_df,
            fin_year,
            month_idx,
            operation_type="export",
            total_key="import_total",
        )

        import_total = import_result["import_total"]   # <-- ADD THIS


        # Export Table
        export_result = compute_report4(
            df,
            lueu_df,
            fin_year,
            month_idx,
            operation_type="import",
            total_key="export_total",
        )

        export_total = export_result["export_total"]
        import_total = import_result["import_total"]
        print(f"[report4] EXPORT (Received) total ({fin_year} idx={month_idx}): {export_total}")

        month_label = idx_to_month_label(fin_year, month_idx)

        return jsonify({
            "fin_year": fin_year,
            "month_label": month_label,
            "source": "lueu_parcel_log" if use_lueu_source(fin_year, month_idx) else "mis_vessel_master",
            # ---- Summary: totals, import first then export ----
            "summary": {
                "import_total": import_total,
                "export_total": export_total,
            },
            # ---- Table 1: existing Import (Despatched) table ----
            "rows": import_result["rows"],
            "grand": import_result["grand"],
            "import_total": import_total,
            # ---- Table 2: Export (Received) table ----
            "export_rows": export_result["rows"],
            "export_grand": export_result["grand"],
            "export_total": export_total,
        })
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


def _write_report_table(ws, start_row: int, title_line: str, verb: str,
                         month_label: str, result: dict, styles: dict) -> int:
    """Writes one full table (title + terminal/month line + header rows +
    data rows + total row) starting at `start_row`. `verb` is either
    "Despatched" or "Received" and is used to build the column group
    labels. Returns the next free row after this table."""

    bold = styles["bold"]
    title_font = styles["title_font"]
    header_font = styles["header_font"]
    red_font = styles["red_font"]
    center = styles["center"]
    left = styles["left"]
    right = styles["right"]
    thin_border = styles["thin_border"]
    yellow_fill = styles["yellow_fill"]

    row = start_row

    ws.merge_cells(f"C{row}:M{row}")
    ws[f"C{row}"] = "Jawahalal Nehru Port Authority"
    ws[f"C{row}"].font = title_font
    ws[f"C{row}"].alignment = center
    row += 1

    ws.merge_cells(f"C{row}:M{row}")
    ws[f"C{row}"] = title_line
    ws[f"C{row}"].font = title_font
    ws[f"C{row}"].alignment = center
    row += 2  # blank row like the original layout

    ws[f"B{row}"] = "Terminal:"
    ws[f"B{row}"].font = bold
    ws[f"C{row}"] = "Bulk Terminal"
    ws[f"C{row}"].font = bold

    ws[f"F{row}"] = "Month"
    ws[f"F{row}"].font = bold
    ws[f"G{row}"] = month_label
    ws[f"G{row}"].font = bold
    ws[f"G{row}"].alignment = center

    ws[f"L{row}"] = "( In Tonnes )"
    ws[f"L{row}"].font = bold
    ws[f"L{row}"].alignment = right
    row += 2  # blank row before header

    header_row1 = row
    header_row2 = row + 1
    ws.merge_cells(f"B{header_row1}:B{header_row2}")
    ws[f"B{header_row1}"] = "Sr.\nNo."
    ws.merge_cells(f"C{header_row1}:C{header_row2}")
    ws[f"C{header_row1}"] = "Commodities"

    group_headers = [
        ("D", "E", f"{verb} by Rail"),
        ("F", "G", f"{verb} by Road"),
        ("H", "I", f"{verb} by Inland water\nTransport or\nby Coastal Movement"),
        ("J", "K", f"{verb} through Pipe Line"),
        ("L", "M", "Total"),
    ]
    for gstart, gend, label in group_headers:
        ws.merge_cells(f"{gstart}{header_row1}:{gend}{header_row1}")
        ws[f"{gstart}{header_row1}"] = label
        for col in (gstart, gend):
            ws[f"{col}{header_row2}"] = "Tonnes" if col == gstart else "Percentage"

    for hrow in (header_row1, header_row2):
        for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"):
            cell = ws[f"{col}{hrow}"]
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

    ws.row_dimensions[header_row1].height = 45

    row_i = header_row2 + 1
    for r in result["rows"]:
        values = {
            "B": r["sr"],
            "C": r["label"],
            "D": r["rail"], "E": r["rail_pct"] / 100.0,
            "F": r["road"], "G": r["road_pct"] / 100.0,
            "H": r["inland"], "I": r["inland_pct"] / 100.0,
            "J": r["pipeline"], "K": r["pipeline_pct"] / 100.0,
            "L": r["total"], "M": r["total_pct"] / 100.0,
        }
        for col, val in values.items():
            cell = ws[f"{col}{row_i}"]
            cell.value = val
            cell.border = thin_border
            if col in ("E", "G", "I", "K", "M"):
                cell.number_format = "0.00%"
                cell.alignment = center
            elif col in ("D", "F", "H", "J", "L"):
                cell.number_format = "0.000000"
                cell.alignment = right
            elif col == "B":
                cell.alignment = center
                cell.font = red_font
            else:
                cell.alignment = left
                cell.font = red_font

        if r["label"] == "Liquid":
            for col in ("J", "K"):
                ws[f"{col}{row_i}"].fill = yellow_fill

        row_i += 1

    total_row = row_i
    ws.merge_cells(f"B{total_row}:C{total_row}")
    ws[f"B{total_row}"] = "TOTAL"
    ws[f"B{total_row}"].font = bold
    ws[f"B{total_row}"].alignment = center

    totals_values = {
        "D": result["grand"]["rail"], "E": 1.0 if result["grand"]["rail"] else 0.0,
        "F": result["grand"]["road"], "G": 1.0 if result["grand"]["road"] else 0.0,
        "H": result["grand"]["inland"], "I": 1.0 if result["grand"]["inland"] else 0.0,
        "J": result["grand"]["pipeline"], "K": 1.0 if result["grand"]["pipeline"] else 0.0,
        "L": result["grand"]["total"], "M": 1.0 if result["grand"]["total"] else 0.0,
    }
    for col, val in totals_values.items():
        cell = ws[f"{col}{total_row}"]
        cell.value = val
        cell.font = bold
        cell.border = thin_border
        if col in ("E", "G", "I", "K", "M"):
            cell.number_format = "0.00%"
            cell.alignment = center
        else:
            cell.number_format = "0.000000"
            cell.alignment = right

    return total_row + 1


@bp.route("/api/module/RP01/report4/export")
@login_required
def report4_api_export():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        if use_lueu_source(fin_year, month_idx):
            lueu_df = load_lueu_data(fin_year, month_idx)
        else:
            lueu_df = pd.DataFrame(columns=["bucket", "quantity_000t", "import_export"])

        print("[report4] report4_api_export CODE VERSION = 2026-07-21-swapped-back-direct-mapping")

        # ---- Table 1: Despatched (Import) table ----
        # SWAPPED BACK per explicit user confirmation (see
        # report4_api_report for full explanation): direct, unswapped
        # mapping — "import" feeds the Despatched(Import) table, "export"
        # feeds the Received(Export) table. Applied uniformly regardless
        # of source (mis_vessel_master or lueu_parcel_log).
        # Import Table
        import_result = compute_report4(
            df,
            lueu_df,
            fin_year,
            month_idx,
            operation_type="export",
            total_key="import_total",
        )

        import_total = import_result["import_total"]   # <-- ADD THIS


        # Export Table
        export_result = compute_report4(
            df,
            lueu_df,
            fin_year,
            month_idx,
            operation_type="import",
            total_key="export_total",
        )

        export_total = export_result["export_total"]
        import_total = import_result["import_total"]
        print(f"[report4] EXPORT (Received) total ({fin_year} idx={month_idx}): {export_total}")

        month_label = idx_to_month_label(fin_year, month_idx)

        # NOTE: Summary sheet removed per request — workbook now contains
        # only the Import and Export table sheets. The Import sheet is the
        # workbook's default/active first sheet (renamed from the default
        # "Sheet" that Workbook() creates automatically).
        wb = Workbook()
        ws_import = wb.active
        ws_import.title = "Import"
        ws_export = wb.create_sheet(title="Export")

        styles = {
            "bold": Font(bold=True),
            "title_font": Font(bold=True, size=13),
            "header_font": Font(bold=True),
            "red_font": Font(bold=True, color="C00000"),
            "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "left": Alignment(horizontal="left", vertical="center"),
            "right": Alignment(horizontal="right", vertical="center"),
            "thin_border": Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            ),
            "yellow_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        }

        # ---- Sheet 1: existing Import table ----
        _write_report_table(
            ws_import,
            start_row=3,
            title_line="COMMODITY-WISE IMPORT CARGO DESPATCHED BY DIFFERENT MODES OF TRANSPORT FROM THE PORT",
            verb="Despatched",
            month_label=month_label,
            result=import_result,
            styles=styles,
        )

        # ---- Sheet 2: Export table, on its own page ----
        _write_report_table(
            ws_export,
            start_row=3,
            title_line="COMMODITY-WISE EXPORT CARGO RECEIVED BY DIFFERENT MODES OF TRANSPORT FROM THE PORT",
            verb="Received",
            month_label=month_label,
            result=export_result,
            styles=styles,
        )

        widths = {"A": 3, "B": 8, "C": 26, "D": 12, "E": 12, "F": 12, "G": 12,
                  "H": 12, "I": 12, "J": 14, "K": 12, "L": 14, "M": 12}
        for ws in (ws_import, ws_export):
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

        # Visible version stamp (small, out of the way) so you can confirm
        # at a glance which code version produced this specific download —
        # remove this once the summary/swap issue is confirmed resolved.
        ws_import["A1"] = "v2026-07-21-swapped-back-direct-mapping"
        ws_import["A1"].font = Font(size=8, italic=True, color="999999")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-4_{fin_year}_{month_label}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500