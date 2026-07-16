"""
Report-1 — Principal Commodity Report
Flask Blueprint version.

Primary source: mis_vessel_master (Postgres).
Fallback source: for any (fin_year, month) that has ZERO rows in
mis_vessel_master, figures are pulled instead from the live LUEU01
logging pipeline (lueu_parcel_log -> ldud_parcel_ops) so the current
month can show real data even before that month's mis_vessel_master
upload has been done. mis_vessel_master always wins for any period
where it actually has rows.
"""

import io
import traceback
from functools import wraps
from datetime import datetime, date

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, get_cursor

from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
CAL_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

CATEGORY_ORDER = [
    {"key": "POL (Crude, Products and LPG/LNG)", "sub": False},
    {"key": "Other liquids", "sub": False},
    {"key": "Iron ore incl.Iron ore pallets", "sub": False},
    {"key": "Fertilizers- Finished", "sub": False},
    {"key": "Fertilizers -Raw Material (PH ACID)", "sub": False},
    {"key": "Thermal and Steam Coal", "sub": False},
    {"key": "Cooking coal and other coals", "sub": False},
    {"key": "Containers- Tonnage", "sub": False},
    {"key": "Containers- TEUs", "sub": False},
    {"key": "Other Misc. cargo", "sub": False},
    {"key": "A. Cement", "sub": True},
    {"key": "B. Break Bulk/General cargo", "sub": True},
]

# NOTE: this maps mis_vessel_master.category values -> the report's
# principal-commodity buckets above. mis_vessel_master also has
# `category1` and `new_cat` columns — if `category` values below don't
# actually match what's in your data, tell me which column/values to use
# and I'll adjust this map.
CATEGORY_MAP = {
    "POL": "POL (Crude, Products and LPG/LNG)",
    "POL Black": "POL (Crude, Products and LPG/LNG)",
    "Other Liquid": "Other liquids",
    "Edible Oil": "Other liquids",
    "Chemical": "Other liquids",
    "Ph.Acid": "Fertilizers -Raw Material (PH ACID)",
}

# Maps free-text ldud_parcel_ops.cargo_name (from the live LUEU01 logging
# pipeline, used as a fallback source) straight to Report-1's bucket
# labels. Add entries here as new cargo names show up in that pipeline —
# anything not listed gets dropped (with a console warning) rather than
# crashing the report.
LIVE_CARGO_TO_BUCKET = {

    # =========================
    # POL (Crude, Products and LPG/LNG)
    # =========================
    "FURNACE OIL": "POL (Crude, Products and LPG/LNG)",
    "CARBAN BLACK FEED STOCK": "POL (Crude, Products and LPG/LNG)",

    # =========================
    # Other liquids
    # =========================
    "(SQ100HS)": "Other liquids",
    "ACETIC ACID": "Other liquids",
    "ACETONE": "Other liquids",
    "ARAMCO ALTRA 4 (SS100H)": "Other liquids",
    "ARAMCO ALTRA 6 (SS150H)": "Other liquids",
    "BASE OIL": "Other liquids",
    "BASE OIL 4CST": "Other liquids",
    "BASE OIL 6CST": "Other liquids",
    "BASE OIL ARAMCO PRIMA 150": "Other liquids",
    "BASE OIL ARAMCO PRIMA 500": "Other liquids",
    "BASE OIL KIXX LUBO 150N": "Other liquids",
    "BASE OIL KIXX LUBO 600N": "Other liquids",
    "BUTYL ACETATE": "Other liquids",
    "GLYCERINE": "Other liquids",
    "GTL BASE OIL QHVI 4": "Other liquids",
    "HVI 120": "Other liquids",
    "HVI 650": "Other liquids",
    "ISOPROPYL ALCOHOL": "Other liquids",
    "LUBE OIL": "Other liquids",
    "MDC": "Other liquids",
    "METHANOL": "Other liquids",
    "METHELENE CHOLORIDE": "Other liquids",
    "METHYL ETHYL KETONE": "Other liquids",
    "N BUTANOL": "Other liquids",
    "N BUTYL ACETATE": "Other liquids",
    "NBA": "Other liquids",
    "NITRIC ACID": "Other liquids",
    "OLEIC ACID": "Other liquids",
    "PHENOL": "Other liquids",
    "SHELL 150 N(DAESAN)": "Other liquids",
    "SHELL 150N": "Other liquids",
    "SHELL 500N": "Other liquids",
    "STYRENE MONOMER": "Other liquids",
    "TOULENE": "Other liquids",
    "VINYL ACETATE MONOMER": "Other liquids",

    # =========================
    # Edible Oil (Report-1 groups these under Other liquids)
    # =========================
    "CRUDE DEGUMMED SOYABEAN OIL": "Other liquids",
    "CRUDE PALM KERNEL OIL": "Other liquids",
    "CRUDE PALM KERNEL OIL- EDIBLE GRADE": "Other liquids",
    "CRUDE PALM OIL": "Other liquids",
    "CRUDE PALM OIL - EDIBLE GRADE": "Other liquids",
    "CRUDE PALM OIL - MB": "Other liquids",
    "CRUDE SUNFLOWER SEED OIL": "Other liquids",
    "RBD PALM OLEIN": "Other liquids",
    "REFINED GLYCERINE": "Other liquids",
    "SUNFLOWER OIL": "Other liquids",

    # =========================
    # Fertilizers - Raw Material (PH ACID)
    # =========================
    "PHOSPHORIC ACID": "Fertilizers -Raw Material (PH ACID)",
}


class ReportDataError(Exception):
    """Raised for any problem loading/validating the report's source data.
    Caught by the route handlers and turned into a clean JSON error response."""
    pass


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


def month_str_to_idx(month_str: str) -> int:
    """'Apr-26' -> 0, 'Dec-24' -> 8, etc. Matches MONTH_NAMES order (FY Apr..Mar)."""
    abbrev = str(month_str).split("-")[0].strip()
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized value in mis_vessel_master.month: '{month_str}' "
            f"(expected something like 'Apr-26')"
        )


def _entry_date_to_fy_month(d):
    """Real calendar date (from lueu_parcel_log.entry_date) -> (fin_year,
    fy_month_idx), using the same Apr-Mar FY convention as the rest of the
    report. e.g. 2026-07-12 -> ('2026-27', 3)."""
    if isinstance(d, date):
        dt = d
    else:
        dt = datetime.strptime(str(d)[:10], "%Y-%m-%d").date()

    if dt.month >= 4:
        fy_start = dt.year
    else:
        fy_start = dt.year - 1
    fin_year = f"{fy_start}-{str(fy_start + 1)[-2:]}"

    mn = CAL_MONTH_ABBR[dt.month - 1]
    fy_month_idx = MONTH_NAMES.index(mn)
    return fin_year, fy_month_idx


def _classify_live_cargo(cargo_name):
    if not cargo_name:
        return None

    conn = get_db()
    try:
        cur = get_cursor(conn)

        cur.execute("""
            SELECT cargo_category
            FROM vessel_cargo
            WHERE UPPER(TRIM(cargo_name)) = UPPER(TRIM(%s))
            LIMIT 1
        """, (cargo_name,))

        row = cur.fetchone()

        if not row:
            return None

        category = str(row["cargo_category"]).strip().upper()

        CATEGORY_LOOKUP = {
            "POL": "POL (Crude, Products and LPG/LNG)",
            "POL-BLACK": "POL (Crude, Products and LPG/LNG)",
            "OTHER LIQUID": "Other liquids",
            "OTHER LIQUIDS": "Other liquids",
            "EDIBLE OIL": "Other liquids",
            "FERTILIZERS": "Fertilizers -Raw Material (PH ACID)",
        }

        return CATEGORY_LOOKUP.get(category)

    finally:
        conn.close()


def _load_live_pipeline_data():
    """Fallback source: real-time LUEU01 logging pipeline
    (lueu_parcel_log -> ldud_parcel_ops). Only used to fill in months that
    have ZERO rows in mis_vessel_master — see load_data() below."""
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT l.entry_date, po.cargo_name, l.quantity
            FROM lueu_parcel_log l
            JOIN ldud_parcel_ops po ON po.id = l.parcel_op_id
            WHERE l.is_deleted IS NOT TRUE
              AND l.entry_date IS NOT NULL
              AND l.quantity IS NOT NULL
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    empty = pd.DataFrame(columns=["fin_year", "fy_month_idx", "cargo_sub_category", "quantity_000t"])
    if not rows:
        return empty

    df = pd.DataFrame(rows)
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0.0)

    fy_list, idx_list = [], []
    for d in df["entry_date"]:
        fy, idx = _entry_date_to_fy_month(d)
        fy_list.append(fy)
        idx_list.append(idx)
    df["fin_year"] = fy_list
    df["fy_month_idx"] = idx_list

    df["cargo_sub_category"] = df["cargo_name"].apply(_classify_live_cargo)

    unmapped = sorted(df.loc[df["cargo_sub_category"].isna(), "cargo_name"].dropna().unique().tolist())
    if unmapped:
        print("REPORT1 WARNING: live-pipeline cargo_name values with no bucket mapping, dropped:", unmapped)

    df = df.dropna(subset=["cargo_sub_category"])
    if df.empty:
        return empty

    df["quantity_000t"] = df["quantity"] / 1000.0

    return (
        df.groupby(["fin_year", "fy_month_idx", "cargo_sub_category"], as_index=False)["quantity_000t"]
        .sum()
    )


def load_data() -> pd.DataFrame:
    """Primary source: mis_vessel_master. For any (fin_year, month) that has
    NO rows at all in mis_vessel_master, fall back to the live LUEU01
    logging pipeline for that period only -- mis_vessel_master always wins
    where it has data."""
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT fin_year, month, category, quantity
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month IS NOT NULL
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        mv_df = pd.DataFrame(columns=["fin_year", "fy_month_idx", "cargo_sub_category", "quantity_000t"])
    else:
        mv_df = pd.DataFrame(rows)

        missing_cols = [c for c in ("fin_year", "month", "category", "quantity") if c not in mv_df.columns]
        if missing_cols:
            raise ReportDataError(f"Query result is missing column(s): {', '.join(missing_cols)}")

        mv_df["fin_year"] = mv_df["fin_year"].str.strip()
        mv_df["category"] = mv_df["category"].astype(str).str.strip()
        mv_df["quantity"] = pd.to_numeric(mv_df["quantity"], errors="coerce").fillna(0.0)
        mv_df["fy_month_idx"] = mv_df["month"].apply(month_str_to_idx)
        mv_df["cargo_sub_category"] = mv_df["category"].map(CATEGORY_MAP)

        unmapped = sorted(mv_df.loc[mv_df["cargo_sub_category"].isna(), "category"].unique().tolist())
        if unmapped:
            print("REPORT1 WARNING: unmapped mis_vessel_master category values dropped:", unmapped)

        mv_df = mv_df.dropna(subset=["cargo_sub_category"])
        mv_df["quantity_000t"] = mv_df["quantity"] / 1000.0
        mv_df = mv_df[["fin_year", "fy_month_idx", "cargo_sub_category", "quantity_000t"]]

    # ---- which (fin_year, month) periods does mis_vessel_master actually
    # cover? Only periods with ZERO rows there fall back to the live pipeline. ----
    covered_periods = set(zip(mv_df["fin_year"], mv_df["fy_month_idx"]))

    live_df = _load_live_pipeline_data()
    if not live_df.empty:
        live_df = live_df[
            ~live_df.apply(lambda r: (r["fin_year"], r["fy_month_idx"]) in covered_periods, axis=1)
        ]

    combined = pd.concat([mv_df, live_df], ignore_index=True)

    if combined.empty:
        raise ReportDataError(
            "No usable rows found in mis_vessel_master or the live LUEU01 pipeline."
        )

    # re-aggregate in case both sources ever contributed to the same
    # (fin_year, fy_month_idx, bucket) -- shouldn't happen given the period
    # filter above, but keeps totals correct if it ever does
    combined = (
        combined.groupby(["fin_year", "fy_month_idx", "cargo_sub_category"], as_index=False)["quantity_000t"]
        .sum()
    )

    return combined


def _get_df_and_years():
    df = load_data()
    years = sorted(df["fin_year"].unique().tolist())
    return df, years


def compute_totals(df, fin_year: str, month_idx: int):
    subset = df[df["fin_year"] == fin_year]
    for_month = subset[subset["fy_month_idx"] == month_idx]
    upto_month = subset[subset["fy_month_idx"] <= month_idx]

    for_month_sums = for_month.groupby("cargo_sub_category")["quantity_000t"].sum().to_dict()
    upto_month_sums = upto_month.groupby("cargo_sub_category")["quantity_000t"].sum().to_dict()

    for_month_out = {c["key"]: round(for_month_sums.get(c["key"], 0.0), 6) for c in CATEGORY_ORDER}
    upto_month_out = {c["key"]: round(upto_month_sums.get(c["key"], 0.0), 6) for c in CATEGORY_ORDER}

    return {
        "for_month": for_month_out,
        "upto_month": upto_month_out,
        "total_for_month": round(sum(for_month_out.values()), 6),
        "total_upto_month": round(sum(upto_month_out.values()), 6),
    }


@bp.route("/module/RP01/report1/")
@login_required
def report1_index():
    return render_template("report1/report1.html", port_name="")


@bp.route("/api/module/RP01/report1/meta")
@login_required
def api_meta():
    try:
        _, years = _get_df_and_years()
        months = {fy: month_options_for(fy) for fy in years}
        return jsonify({"years": years, "months": months, "port_name": ""})
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report1/report")
@login_required
def api_report():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        totals = compute_totals(df, fin_year, month_idx)
        month_label = next(o["label"] for o in month_options_for(fin_year) if o["idx"] == month_idx)

        rows = []
        sr = 1
        for c in CATEGORY_ORDER:
            rows.append({
                "sr": None if c["sub"] else sr,
                "label": c["key"],
                "sub": c["sub"],
                "for_month": totals["for_month"][c["key"]],
                "upto_month": totals["upto_month"][c["key"]],
            })
            if not c["sub"]:
                sr += 1

        subset = df[df["fin_year"] == fin_year]
        debug_info = {
            "received_fin_year": fin_year,
            "received_month_idx": month_idx,
            "distinct_month_idx_present_for_fin_year": sorted(subset["fy_month_idx"].unique().tolist()),
            "row_count_matching_month_idx": int((subset["fy_month_idx"] == month_idx).sum()),
            "row_count_upto_month_idx": int((subset["fy_month_idx"] <= month_idx).sum()),
        }
        print("REPORT1 DEBUG:", debug_info)

        return jsonify({
            "port_name": "",
            "fin_year": fin_year,
            "month_label": month_label,
            "rows": rows,
            "total_for_month": totals["total_for_month"],
            "total_upto_month": totals["total_upto_month"],
            "debug": debug_info,
        })
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report1/export")
@login_required
def api_export():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        port_name = request.args.get("port_name", "JNPA")
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        totals = compute_totals(df, fin_year, month_idx)
        month_label = next(o["label"] for o in month_options_for(fin_year) if o["idx"] == month_idx)

        wb = Workbook()
        ws = wb.active
        ws.title = "Report-1"

        # ---- styles -------------------------------------------------
        bold = Font(bold=True)
        bold_blue = Font(bold=True, color="1F4E78")
        title_font = Font(bold=True, underline="single", color="1F4E78", size=12)
        label_font = Font(bold=True, color="1F4E78")
        commodity_font = Font(color="7B241C")  # maroon, matches non-sub commodity rows
        header_font = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        thick_green = Side(style="medium", color="1E7145")
        title_border = Border(left=thick_green, right=thick_green, top=thick_green, bottom=thick_green)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # ---- row 2: appendix label -----------------------------------
        ws["E2"] = "Appendix: 2"
        ws["E2"].font = bold
        ws["E2"].alignment = right

        # ---- row 4: title box ------------------------------------------
        ws.merge_cells("B4:E4")
        ws["B4"] = "TRAFFIC HANDLED BY PRINCIPLE COMMODITIES (PROVISIONAL)"
        ws["B4"].font = title_font
        ws["B4"].alignment = center
        for col in ("B", "C", "D", "E"):
            ws[f"{col}4"].border = title_border

        # ---- rows 6-8: port / year / period -----------------------------
        ws.merge_cells("C6:D6")
        ws["C6"] = "NAME OF THE PORT:"
        ws["C6"].font = label_font
        ws["C6"].alignment = center
        ws["E6"] = port_name
        ws["E6"].font = bold
        ws["E6"].alignment = center

        ws.merge_cells("C7:D7")
        ws["C7"] = "YEAR :"
        ws["C7"].font = label_font
        ws["C7"].alignment = center
        ws["E7"] = fin_year
        ws["E7"].font = bold
        ws["E7"].alignment = center

        ws.merge_cells("C8:D8")
        ws["C8"] = "PERIOD: FOR THE MONTH OF"
        ws["C8"].font = label_font
        ws["C8"].alignment = center
        ws["E8"] = month_label
        ws["E8"].font = bold
        ws["E8"].alignment = center

        # ---- row 9: units label ------------------------------------------
        ws["E9"] = "('000 TONNES)"
        ws["E9"].alignment = right

        # ---- row 10: table header -----------------------------------------
        header_row = 10
        headers = {"B": "SR. NO.", "C": "PRINCIPLE COMMODITY", "D": "FOR THE MONTH", "E": "UP TO MONTH"}
        for col, h in headers.items():
            cell = ws[f"{col}{header_row}"]
            cell.value = h
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

        # ---- rows 11-22: data --------------------------------------------
        row_i = header_row + 1
        sr = 1
        for c in CATEGORY_ORDER:
            label = c["key"]
            for_m = totals["for_month"][label]
            upto_m = totals["upto_month"][label]

            b_cell = ws[f"B{row_i}"]
            c_cell = ws[f"C{row_i}"]
            d_cell = ws[f"D{row_i}"]
            e_cell = ws[f"E{row_i}"]

            b_cell.value = None if c["sub"] else sr
            c_cell.value = ("    " + label) if c["sub"] else label
            d_cell.value = for_m
            e_cell.value = upto_m
            d_cell.number_format = "0.000000"
            e_cell.number_format = "0.000000"

            b_cell.alignment = center
            c_cell.alignment = left
            d_cell.alignment = right
            e_cell.alignment = right

            b_cell.font = bold_blue
            c_cell.font = commodity_font
            d_cell.font = commodity_font
            e_cell.font = commodity_font

            for cell in (b_cell, c_cell, d_cell, e_cell):
                cell.border = thin_border

            # highlight rows that actually have data, like the sample sheet
            if for_m or upto_m:
                for cell in (b_cell, c_cell, d_cell, e_cell):
                    cell.fill = yellow_fill

            if not c["sub"]:
                sr += 1
            row_i += 1

        # ---- row 23: total --------------------------------------------
        total_row = row_i
        ws[f"C{total_row}"] = "Total"
        ws[f"C{total_row}"].font = bold
        ws[f"C{total_row}"].alignment = center
        ws[f"D{total_row}"] = totals["total_for_month"]
        ws[f"D{total_row}"].font = bold
        ws[f"D{total_row}"].alignment = right
        ws[f"D{total_row}"].number_format = "0.000000"
        ws[f"E{total_row}"] = totals["total_upto_month"]
        ws[f"E{total_row}"].font = bold
        ws[f"E{total_row}"].alignment = right
        ws[f"E{total_row}"].number_format = "0.000000"
        for col in ("B", "C", "D", "E"):
            ws[f"{col}{total_row}"].border = thin_border

        # ---- note row (2 blank rows after total) --------------------------
        note_row = total_row + 2
        ws.merge_cells(f"B{note_row}:E{note_row}")
        ws[f"B{note_row}"] = "Note: Other liquids Include chemicals, edible oil, molasses etc."
        ws[f"B{note_row}"].alignment = left

        # ---- signature (2 blank rows after note) --------------------------
        sign_row = note_row + 3
        ws[f"E{sign_row}"] = "Sr. Manager (Traffic)"
        ws[f"E{sign_row}"].font = bold
        ws[f"E{sign_row}"].alignment = center

        # ---- column widths -------------------------------------------------
        widths = {"A": 3, "B": 10, "C": 42, "D": 18, "E": 18}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-1_{fin_year}_{month_label}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500