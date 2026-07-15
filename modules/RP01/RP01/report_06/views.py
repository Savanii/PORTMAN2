"""
RP01 Report-1 — Overseas/Coastal Cargo Traffic Handled (monthly, by commodity)

DATA SOURCES:
  1) mis_vessel_master — reconciled monthly MIS table (unchanged, see below).
     Used for any month that has already been reconciled/uploaded.

  2) LIVE FALLBACK (NEW) — when the requested month has zero rows in
     mis_vessel_master (typically the current, not-yet-reconciled month),
     we build the same shape of data live from the operational chain:

         vcn_header (1)
           -> ldud_header      (via ldud_header.vcn_id = vcn_header.id)
             -> ldud_parcel_ops  (via ldud_parcel_ops.ldud_id = ldud_header.id)
               -> lueu_parcel_log  (via lueu_parcel_log.parcel_op_id = ldud_parcel_ops.id)

     Field mapping confirmed against live schema (2026-07):
       overseas_coastal -> vcn_header.vessel_run_type
                            'Foreign' -> 'Overseas', 'Costal' -> 'Costal'
       import_export     -> vcn_header.operation_type ('Import' / 'Export')
       category           -> vessel_cargo.cargo_category, matched against
                              vcn_header.cargo_type. Match is NOT always exact
                              (e.g. cargo_type 'BASE OIL' has no exact row in
                              vessel_cargo, and some cargo_type values contain
                              multiple cargoes comma-separated, e.g.
                              'FO [E], FURNACE OIL'). We therefore try an exact
                              match first, then a fuzzy LIKE match, and log
                              anything still unresolved into '_debug.unmapped_categories'
                              same as the reconciled path -> falls to OTHER BULK.
       quantity            -> SUM(lueu_parcel_log.quantity), raw MT

     KNOWN GAP — foreign_indian (IF/FF split):
       No table in the live operational schema captures vessel flag
       (Indian-registered vs foreign-registered). This is confirmed absent
       from vcn_header, ldud_header, ldud_parcel_ops, lueu_parcel_log,
       vcn_consigners, and vessel_cargo as of 2026-07-15.
       Until this is captured somewhere (or a new column is added), all live
       rows are defaulted to foreign_indian='I' and the response includes
       '_debug.foreign_indian_is_placeholder = True' so this is never
       mistaken for confirmed data. If it turns out this is entered manually
       somewhere before month-end reconciliation, update _fetch_live_rows()
       below to source it from there instead of defaulting it.

  mis_vessel_master fields (unchanged):
    overseas_coastal  -> 'Overseas' or 'Costal'  (NOTE: 'Costal' is the
                          actual spelling stored in this table — a typo
                          in the source system, matched here as-is)
    foreign_indian     -> 'I' = Indian Flag, 'F' = Foreign Flag
    import_export       -> 'Import' or 'Export'
    quantity            -> numeric, raw MT (divided by 1000 for '000 Tonnes)
    category            -> cargo category text; mapped below to this
                            report's 10 commodity buckets
    month                -> text like 'Jun-26' (Mon-YY), NOT a real date
                            column — matched by exact string, built from
                            the requested 'YYYY-MM' filter

CATEGORY MAPPING (confirm/adjust CATEGORY_MAP below if wrong):
  POL          -> POL-PRODUCTS
  POL Black    -> POL-CRUDE
  Edible Oil   -> EDIBLE OIL
  Other Liquid -> OTHER LIQUIDS
  Chemical     -> OTHER LIQUIDS
  Ph.Acid      -> OTHER LIQUIDS
  (anything else, including NULL) -> OTHER BULK
  No current data maps to LPG / FARM LIQUIDS / MOLASSES / CEMENT /
  CONTAINER — those rows will correctly show 0.00 until such cargo exists.
"""

from functools import wraps

from flask import request, jsonify, render_template, session, redirect, url_for, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .. import bp   # shared RP01 blueprint
from database import get_db, get_cursor, get_user_permissions

MODULE_CODE = 'RP01'

COMMODITIES = [
    'POL-CRUDE', 'POL-PRODUCTS', 'LPG', 'OTHER LIQUIDS', 'FARM LIQUIDS',
    'EDIBLE OIL', 'MOLASSES', 'CEMENT', 'OTHER BULK', 'CONTAINER',
]

CATEGORY_MAP = {
    'POL': 'POL-PRODUCTS',
    'POL BLACK': 'POL-PRODUCTS',
    'EDIBLE OIL': 'EDIBLE OIL',
    'OTHER LIQUID': 'OTHER LIQUIDS',
    'CHEMICAL': 'OTHER LIQUIDS',
    'PH.ACID': 'OTHER LIQUIDS',
}

_MONTH_ABBR = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_perms():
    if session.get('is_admin'):
        return {'can_read': 1, 'can_add': 1, 'can_edit': 1, 'can_delete': 1}
    return get_user_permissions(session.get('user_id'), MODULE_CODE)


@bp.route("/module/RP01/report_06/")
@login_required
def report1_page():
    perms = get_perms()
    if not perms.get("can_read"):
        return render_template("no_access.html"), 403
    return render_template("report_06/report_06.html", permissions=perms)


def _month_label(month_str):
    """
    'YYYY-MM' (e.g. '2026-06') -> 'Mon-YY' (e.g. 'Jun-26') matching the
    literal text stored in mis_vessel_master.month.
    """
    y, m = month_str.split('-')
    y, m = int(y), int(m)
    return f"{_MONTH_ABBR[m - 1]}-{str(y)[2:]}"


def _empty_bucket():
    return {
        'ov_imp_if': 0.0, 'ov_imp_ff': 0.0, 'ov_exp_if': 0.0, 'ov_exp_ff': 0.0,
        'co_imp_if': 0.0, 'co_imp_ff': 0.0, 'co_exp_if': 0.0, 'co_exp_ff': 0.0,
    }


def _aggregate(lines, debug=False, month_label=None, foreign_indian_placeholder=False):
    """
    Shared aggregation logic — takes normalized line dicts with keys:
      overseas_coastal, foreign_indian, import_export, category, quantity
    and produces the rows/totals structure used by both data sources.
    """
    matrix = {c: _empty_bucket() for c in COMMODITIES}
    lines_with_qty = 0
    unmapped_categories = set()

    for ln in lines:
        qty_mt = float(ln.get('quantity') or 0)
        if qty_mt <= 0:
            continue
        lines_with_qty += 1
        qty_k = qty_mt / 1000.0   # tonnes -> '000 Tonnes

        cat_raw = (ln.get('category') or '').strip().upper()
        commodity = CATEGORY_MAP.get(cat_raw)
        if commodity is None:
            if cat_raw:
                unmapped_categories.add(ln.get('category'))
            commodity = 'OTHER BULK'

        oc = (ln.get('overseas_coastal') or '').strip().lower()
        is_coastal = oc.startswith('cost') or oc.startswith('coast')  # 'Costal' typo-safe

        fi = (ln.get('foreign_indian') or '').strip().upper()
        is_indian = fi == 'I'

        ie = (ln.get('import_export') or '').strip().lower()
        is_export = ie == 'export'

        key = ('co' if is_coastal else 'ov') + ('_exp_' if is_export else '_imp_') + ('if' if is_indian else 'ff')
        matrix[commodity][key] += qty_k

    rows = []
    grand = _empty_bucket()
    for c in COMMODITIES:
        b = matrix[c]
        ov_total = b['ov_imp_if'] + b['ov_imp_ff'] + b['ov_exp_if'] + b['ov_exp_ff']
        co_total = b['co_imp_if'] + b['co_imp_ff'] + b['co_exp_if'] + b['co_exp_ff']
        rows.append({
            'commodity': c,
            'ov_imp_if': round(b['ov_imp_if'], 2), 'ov_imp_ff': round(b['ov_imp_ff'], 2),
            'ov_exp_if': round(b['ov_exp_if'], 2), 'ov_exp_ff': round(b['ov_exp_ff'], 2),
            'ov_total': round(ov_total, 2),
            'co_imp_if': round(b['co_imp_if'], 2), 'co_imp_ff': round(b['co_imp_ff'], 2),
            'co_exp_if': round(b['co_exp_if'], 2), 'co_exp_ff': round(b['co_exp_ff'], 2),
            'co_total': round(co_total, 2),
            'grand_total': round(ov_total + co_total, 2),
        })
        for k in grand:
            grand[k] += b[k]

    gov_total = grand['ov_imp_if'] + grand['ov_imp_ff'] + grand['ov_exp_if'] + grand['ov_exp_ff']
    gco_total = grand['co_imp_if'] + grand['co_imp_ff'] + grand['co_exp_if'] + grand['co_exp_ff']
    totals = {
        'ov_imp_if': round(grand['ov_imp_if'], 3), 'ov_imp_ff': round(grand['ov_imp_ff'], 3),
        'ov_exp_if': round(grand['ov_exp_if'], 3), 'ov_exp_ff': round(grand['ov_exp_ff'], 3),
        'ov_total': round(gov_total, 3),
        'co_imp_if': round(grand['co_imp_if'], 3), 'co_imp_ff': round(grand['co_imp_ff'], 3),
        'co_exp_if': round(grand['co_exp_if'], 3), 'co_exp_ff': round(grand['co_exp_ff'], 3),
        'co_total': round(gco_total, 3),
        'grand_total': round(gov_total + gco_total, 3),
    }

    debug_block = None
    if debug:
        debug_block = {
            'month_label_matched': month_label,
            'rows_fetched': len(lines),
            'lines_with_qty_gt_0': lines_with_qty,
            'unmapped_categories': list(unmapped_categories),
        }
        if foreign_indian_placeholder:
            debug_block['foreign_indian_is_placeholder'] = True
            debug_block['foreign_indian_note'] = (
                "Live source has no vessel-flag column; all rows defaulted "
                "to 'I' until this is confirmed or captured upstream."
            )

    return rows, totals, debug_block


def _fetch_reconciled_rows(month_label):
    """mis_vessel_master path — unchanged."""
    conn = get_db()
    cur = get_cursor(conn)
    cur.execute("""
        SELECT overseas_coastal, foreign_indian, import_export, category, quantity
        FROM mis_vessel_master
        WHERE month = %s
    """, [month_label])
    lines = [dict(r) for r in cur.fetchall()]
    conn.close()
    return lines


def _fetch_live_rows(year_str, month_str):
    """
    LIVE fallback for the current / not-yet-reconciled month.
    Reads from vcn_header -> ldud_header -> ldud_parcel_ops -> lueu_parcel_log.

    NOTE: foreign_indian is not available in this chain (see module docstring)
    and is defaulted to 'I' by the caller via _aggregate's placeholder flag.
    category is matched against vessel_cargo.cargo_category with an exact
    match first, falling back to a fuzzy LIKE match.
    """
    y, m = month_str.split('-')
    y, m = int(y), int(m)
    # month range as text-comparable dates (doc_date is stored as text 'YYYY-MM-DD')
    start = f"{y:04d}-{m:02d}-01"
    if m == 12:
        end = f"{y + 1:04d}-01-01"
    else:
        end = f"{y:04d}-{m + 1:02d}-01"

    conn = get_db()
    cur = get_cursor(conn)
    cur.execute("""
        SELECT
            vh.vessel_run_type   AS overseas_coastal,
            vh.operation_type    AS import_export,
            vh.cargo_type        AS cargo_type_raw,
            COALESCE(vc_exact.cargo_category, vc_fuzzy.cargo_category) AS category,
            SUM(lpl.quantity)    AS quantity
        FROM vcn_header vh
        JOIN ldud_header ldh         ON ldh.vcn_id = vh.id
        JOIN ldud_parcel_ops lpo     ON lpo.ldud_id = ldh.id
        JOIN lueu_parcel_log lpl     ON lpl.parcel_op_id = lpo.id
        LEFT JOIN vessel_cargo vc_exact
               ON UPPER(vc_exact.cargo_name) = UPPER(vh.cargo_type)
        LEFT JOIN vessel_cargo vc_fuzzy
               ON vc_exact.id IS NULL
              AND UPPER(vh.cargo_type) LIKE '%%' || UPPER(vc_fuzzy.cargo_name) || '%%'
        WHERE vh.doc_date >= %s AND vh.doc_date < %s
        GROUP BY vh.id, vh.vessel_run_type, vh.operation_type, vh.cargo_type,
                 vc_exact.cargo_category, vc_fuzzy.cargo_category
    """, [start, end])
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    lines = []
    for r in rows:
        lines.append({
            'overseas_coastal': r.get('overseas_coastal'),
            'foreign_indian': 'I',   # PLACEHOLDER — see module docstring
            'import_export': r.get('import_export'),
            'category': r.get('category') or r.get('cargo_type_raw'),
            'quantity': r.get('quantity'),
        })
    return lines


def get_report1_data(year_str, month_str, debug=False):
    """
    year_str  : display-only FY label, e.g. '2026-27' (NOT used for filtering)
    month_str : the ACTUAL filter key, 'YYYY-MM', e.g. '2026-06' for Jun-26
    debug     : if True, includes a '_debug' block with raw fetch counts and
                any category values that fell through to OTHER BULK, so an
                unmapped category can be spotted without needing DB access.

    Tries the reconciled mis_vessel_master table first. If that month has
    zero rows (not yet uploaded — typically the current/open month), falls
    back to building the same shape of data live from operational tables.
    """
    if not month_str or len(month_str.split('-')) != 2:
        raise ValueError(f"month must be 'YYYY-MM', got: {month_str!r}")

    month_label = _month_label(month_str)

    lines = _fetch_reconciled_rows(month_label)
    source = 'reconciled'
    fi_placeholder = False

    if not lines:
        lines = _fetch_live_rows(year_str, month_str)
        source = 'live'
        fi_placeholder = True

    rows, totals, debug_block = _aggregate(
        lines, debug=debug, month_label=month_label,
        foreign_indian_placeholder=fi_placeholder,
    )

    result = {
        'rows': rows,
        'totals': totals,
        'year': year_str,
        'month': month_str,
        'source': source,   # 'reconciled' or 'live' — surfaced so UI can show a badge
    }
    if debug_block is not None:
        result['_debug'] = debug_block
    return result


@bp.route('/api/module/RP01/report1/data')
@login_required
def report1_data():
    year = request.args.get('year')
    month = request.args.get('month')   # MUST be 'YYYY-MM', e.g. 2026-06
    debug = request.args.get('debug') == '1'
    if not year or not month:
        return jsonify({'error': 'year and month are required'}), 400
    try:
        report = get_report1_data(year, month, debug=debug)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    return jsonify(report)


@bp.route('/api/module/RP01/report1/export/excel')
@login_required
def report1_export_excel():
    year = request.args.get('year')
    month = request.args.get('month')
    report = get_report1_data(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = 'M-I'

    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    yellow = PatternFill('solid', fgColor='FFFF00')
    grey = PatternFill('solid', fgColor='D9D9D9')
    lightblue = PatternFill('solid', fgColor='DCE6F1')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:K1')
    ws['A1'] = 'OVERSEAS-COASTAL CARGO TRAFFIC HANDLED'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws['A2'] = 'PORT : JAWAHARLAL NEHRU PORT AUTHORITY'
    ws['A2'].font = bold
    ws['H2'] = 'Year :'
    ws['I2'] = report['year']
    ws['H3'] = 'Month :'
    ws['I3'] = report['month']
    ws['I3'].fill = yellow
    if report.get('source') == 'live':
        ws['H4'] = 'Note :'
        ws['I4'] = 'PROVISIONAL (live, not yet reconciled)'
        ws['I4'].font = Font(bold=True, color='CC0000')

    r = 5
    ws.cell(r, 1, 'COMMODITY').font = bold
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=1)
    ws.cell(r, 2, 'OVERSEAS').font = bold
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(r, 6, 'OVERSEAS Total').font = bold
    ws.merge_cells(start_row=r, start_column=6, end_row=r + 2, end_column=6)
    ws.cell(r, 7, 'COASTAL').font = bold
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)
    ws.cell(r, 11, 'COASTAL Total').font = bold
    ws.merge_cells(start_row=r, start_column=11, end_row=r + 2, end_column=11)
    ws.cell(r, 12, 'Grand Total').font = bold
    ws.merge_cells(start_row=r, start_column=12, end_row=r + 2, end_column=12)
    for col in [2, 6, 7, 11, 12]:
        ws.cell(r, col).fill = yellow if col in (2, 7) else grey
        ws.cell(r, col).alignment = center

    r2 = r + 1
    ws.cell(r2, 2, 'IMPORT').font = bold; ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=3)
    ws.cell(r2, 4, 'EXPORT').font = bold; ws.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=5)
    ws.cell(r2, 7, 'IMPORT').font = bold; ws.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=8)
    ws.cell(r2, 9, 'EXPORT').font = bold; ws.merge_cells(start_row=r2, start_column=9, end_row=r2, end_column=10)
    for col in [2, 4, 7, 9]:
        ws.cell(r2, col).fill = yellow
        ws.cell(r2, col).alignment = center

    r3 = r2 + 1
    labels3 = {2: 'IF', 3: 'FF', 4: 'IF', 5: 'FF', 7: 'IF', 8: 'FF', 9: 'IF', 10: 'FF'}
    for col, lbl in labels3.items():
        ws.cell(r3, col, lbl).font = bold
        ws.cell(r3, col).fill = yellow
        ws.cell(r3, col).alignment = center

    for cc in range(1, 13):
        for rr in (r, r2, r3):
            ws.cell(rr, cc).border = border

    r = r3 + 1
    for row in report['rows']:
        vals = [row['commodity'], row['ov_imp_if'], row['ov_imp_ff'], row['ov_exp_if'], row['ov_exp_ff'],
                row['ov_total'], row['co_imp_if'], row['co_imp_ff'], row['co_exp_if'], row['co_exp_ff'],
                row['co_total'], row['grand_total']]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(r, ci, v if v else (row['commodity'] if ci == 1 else 0.00))
            cell.border = border
            cell.alignment = center
            if ci == 6 or ci == 11:
                cell.fill = grey
        r += 1

    t = report['totals']
    vals = ['Grand Total', t['ov_imp_if'], t['ov_imp_ff'], t['ov_exp_if'], t['ov_exp_ff'], t['ov_total'],
            t['co_imp_if'], t['co_imp_ff'], t['co_exp_if'], t['co_exp_ff'], t['co_total'], t['grand_total']]
    for ci, v in enumerate(vals, start=1):
        cell = ws.cell(r, ci, v)
        cell.font = bold
        cell.border = border
        cell.alignment = center
        cell.fill = lightblue

    ws.column_dimensions['A'].width = 18
    for i in range(2, 13):
        ws.column_dimensions[get_column_letter(i)].width = 13

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"Report1_{report['year']}_{report['month']}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )