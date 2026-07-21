"""
Report-9 — Traffic / Vessels Handled at BPCL-BT/Anchorage
Flask Blueprint version. Reads directly from mis_vessel_master (Postgres).

Mirrors the structure of report1.py so it can be dropped into the same
module pattern (e.g. modules/RP09/report9/routes.py) and registered the
same way report1 is.

DATA-SOURCE / FALLBACK NOTES (mirrors report3/report8's pattern):
- Primary source is mis_vessel_master, exactly as before.
- For any (fin_year, fy_month_idx) period that has ZERO rows in
  mis_vessel_master, figures for that period only are pulled instead from
  the live LUEU01 pipeline (vcn_header / ldud_header / ldud_parcel_ops /
  lueu_parcel_log). mis_vessel_master always wins for periods where it has
  data; the live pipeline is purely a gap-filler for periods it hasn't
  reached yet.
- In the live pipeline, berth_no and import_export are NOT on
  mis_vessel_master's schema — they live on vcn_header as
  vcn_header.berth_name and vcn_header.operation_type respectively.
  Everything else (quantity, entry_date for period bucketing) comes from
  lueu_parcel_log, joined down through ldud_parcel_ops / ldud_header to
  vcn_header, the same join path used in report3's
  _load_live_pipeline_data().
"""

import io
import datetime
from datetime import date
import traceback
from functools import wraps

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from .. import bp

from database import get_db, get_cursor

# NOTE: single dot — bp lives in THIS package's __init__.py (modules/RP09/__init__.py),
# not two levels up. This is the line that gets corrupted if it's ever line-wrapped
# during copy/paste — keep it on one line.



def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
CAL_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ---------------------------------------------------------------------------
# Berth layout (order matches the sample sheet: individual berths, then the
# subtotal / grouping rows that roll them up).
#
# NOTE: this assumes mis_vessel_master.berth_no (and, for the live-pipeline
# fallback, vcn_header.berth_name) holds values exactly like "LB-01",
# "LB-01[N]", "LB-01[S]", "LB-02", "LB-03", "LB-04", "INN ANCHORAGE". If the
# actual stored text differs (spacing, casing, abbreviation), tell me the
# real values and I'll adjust BERTH_ROWS below.
# ---------------------------------------------------------------------------
BERTH_ROWS = [
    {"key": "LB-01",            "type": "berth"},
    {"key": "LB-01[N]",         "type": "berth"},
    {"key": "LB-01[S]",         "type": "berth"},
    {"key": "LB-02",            "type": "berth"},
    {"key": "LB-01 & LB-02",    "type": "subtotal", "berths": ["LB-01", "LB-01[N]", "LB-01[S]", "LB-02"]},
    {"key": "LB-03",            "type": "berth"},
    {"key": "LB-04",            "type": "berth"},
    {"key": "LB-03 & LB-04",    "type": "subtotal", "berths": ["LB-03", "LB-04"]},
    {"key": "LIQUID TERMINAL",  "type": "subtotal", "berths": ["LB-01", "LB-01[N]", "LB-01[S]", "LB-02", "LB-03", "LB-04"]},
    {"key": "INN ANCHORAGE",    "type": "berth"},
    {"key": "TOTAL",            "type": "total", "berths": ["LB-01", "LB-01[N]", "LB-01[S]", "LB-02", "LB-03", "LB-04", "INN ANCHORAGE"]},
]

ALL_BERTH_NAMES = [r["key"] for r in BERTH_ROWS if r["type"] == "berth"]


class ReportDataError(Exception):
    """Raised for any problem loading/validating the report's source data.
    Caught by the route handlers and turned into a clean JSON error response."""
    pass


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def _dt_to_fy_month(dt):
    """A real datetime/date -> (fin_year, fy_month_idx), Apr-Mar FY convention.
    e.g. 2026-07-12 -> ('2026-27', 3)."""
    d = dt.date() if isinstance(dt, datetime.datetime) else dt
    fy_start = d.year if d.month >= 4 else d.year - 1
    fin_year = f"{fy_start}-{str(fy_start + 1)[-2:]}"
    mn = CAL_MONTH_ABBR[d.month - 1]
    return fin_year, MONTH_NAMES.index(mn)


def _parse_dt(v):
    """Parse the live pipeline's entry_date / free-text datetime values.
    Returns None if blank/unparseable."""
    if not v:
        return None
    if isinstance(v, (datetime.datetime, date)):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.datetime.strptime(s.replace("T", " ")[:16], "%Y-%m-%d %H:%M")
    except ValueError:
        try:
            return datetime.datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None


def _norm_import_export(v: str) -> str:
    return (v or "").strip().lower()


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts


def month_str_to_idx(month_str: str) -> int:
    """'Jun-26' -> 2, 'Dec-24' -> 8, etc. Matches MONTH_NAMES order (FY Apr..Mar)."""
    abbrev = str(month_str).split("-")[0].strip()
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized value in mis_vessel_master.month: '{month_str}' "
            f"(expected something like 'Jun-26')"
        )


# ── ADD: fallback source -- live LUEU01 logging pipeline. Returns the same
# shape as the mis_vessel_master loader (fin_year, fy_month_idx, berth_no,
# vcn_no, import_export, quantity) so it can be concatenated directly. Only
# used for (fin_year, fy_month_idx) periods that have ZERO rows in
# mis_vessel_master -- see load_data() below. berth_no/import_export come
# from vcn_header.berth_name / vcn_header.operation_type. ──
def _load_live_pipeline_data() -> pd.DataFrame:
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT l.entry_date, h.id AS vcn_id, h.berth_name, h.operation_type, l.quantity
            FROM lueu_parcel_log l
            JOIN ldud_parcel_ops po ON po.id = l.parcel_op_id
            JOIN ldud_header ld ON ld.id = po.ldud_id
            JOIN vcn_header h ON h.id = ld.vcn_id
            WHERE l.is_deleted IS NOT TRUE
              AND l.entry_date IS NOT NULL
              AND l.quantity IS NOT NULL
        """)
        log_rows = cur.fetchall()
    finally:
        conn.close()

    empty = pd.DataFrame(columns=["fin_year", "fy_month_idx", "berth_no", "vcn_no", "import_export", "quantity"])
    if not log_rows:
        return empty

    ldf = pd.DataFrame(log_rows)
    ldf["quantity"] = pd.to_numeric(ldf["quantity"], errors="coerce").fillna(0.0)
    ldf["berth_no"] = ldf["berth_name"].astype(str).str.strip()
    ldf["import_export"] = ldf["operation_type"].apply(_norm_import_export)
    ldf["vcn_no"] = ldf["vcn_id"].astype(str)

    ldf["entry_dt"] = ldf["entry_date"].apply(_parse_dt)
    ldf = ldf.dropna(subset=["entry_dt"])
    if ldf.empty:
        return empty

    fy_list, idx_list = [], []
    for dt in ldf["entry_dt"]:
        fy, idx = _dt_to_fy_month(dt)
        fy_list.append(fy)
        idx_list.append(idx)
    ldf["fin_year"] = fy_list
    ldf["fy_month_idx"] = idx_list

    return ldf[["fin_year", "fy_month_idx", "berth_no", "vcn_no", "import_export", "quantity"]].copy()


def load_data() -> pd.DataFrame:
    """Returns df with columns [fin_year, fy_month_idx, berth_no, vcn_no,
    import_export, quantity].
    Primary source: mis_vessel_master, exactly as before.
    Fallback: for any (fin_year, fy_month_idx) period with ZERO rows in
    mis_vessel_master, figures are pulled instead from the live LUEU01
    pipeline for that period only. mis_vessel_master always wins for
    periods where it has data.
    """
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT fin_year, month, berth_no, vcn_no, import_export, quantity
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month IS NOT NULL
        """)
        rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        mv_df = pd.DataFrame(columns=["fin_year", "fy_month_idx", "berth_no", "vcn_no", "import_export", "quantity"])
    else:
        df = pd.DataFrame(rows)

        missing_cols = [c for c in ("fin_year", "month", "berth_no", "vcn_no", "import_export", "quantity")
                         if c not in df.columns]
        if missing_cols:
            raise ReportDataError(f"Query result is missing column(s): {', '.join(missing_cols)}")

        df["fin_year"] = df["fin_year"].astype(str).str.strip()
        df["berth_no"] = df["berth_no"].astype(str).str.strip()
        df["vcn_no"] = df["vcn_no"].astype(str).str.strip()
        df["import_export"] = df["import_export"].astype(str).str.strip().str.lower()
        df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0.0)
        df["fy_month_idx"] = df["month"].apply(month_str_to_idx)

        mv_df = df[["fin_year", "fy_month_idx", "berth_no", "vcn_no", "import_export", "quantity"]].copy()

    # ---- which (fin_year, fy_month_idx) periods does mis_vessel_master
    # actually cover? Only periods with ZERO rows there fall back. ----
    covered_periods = set(zip(mv_df["fin_year"], mv_df["fy_month_idx"]))

    live_df = _load_live_pipeline_data()
    if not live_df.empty:
        live_df = live_df[
            ~live_df.apply(lambda r: (r["fin_year"], r["fy_month_idx"]) in covered_periods, axis=1)
        ]

    df_all = pd.concat([mv_df, live_df], ignore_index=True)

    if df_all.empty:
        raise ReportDataError(
            "No usable rows found in mis_vessel_master or the live LUEU01 pipeline."
        )

    unmapped = sorted(set(df_all["berth_no"].unique().tolist()) - set(ALL_BERTH_NAMES))
    if unmapped:
        # Not fatal — just means some berths in the DB aren't part of this
        # report's layout (e.g. other terminals). We simply ignore them.
        print("REPORT9: berth_no values not in BERTH_ROWS layout (ignored):", unmapped)

    return df_all


def _get_df_and_years():
    df = load_data()
    years = sorted(df["fin_year"].unique().tolist())
    return df, years


def period_stats(subset: pd.DataFrame, berths: list) -> dict:
    b = subset[subset["berth_no"].isin(berths)]
    vsls = int(b["vcn_no"].nunique())
    imp = round(float(b.loc[b["import_export"] == "import", "quantity"].sum()), 3)
    exp = round(float(b.loc[b["import_export"] == "export", "quantity"].sum()), 3)
    tot = round(float(b["quantity"].sum()), 3)
    return {"vsls": vsls, "import": imp, "export": exp, "total": tot}


def compute_totals(df: pd.DataFrame, fin_year: str, month_idx: int) -> list:
    fy_subset = df[df["fin_year"] == fin_year]
    month_subset = fy_subset[fy_subset["fy_month_idx"] == month_idx]
    upto_subset = fy_subset[fy_subset["fy_month_idx"] <= month_idx]

    rows = []
    for r in BERTH_ROWS:
        berths = r["berths"] if r["type"] != "berth" else [r["key"]]
        rows.append({
            "label": r["key"],
            "type": r["type"],
            "month": period_stats(month_subset, berths),
            "upto": period_stats(upto_subset, berths),
        })
    return rows


@bp.route("/module/RP01/report9/")
@login_required
def report9_index():
    return render_template("/report9/report09.html", port_name="")


@bp.route("/api/module/RP01/report9/meta")
@login_required
def report9_api_meta():
    try:
        _, years = _get_df_and_years()
        months = {fy: month_options_for(fy) for fy in years}
        return jsonify({"years": years, "months": months, "port_name": ""})
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report9/report")
@login_required
def report9_api_report():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        rows = compute_totals(df, fin_year, month_idx)
        month_label = next(o["label"] for o in month_options_for(fin_year) if o["idx"] == month_idx)

        return jsonify({
            "fin_year": fin_year,
            "month_label": month_label,
            "rows": rows,
        })
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report9/export")
@login_required
def report9_api_export():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1])
        month_idx = int(request.args.get("month_idx", 2))
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        rows = compute_totals(df, fin_year, month_idx)
        month_label = next(o["label"] for o in month_options_for(fin_year) if o["idx"] == month_idx)

        wb = Workbook()
        ws = wb.active
        ws.title = "Report-9"

        # ---- styles -------------------------------------------------
        bold = Font(bold=True)
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        title_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        subtotal_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        n_data_rows = len(BERTH_ROWS)
        last_col = 9  # A=berths label col(2 cols incl border) .. we'll use B..I (8 data cols) -> B..I = 8 cols

        # Columns: A=Berths, B=Vsls(month), C=Import(month), D=Export(month), E=Total(month),
        #          F=Vsls(upto), G=Import(upto), H=Export(upto), I=Total(upto)

        # ---- date stamp (row 1) --------------------------------------
        ws["I1"] = datetime.date.today().strftime("%d-%b-%y")
        ws["I1"].alignment = right

        # ---- title (row 3) --------------------------------------------
        ws.merge_cells("A3:I3")
        ws["A3"] = "TRAFFIC / VESSELS HANDLED at BPCL-BT/Anchorage"
        ws["A3"].font = title_font
        ws["A3"].alignment = center
        ws["A3"].fill = title_fill

        # ---- header rows (5,6,7) ---------------------------------------
        r1, r2, r3 = 5, 6, 7

        ws.merge_cells(f"A{r1}:A{r3}")
        ws[f"A{r1}"] = "BERTHS"
        ws[f"A{r1}"].font = header_font
        ws[f"A{r1}"].alignment = center

        ws.merge_cells(f"B{r1}:E{r1}")
        ws[f"B{r1}"] = month_label
        ws[f"B{r1}"].font = header_font
        ws[f"B{r1}"].alignment = center
        ws[f"B{r1}"].fill = yellow_fill

        ws.merge_cells(f"F{r1}:I{r1}")
        ws[f"F{r1}"] = fin_year
        ws[f"F{r1}"].font = header_font
        ws[f"F{r1}"].alignment = center

        ws.merge_cells(f"B{r2}:B{r3}")
        ws[f"B{r2}"] = "Vsls"
        ws[f"B{r2}"].font = header_font
        ws[f"B{r2}"].alignment = center

        ws.merge_cells(f"C{r2}:E{r2}")
        ws[f"C{r2}"] = "Quantity"
        ws[f"C{r2}"].font = header_font
        ws[f"C{r2}"].alignment = center

        ws.merge_cells(f"F{r2}:F{r3}")
        ws[f"F{r2}"] = "Vsls"
        ws[f"F{r2}"].font = header_font
        ws[f"F{r2}"].alignment = center

        ws.merge_cells(f"G{r2}:I{r2}")
        ws[f"G{r2}"] = "Quantity"
        ws[f"G{r2}"].font = header_font
        ws[f"G{r2}"].alignment = center

        for col, label in (("C", "Import"), ("D", "Export"), ("E", "Total"),
                           ("G", "Import"), ("H", "Export"), ("I", "Total")):
            cell = ws[f"{col}{r3}"]
            cell.value = label
            cell.font = header_font
            cell.alignment = center

        for row in (r1, r2, r3):
            for col_idx in range(1, last_col + 1):
                ws.cell(row=row, column=col_idx).border = thin_border

        # ---- data rows --------------------------------------------------
        row_i = r3 + 1
        for r in rows:
            label_cell = ws[f"A{row_i}"]
            label_cell.value = ("    " + r["label"]) if r["type"] == "berth" else r["label"]
            label_cell.alignment = left if r["type"] == "berth" else center
            label_cell.font = bold if r["type"] != "berth" else Font()

            m, u = r["month"], r["upto"]
            values = [m["vsls"], m["import"], m["export"], m["total"],
                      u["vsls"], u["import"], u["export"], u["total"]]
            cols = ["B", "C", "D", "E", "F", "G", "H", "I"]

            for col, val in zip(cols, values):
                cell = ws[f"{col}{row_i}"]
                cell.value = val
                cell.alignment = right
                if col in ("B", "F"):
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.000"
                if r["type"] != "berth":
                    cell.font = bold

            row_fill = None
            if r["type"] == "subtotal":
                row_fill = subtotal_fill
            elif r["type"] == "total":
                row_fill = total_fill

            for col_idx in range(1, last_col + 1):
                cell = ws.cell(row=row_i, column=col_idx)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

            row_i += 1

        # ---- column widths -------------------------------------------------
        widths = {"A": 20, "B": 8, "C": 12, "D": 12, "E": 12, "F": 8, "G": 12, "H": 12, "I": 12}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-9_BPCL-BT_{fin_year}_{month_label}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except ValueError as e:
        return jsonify({"error": f"Invalid parameter: {e}"}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500