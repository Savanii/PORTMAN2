"""
Report-11 — Bulk Terminal Performance Report
Flask Blueprint version.

Source: mis_vessel_master (single table — all timing/quantity columns already
present as numeric, per-vessel-call, values in DAYS for the timing columns).

Section A  -> raw sums pulled straight from DB for the selected fin_year
              (all 12 months, Apr-Mar, + FY total column), written as literal
              numbers (these are the "inputs").
Section B  -> derived productivity parameters. Written as REAL EXCEL FORMULAS
              referencing Section A cells in the same month-column (verified
              against the uploaded workbook's actual Apr/May/Jun-26 numbers).
Section C  -> same derived parameters expressed in Hours instead of Days.
              Also real Excel formulas referencing A/B cells.
Reasons block -> static legend, not data-driven, reproduced verbatim.

Known gaps (no data source in current schema -> always 0 / blank):
    - Vessel Discharge/Load (TEUs), Tonnage (row A3), Crane deployed hours,
      Idle time at NON-working berth (Port/Non-port a/c), Shifting time,
      Total moves, Total TEUs for crane productivity, Rail load/discharge,
      No. of rakes handled, and anything TEU-based (Container throughput,
      Gross Crane Productivity, TRT per 1000 TEUs).
    - No. of berths (row A22) has no DB column -> constant NO_OF_BERTHS
      below (default 2). Change it if the terminal's berth count differs.

Cargo -> broad section (LIQUID / DRY BULK / BREAK BULK) classification is
best-effort keyword matching on mis_vessel_master.cargo (free text). Only
Liquid-terminal cargo names have been verified against real data so far
(BASE OIL, ACETIC ACID, FO/FO[E], FURNACE OIL, etc.). Unmapped cargo values
are dropped with a console warning rather than guessed.
"""

import calendar
import io
import traceback
from functools import wraps

import pandas as pd

from flask import jsonify, request, render_template, send_file, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re
from datetime import datetime, timedelta

from database import get_db, get_cursor

from .. import bp


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


MONTH_NAMES = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

# No DB column carries this -> constant assumption. Adjust if wrong.
NO_OF_BERTHS = 2


class ReportDataError(Exception):
    pass


def fy_start_year(fin_year: str) -> int:
    return int(fin_year.split("-")[0])


def month_options_for(fin_year: str):
    start_y = fy_start_year(fin_year)
    opts = []
    for idx, mn in enumerate(MONTH_NAMES):
        yy = start_y if idx < 9 else start_y + 1
        opts.append({"idx": idx, "label": f"{mn}-{str(yy % 100).zfill(2)}"})
    return opts
def _days_between(t1, t2):
    """Days (float) from t1 -> t2. 0 if either is missing or negative."""
    if t1 is None or t2 is None:
        return 0.0
    delta = (t2 - t1).total_seconds() / 86400.0
    return delta if delta > 0 else 0.0
def _classify_delay_reason(delay_name):
    """Matches a lueu_parcel_log.delay_name string against the same
    REASONS_PORT / REASONS_NON_PORT legends used in the Excel export.
    Returns 'port', 'non_port', or None if it can't tell / blank."""
    name = str(delay_name or "").strip().upper()
    if not name:
        return None

    port_keywords = (
        "BERTH", "TUG", "PILOT", "EQUIPMENT", "BREAKDOWN", "WORKER",
        "STRIKE", "STOPPAGE", "POWER FAILURE", "LABOUR HOLIDAY",
        "NIGHT NAVIGATION", "DRAFT RESTRICTION",
    )
    non_port_keywords = (
        "SHIP", "SHIPPER", "AGENT", "CARGO", "DEPARTURE", "WEATHER",
        "STORAGE", "TIDAL", "DOCUMENT", "POWER FAILURE GRID", "SCHEDULE",
    )

    if any(k in name for k in port_keywords):
        return "port"
    if any(k in name for k in non_port_keywords):
        return "non_port"
    return None


def _fetch_live_idle_by_parcel_op(cur, parcel_op_ids):
    """Sums idle/delay hours (as day-fractions) from lueu_parcel_log per
    parcel_op_id, split into port / non-port buckets by delay_name."""
    idle = {}  # parcel_op_id -> {"port": days, "non_port": days}
    if not parcel_op_ids:
        return idle

    cur.execute("""
        SELECT parcel_op_id, entry_date, from_time, to_time, delay_name
        FROM lueu_parcel_log
        WHERE COALESCE(is_deleted, FALSE) = FALSE
          AND parcel_op_id = ANY(%s)
          AND delay_name IS NOT NULL
          AND delay_name <> ''
    """, (list(parcel_op_ids),))

    for r in cur.fetchall():
        bucket = _classify_delay_reason(r["delay_name"])
        if bucket is None:
            continue
        try:
            entry_date = str(r["entry_date"]).strip()
            ft = datetime.strptime(f"{entry_date} {r['from_time']}", "%Y-%m-%d %H:%M")
            tt = datetime.strptime(f"{entry_date} {r['to_time']}", "%Y-%m-%d %H:%M")
            if tt <= ft:
                tt += timedelta(days=1)  # spans midnight
            days = (tt - ft).total_seconds() / 86400.0
        except (ValueError, TypeError):
            continue

        pid = r["parcel_op_id"]
        idle.setdefault(pid, {"port": 0.0, "non_port": 0.0})
        idle[pid][bucket] += days

    return idle


def _fetch_live_rows(cur):
    """Current-month vessel-call rows built from LDUD/VCN, in the same
    shape as the mis_vessel_master rows, for months not yet migrated
    into mis_vessel_master."""
    cur.execute("""
        SELECT
            po.id AS parcel_op_id,
            to_char(current_date, 'YYYY') || '-' ||
                right(to_char(current_date + interval '1 year', 'YYYY'), 2) AS fin_year,
            to_char(current_date, 'Mon-YY') AS month,
            vh.berth_name AS berth_no,
            vh.operation_type AS import_export,
            po.cargo_name AS cargo,
            po.quantity AS quantity,
            lh.nor_tendered AS nor_tendered,
            lh.nor_accepted AS nor_accepted,
            lh.alongside_datetime AS alongside_datetime,
            lh.cast_off_datetime AS cast_off_datetime,
            lh.pilot_pickup_time AS pilot_pickup_time,
            lh.pilot_board_departure AS pilot_board_departure
        FROM ldud_parcel_ops po
        JOIN ldud_header lh ON lh.id = po.ldud_id
        JOIN vcn_header vh ON vh.id = lh.vcn_id
        WHERE to_char(current_date, 'Mon-YY') = to_char(current_date, 'Mon-YY')
    """)
    raw = cur.fetchall()
    if not raw:
        return []

    parcel_op_ids = [r["parcel_op_id"] for r in raw]
    idle = _fetch_live_idle_by_parcel_op(cur, parcel_op_ids)

    rows = []
    for r in raw:
        nor_tendered = _parse_ts(r["nor_tendered"])
        nor_accepted = _parse_ts(r["nor_accepted"])
        alongside = _parse_ts(r["alongside_datetime"])
        cast_off = _parse_ts(r["cast_off_datetime"])
        pilot_pickup = _parse_ts(r["pilot_pickup_time"])
        pilot_departure = _parse_ts(r["pilot_board_departure"])

        waiting_non_port = _days_between(nor_tendered, nor_accepted)
        waiting_port = _days_between(nor_accepted, alongside)
        stay_at_berth = _days_between(alongside, cast_off)
        inward_movement = _days_between(pilot_pickup, alongside)
        outward_movement = _days_between(pilot_departure, cast_off)

        idle_bucket = idle.get(r["parcel_op_id"], {"port": 0.0, "non_port": 0.0})

        rows.append({
            "fin_year": r["fin_year"],
            "month": r["month"],
            "berth_no": r["berth_no"],
            "import_export": r["import_export"],
            "cargo": r["cargo"],
            "quantity": r["quantity"],
            "pre_berthing_waiting": waiting_port + waiting_non_port,
            "waiting_port": waiting_port,
            "waiting_non_port": waiting_non_port,
            "stay_at_berth": stay_at_berth,
            "inward_movement": inward_movement,
            "outward_movement": outward_movement,
            "non_working_port": idle_bucket["port"],
            "non_working_non_port": idle_bucket["non_port"],
        })
    return rows
def _parse_ts(val):
    """LDUD header fields are free-text ISO-ish datetimes ('2026-07-12T14:40')
    or blank. Returns a datetime or None."""
    if not val or not str(val).strip():
        return None
    val = str(val).strip()
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None
def month_str_to_idx(month_str: str) -> int:
    abbrev = str(month_str).split("-")[0].strip()
    try:
        return MONTH_NAMES.index(abbrev)
    except ValueError:
        raise ReportDataError(
            f"Unrecognized value in mis_vessel_master.month: '{month_str}' "
            f"(expected something like 'Jun-26')"
        )


def days_in_month(fin_year: str, month_idx: int) -> int:
    start_y = fy_start_year(fin_year)
    real_month_num = (month_idx + 3) % 12 + 1          # Apr=4 ... Mar=3
    year = start_y if month_idx < 9 else start_y + 1
    return calendar.monthrange(year, real_month_num)[1]


# ---------------------------------------------------------------------
# Cargo (free text) -> broad section classification
# ---------------------------------------------------------------------
def classify_broad_category(cargo):
    cargo = str(cargo or "").strip().upper()

    liquid_names = (
        "FO", "FO [E]", "CBFS", "FURNACE OIL", "POL CRUDE",
        "CPO", "CPKO", "CPO/CPKO", "CPO/RBDPO", "RBD PALM OLEIN", "EDIBLE OIL",
        "SUNFLOWER OIL", "CDSBO", "CSBO", "CSFO",
        "CHEMICAL", "CHEMICALS", "ACETIC ACID", "A. ACID", "A.ACID", "AACID",
        "VAM", "PHENOL/ACETON/VAM", "PHENOL", "MDC", "MEK", "IPA",
        "ISOPROPYL ALCOHOL", "SM", "STRENE MONOMER", "STYRENE MONOMER",
        "N BUTONAL/TOLUNE", "PHOSPHORIC ACID", "PH.ACID", "PH ACID",
        "BASE OIL",
    )
    if cargo in liquid_names or "LPG" in cargo or "LNG" in cargo or \
       "BASE OIL" in cargo or "LUBE" in cargo or "SHELL" in cargo:
        return "LIQUID"

    dry_bulk_keywords = (
        "IRON ORE", "COAL", "FERTILIZER", "CEMENT", "SALT", "SUGAR",
        "PULSES", "FOOD GRAIN", "TEA", "COFFEE", "SCRAP",
        "CLINKER", "LIMESTONE", "DOLOMITE", "HBI", "FINES",
        "GYPSUM", "BAUXITE", "CLO", "BRBF", "MABU", "VIZAG", "DHAMRA",
    )
    if any(k in cargo for k in dry_bulk_keywords):
        return "DRY BULK"

    break_bulk_keywords = ("IRON AND STEEL", "TIMBER", "LOG", "PROJECT CARGO")
    if any(k in cargo for k in break_bulk_keywords):
        return "BREAK BULK"

    return None


def _direction(import_export):
    ie = str(import_export or "").strip().upper()
    if ie == "IMPORT":
        return "Import"
    if ie == "EXPORT":
        return "Export"
    return None


def load_data() -> pd.DataFrame:
    conn = get_db()
    try:
        cur = get_cursor(conn)
        cur.execute("""
            SELECT
                fin_year,
                month,
                berth_no,
                import_export,
                cargo,
                quantity,
                pre_berthing_waiting,
                waiting_port,
                waiting_non_port,
                stay_at_berth,
                inward_movement,
                outward_movement,
                non_working_port,
                non_working_non_port
            FROM mis_vessel_master
            WHERE fin_year IS NOT NULL
              AND month IS NOT NULL
        """)
        mis_rows = cur.fetchall()

        # ------------------------------------------------------------
        # Only pull live LDUD/LUEU data if the current month isn't
        # already present in mis_vessel_master. Never discard mis_rows.
        # ------------------------------------------------------------
        current_month = pd.Timestamp.today().strftime("%b-%y")
        mis_current = [r for r in mis_rows if str(r["month"]).strip() == current_month]

        live_rows = []
        if not mis_current:
            print("REPORT11: Current month not found in mis_vessel_master")
            print("REPORT11: Loading live LDUD/LUEU data...")
            live_rows = _fetch_live_rows(cur)
        else:
            print("REPORT11: Current month already present in mis_vessel_master — skipping live load")

        rows = list(mis_rows) + list(live_rows)
    finally:
        conn.close()

    cols = [
        "fin_year", "fy_month_idx", "import_export", "quantity",
        "waiting_port", "waiting_non_port", "stay_at_berth",
        "inward_movement", "outward_movement",
        "non_working_port", "non_working_non_port", "broad_category",
    ]
    if not rows:
        return pd.DataFrame(columns=cols)

    df = pd.DataFrame(rows)
    print(f"REPORT11 DEBUG: rows fetched from DB: {len(df)}")

    df["fin_year"] = df["fin_year"].str.strip()
    df["fy_month_idx"] = df["month"].apply(month_str_to_idx)

    numeric_cols = [
        "quantity", "pre_berthing_waiting", "waiting_port", "waiting_non_port",
        "stay_at_berth", "inward_movement", "outward_movement",
        "non_working_port", "non_working_non_port",
    ]
    for c in numeric_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    df["direction"] = df["import_export"].apply(_direction)
    df["broad_category"] = df["cargo"].apply(classify_broad_category)

    unmapped = sorted(df.loc[df["broad_category"].isna(), "cargo"].dropna().unique().tolist())
    if unmapped:
        print("REPORT11 WARNING: Unmapped cargo values (dropped from Dry/Break/Liquid tonnage rows):")
        for c in unmapped:
            print("   ", c)

    return df


def _get_df_and_years():
    df = load_data()
    years = sorted(df["fin_year"].unique().tolist())
    return df, years


# ---------------------------------------------------------------------
# Section A: raw sums per month, straight from mis_vessel_master
# ---------------------------------------------------------------------
def compute_section_a_month(df, fin_year, month_idx):
    m = df[(df["fin_year"] == fin_year) & (df["fy_month_idx"] == month_idx)]

    vessels_sailed = len(m)  # one row per vessel call

    def tonnes(category, direction):
        sub = m[(m["broad_category"] == category) & (m["direction"] == direction)]
        return round(float(sub["quantity"].sum()), 3)

    a = {
        "Vessel Discharge (Including Restow)": 0.0,
        "Vessel Load (Including Restow)": 0.0,
        "Tonnage": 0.0,
        "Vessels Sailed": vessels_sailed,
        "Pre_berthing Waiting Time-on Port a/c (Total)": round(float(m["waiting_port"].sum()) * 24, 3),
        "Pre_berthing Waiting Time-on Non-Port a/c (Total)": round(float(m["waiting_non_port"].sum()) * 24, 3),
        "Total Berth Stay of all vessels (For Berth Productivity)": round(float(m["stay_at_berth"].sum()) * 24, 3),
        "Total Crane deplyoed hours (For Crane Productivity)": 0.0,
        "Vessel Inward movement (Total)": round(float(m["inward_movement"].sum()) * 24, 3),
        "Vessel Outward movement (Total)": round(float(m["outward_movement"].sum()) * 24, 3),
        "Idle time at working berth on Port A/c.": round(float(m["non_working_port"].sum()) * 24, 3),
        "Idle time at working berth on Non-Port A/c.": round(float(m["non_working_non_port"].sum()) * 24, 3),
        "Idle time at Non-working berth on Port A/c.": 0.0,
        "Idle time at Non-working berth on Non-Port A/c.": 0.0,
        "Shifting Time": 0.0,
        "Total No. of Moves for calculating  Berth / Crane Productivity": 0.0,
        "Total No. of TEUs for calculating  Crane Productivity": 0.0,
        "Rail Load": 0.0,
        "Rail Discharge": 0.0,
        "No. of Rakes handled": 0.0,
        "Days in a month": days_in_month(fin_year, month_idx),
        "No. of berths for % berth occupancy": NO_OF_BERTHS,
        "Dry Bulk traffic - Import": tonnes("DRY BULK", "Import"),
        "Dry Bulk traffic - Export": tonnes("DRY BULK", "Export"),
        "Break Bulk traffic - Import": tonnes("BREAK BULK", "Import"),
        "Break Bulk traffic - Export": tonnes("BREAK BULK", "Export"),
        "Liquid - Import": tonnes("LIQUID", "Import"),
        "Liquid - Export": tonnes("LIQUID", "Export"),
    }
    print(
    f"Month={month_idx}",
    "Liquid Import =", a["Liquid - Import"],
    "Liquid Export =", a["Liquid - Export"],
)
    return a



def _safe_div(n, d):
    return (n / d) if d else 0.0


def compute_section_b_month(a):
    total_traffic_tons = (
        a["Dry Bulk traffic - Import"] + a["Dry Bulk traffic - Export"]
        + a["Break Bulk traffic - Import"] + a["Break Bulk traffic - Export"]
        + a["Liquid - Import"] + a["Liquid - Export"]
    )
    vs = a["Vessels Sailed"]
    berth_stay_hrs = a["Total Berth Stay of all vessels (For Berth Productivity)"]

    b = {
        "Vessels Sailed": vs,
        "Total Traffic Throughputs (TEUs)": a["Vessel Discharge (Including Restow)"] + a["Vessel Load (Including Restow)"],
        "Total traffic throughputs (Tons)": round(total_traffic_tons, 3),
        "Parcel Size": 0.0,  # TEU-based, no source
        "Avg. Pre-berthing Waiting Time-Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"],
            24 * vs), 6),
        "Avg. Pre-berthing Waiting Time-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Port a/c (Total)"], 24 * vs), 6),
        "Avg. Pre-berthing Waiting Time-Non-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], 24 * vs), 6),
        "Avg. Berth stay": round(_safe_div(berth_stay_hrs, 24 * vs), 6),
        "Avg. Turn around time - Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"]
            + berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], 24 * vs), 6),
        "Avg. Turn around time - Port A/c.": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + berth_stay_hrs
            + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], 24 * vs), 6),
        "Avg. Turn around time - Non- Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], 24 * vs), 6),
        "Avg. Turn around time - Pilot Boarding to De-boarding-Total": round(_safe_div(
            berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], 24 * vs), 6),
        "Berth Occupancy": round(_safe_div(berth_stay_hrs, a["Days in a month"] * 24 * a["No. of berths for % berth occupancy"]), 6),
        "Idle time": round(_safe_div(
            a["Idle time at working berth on Port A/c."] + a["Idle time at working berth on Non-Port A/c."],
            berth_stay_hrs), 6),
        "Gross Berth Productivity": 0.0,
        "Gross Crane Productivity (Moves)": None,
        "Gross Crane Productivity (TEUs)": None,
        "Ship Output per Day (TEUs)": 0.0,
        "No. of Rakes handled": a["No. of Rakes handled"],
        "Total Rail traffic": a["Rail Load"] + a["Rail Discharge"],
        "% wrt to Total Thoughput": None,
    }
    b["Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c."] = b["Avg. Turn around time - Pilot Boarding to De-boarding-Total"]
    b["Ship Output per Day (Tonnes)"] = round(_safe_div(b["Total traffic throughputs (Tons)"], b["Avg. Berth stay"]), 3) if b["Avg. Berth stay"] else 0.0
    return b


def compute_section_c_month(a, b):
    vs = a["Vessels Sailed"]
    berth_stay_hrs = a["Total Berth Stay of all vessels (For Berth Productivity)"]

    c = {
        "Avg. Pre-berthing Waiting Time-Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], vs), 6),
        "Avg. Pre-berthing Waiting Time-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Port a/c (Total)"], vs), 6),
        "Avg. Pre-berthing Waiting Time-Non-Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], vs), 6),
        "Avg. Berth stay": round(_safe_div(berth_stay_hrs, vs), 6),
        "Avg. Turn around time - Total": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"]
            + berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], vs), 6),
        "Avg. Turn around time - Port A/c.": round(_safe_div(
            a["Pre_berthing Waiting Time-on Port a/c (Total)"] + berth_stay_hrs
            + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], vs), 6),
        "Avg. Turn around time - Non- Port A/c.": round(_safe_div(a["Pre_berthing Waiting Time-on Non-Port a/c (Total)"], vs), 6),
        "Avg. Turn around time - Pilot Boarding to De-boarding-Total": round(_safe_div(
            berth_stay_hrs + a["Vessel Inward movement (Total)"] + a["Vessel Outward movement (Total)"], vs), 6),
        "Avg. TRT for 1000 tonnes (Days)": 0.0,
        "Avg. TRT for 1000 tonnes (Hrs)": 0.0,
        "Avg. TRT for 1000 TEUs (Days)": None,
        "Avg. TRT for 1000 TEUs (Hrs)": None,
    }
    c["Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c."] = c["Avg. Turn around time - Pilot Boarding to De-boarding-Total"]

    tat_total_days = b["Avg. Turn around time - Total"]
    total_tons = b["Total traffic throughputs (Tons)"]
    trt_days = round(_safe_div(tat_total_days * vs * 1000, total_tons), 6) if total_tons else 0.0
    c["Avg. TRT for 1000 tonnes (Days)"] = trt_days
    c["Avg. TRT for 1000 tonnes (Hrs)"] = round(trt_days * 24, 6)
    return c


REASONS_PORT = [
    "1. Non Availability of Berth", "2. Non Availability of Tugs", "3. Non Availability of Pilot",
    "4. Non availability of equipment.", "5. Equipment breakdown", "6. Absence of workers",
    "7. Strike / Stoppage", "8. Power failure", "9. Labour holidays",
    "10. Night navigation restrictions", "11. Draft Restriction", "12. Others",
]
REASONS_NON_PORT = [
    "1. Ships account", "2.Shippers account", "3. Agents options",
    "4. Absence of non-port workers", "5. For want of cargo", "6. Departure formalities",
    "7. Weather restrictions", "8. Lack of storage", "9. Tidal",
    "10. Documents not Ready", "11. Power Failure Grid", "12. Not in Schedule", "13. Others",
]


@bp.route("/module/RP01/report11/")
@login_required
def report11_index():
    return render_template("report11/report11.html", port_name="JJLTPL")


@bp.route("/api/module/RP01/report11/meta")
@login_required
def report11_api_meta():
    try:
        _, years = _get_df_and_years()
        return jsonify({"years": years, "port_name": "JJLTPL"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


@bp.route("/api/module/RP01/report11/report")
@login_required
def report11_api_report():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1] if years else None)
        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        months = month_options_for(fin_year)
        a_by_month, b_by_month, c_by_month = [], [], []
        for mo in months:
            a = compute_section_a_month(df, fin_year, mo["idx"])
            print(
                "EXPORT",
                mo["label"],
                "Liquid Import =", a["Liquid - Import"]
            )
            b = compute_section_b_month(a)
            c = compute_section_c_month(a, b)
            a_by_month.append(a)
            b_by_month.append(b)
            c_by_month.append(c)

        return jsonify({
            "port_name": "JJLTPL",
            "fin_year": fin_year,
            "months": [m["label"] for m in months],
            "section_a": a_by_month,
            "section_b": b_by_month,
            "section_c": c_by_month,
        })
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500


# ---------------------------------------------------------------------
# Excel export — same grid layout as the uploaded workbook:
#   B=SR.NO, C=PARTICULARS, D=Units, E:P=Apr..Mar, Q=FY Total
# ---------------------------------------------------------------------
def _write_row(ws, row_i, sr, label, unit, values, fmt, thin_border, is_formula_row=False):
    ws[f"B{row_i}"] = sr
    ws[f"C{row_i}"] = label
    ws[f"D{row_i}"] = unit
    for i, v in enumerate(values):
        col = get_column_letter(5 + i)  # E=5
        cell = ws[f"{col}{row_i}"]
        cell.value = v
        cell.number_format = fmt
        cell.border = thin_border
    # FY total column (Q = col 17)
    q = ws[f"Q{row_i}"]
    if not is_formula_row:
        q.value = f"=SUM(E{row_i}:P{row_i})"
    q.number_format = fmt
    q.border = thin_border
    for c in ("B", "C", "D"):
        ws[f"{c}{row_i}"].border = thin_border


@bp.route("/api/module/RP01/report11/export")
@login_required
def report11_api_export():
    try:
        df, years = _get_df_and_years()
        fin_year = request.args.get("fin_year", years[-1] if years else None)
        port_name = request.args.get("port_name", "JJLTPL")

        if fin_year not in years:
            return jsonify({"error": f"Unknown fin_year '{fin_year}'. Available: {', '.join(years)}"}), 400

        months = month_options_for(fin_year)
        a_list, b_list, c_list = [], [], []
        for mo in months:
            a = compute_section_a_month(df, fin_year, mo["idx"])
            print(
                mo["label"],
                "Liquid Import =", a["Liquid - Import"],
                "Liquid Export =", a["Liquid - Export"]
            )
            b = compute_section_b_month(a)
            c = compute_section_c_month(a, b)
            a_list.append(a)
            b_list.append(b)
            c_list.append(c)


        wb = Workbook()
        ws = wb.active
        ws.title = "Bulk Terminal Performance"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=13)
        header_font = Font(bold=True)
        section_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center")

        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["B2"] = "Bulk Terminal - Liquid JJLTPL"
        ws["B2"].font = title_font
        ws["B3"] = f"Performance Report {fin_year}"
        ws["B3"].font = bold
        ws["B4"] = f"Port / Terminal - {port_name}"
        ws["B4"].font = bold

        header_row = 6
        ws[f"B{header_row}"] = "SR.NO."
        ws[f"C{header_row}"] = "PARTICULARS"
        ws[f"D{header_row}"] = "Units"
        for i, mo in enumerate(months):
            ws[f"{get_column_letter(5 + i)}{header_row}"] = mo["label"]
        ws[f"Q{header_row}"] = f"FY {fin_year}"
        for col_idx in range(2, 18):
            cell = ws[f"{get_column_letter(col_idx)}{header_row}"]
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

        row_i = header_row + 1

        # ---------------- Section A ----------------
        ws[f"B{row_i}"] = "A)"
        ws[f"C{row_i}"] = "TERMINAL PERFOMANCE (To be filled by Terminals)"
        ws[f"B{row_i}"].font = section_font
        ws[f"C{row_i}"].font = section_font
        row_i += 1

        a_rows = [
            ("Vessel Discharge (Including Restow)", "TEUs", "0.0"),
            ("Vessel Load (Including Restow)", "TEUs", "0.0"),
            ("Tonnage", "Tonnes", "0.0"),
            ("Vessels Sailed", "Nos", "0"),
            ("Pre_berthing Waiting Time-on Port a/c (Total)", "Hrs", "0.00"),
            ("Pre_berthing Waiting Time-on Non-Port a/c (Total)", "Hrs.", "0.00"),
            ("Total Berth Stay of all vessels (For Berth Productivity)", "Hrs", "0.00"),
            ("Total Crane deplyoed hours (For Crane Productivity)", "Hrs", "0.0"),
            ("Vessel Inward movement (Total)", "Hrs", "0.00"),
            ("Vessel Outward movement (Total)", "Hrs", "0.00"),
            ("Idle time at working berth on Port A/c.", "Hrs.", "0.00"),
            ("Idle time at working berth on Non-Port A/c.", "Hrs.", "0.00"),
            ("Idle time at Non-working berth on Port A/c.", "Hrs.", "0.0"),
            ("Idle time at Non-working berth on Non-Port A/c.", "Hrs.", "0.0"),
            ("Shifting Time", "Hrs.", "0.0"),
            ("Total No. of Moves for calculating  Berth / Crane Productivity", "Moves", "0.0"),
            ("Total No. of TEUs for calculating  Crane Productivity", "TEUs", "0.0"),
            ("Rail Load", "TEUs", "0.0"),
            ("Rail Discharge", "TEUs", "0.0"),
            ("No. of Rakes handled", "Nos", "0.0"),
            ("Days in a month", "Nos", "0"),
            ("No. of berths for % berth occupancy", "Nos", "0"),
            ("Dry Bulk traffic - Import", "Tonnes", "0.000"),
            ("Dry Bulk traffic - Export", "Tonnes", "0.000"),
            ("Break Bulk traffic - Import", "Tonnes", "0.000"),
            ("Break Bulk traffic - Export", "Tonnes", "0.000"),
            ("Liquid - Import", "Tonnes", "0.000"),
            ("Liquid - Export", "Tonnes", "0.000"),
        ]
        a_row_excel_num = {}
        for sr, (label, unit, fmt) in enumerate(a_rows, start=1):
            values = [a_list[i][label] for i in range(12)]
            _write_row(ws, row_i, sr, label, unit, values, fmt, thin_border)

            # Constant across all months — averaging (or just showing the constant)
            # is correct; summing to "24" for a 2-berth terminal is wrong.
            if label == "No. of berths for % berth occupancy":
                ws[f"Q{row_i}"] = f"=AVERAGE(E{row_i}:P{row_i})"

            a_row_excel_num[label] = row_i
            row_i += 1

        def A(label, month_i):
            return f"{get_column_letter(5 + month_i)}{a_row_excel_num[label]}"

        row_i += 1
        # ---------------- Section B ----------------
        ws[f"B{row_i}"] = "B)"
        ws[f"C{row_i}"] = "PRODUCTIVITY PARAMETERS (Derived - Not to be filled)"
        ws[f"B{row_i}"].font = section_font
        ws[f"C{row_i}"].font = section_font
        row_i += 1

        b_defs = [
            ("Vessels Sailed", "Nos.", "0", lambda i: f"={A('Vessels Sailed', i)}"),
            ("Total Traffic Throughputs (TEUs)", "TEUs", "0",
             lambda i: f"={A('Vessel Discharge (Including Restow)', i)}+{A('Vessel Load (Including Restow)', i)}"),
            ("Total traffic throughputs (Tons)", "Tons", "0.000",
             lambda i: f"={A('Dry Bulk traffic - Import', i)}+{A('Dry Bulk traffic - Export', i)}"
                       f"+{A('Break Bulk traffic - Import', i)}+{A('Break Bulk traffic - Export', i)}"
                       f"+{A('Liquid - Import', i)}+{A('Liquid - Export', i)}"),
            ("Parcel Size", "TEUs", "0",
             lambda i: f"=IFERROR({A('Vessel Discharge (Including Restow)', i)}+{A('Vessel Load (Including Restow)', i)})/{A('Vessels Sailed', i)},0)"),
            ("Avg. Pre-berthing Waiting Time-Total", "Days", "0.000000",
             lambda i: f"=IFERROR(({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}+{A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)})/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Pre-berthing Waiting Time-Port A/c.", "Days", "0.000000",
             lambda i: f"=IFERROR({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Pre-berthing Waiting Time-Non-Port A/c.", "Days", "0.000000",
             lambda i: f"=IFERROR({A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)}/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Berth stay", "Days", "0.000000",
             lambda i: f"=IFERROR({A('Total Berth Stay of all vessels (For Berth Productivity)', i)}/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Total", "Days", "0.000000",
             lambda i: f"=IFERROR(({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}+{A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)}+{A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Port A/c.", "Days", "0.000000",
             lambda i: f"=IFERROR(({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}+{A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Non- Port A/c.", "Days", "0.000000",
             lambda i: f"=IFERROR({A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)}/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Total", "Days", "0.000000",
             lambda i: f"=IFERROR(({A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/24/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c.", "Days", "0.000000",
             lambda i: f"=IFERROR(({A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/24/{A('Vessels Sailed', i)},0)"),
            ("Berth Occupancy", "%", "0.00%",
             lambda i: f"=IFERROR({A('Total Berth Stay of all vessels (For Berth Productivity)', i)}/({A('Days in a month', i)}*24*{A('No. of berths for % berth occupancy', i)}),0)"),
            ("Idle time", "%", "0.00%",
             lambda i: f"=IFERROR(({A('Idle time at working berth on Port A/c.', i)}+{A('Idle time at working berth on Non-Port A/c.', i)})/{A('Total Berth Stay of all vessels (For Berth Productivity)', i)},0)"),
            ("Gross Berth Productivity", "Moves /Hrs", "0", lambda i: 0),
            ("Gross Crane Productivity (Moves)", "Moves /Hrs", "0", lambda i: None),
            ("Gross Crane Productivity (TEUs)", "TEUs/Hrs", "0", lambda i: None),
            ("Ship Output per Day (TEUs)", "TEUs", "0", lambda i: 0),
        ]
        b_row_excel_num = {}
        for sr, (label, unit, fmt, formula_fn) in enumerate(b_defs, start=1):
            values = [formula_fn(i) for i in range(12)]
            _write_row(ws, row_i, sr, label, unit, values, fmt, thin_border, is_formula_row=True)

            # Anything labeled "Avg." (or ratio-type rows) must be averaged across
            # months for the FY column, never summed — summing 12 monthly averages
            # produces a meaningless inflated number.
            if label.startswith("Avg.") or label in ("Berth Occupancy", "Idle time"):
                ws[f"Q{row_i}"] = f"=AVERAGE(E{row_i}:P{row_i})"
            else:
                ws[f"Q{row_i}"] = f"=SUM(E{row_i}:P{row_i})"

            b_row_excel_num[label] = row_i
            row_i += 1

        def B(label, month_i):
            return f"{get_column_letter(5 + month_i)}{b_row_excel_num[label]}"

        # Ship Output per Day (Tonnes) needs a B-row reference -> add after B defined
        sr_next = len(b_defs) + 1
        values = [
            f"=IFERROR({A('Dry Bulk traffic - Import', i)}+{A('Dry Bulk traffic - Export', i)}"
            f"+{A('Break Bulk traffic - Import', i)}+{A('Break Bulk traffic - Export', i)}"
            f"+{A('Liquid - Import', i)}+{A('Liquid - Export', i)})/{B('Avg. Berth stay', i)},0)"
            for i in range(12)
        ]
        _write_row(ws, row_i, sr_next, "Ship Output per Day (Tonnes)", "Tonnes", values, "0.000", thin_border, is_formula_row=True)

        # FY total = FY total tonnage ÷ FY average berth stay (both already correct
        # in their own Q columns) — not a straight SUM of 12 monthly rates.
        ws[f"Q{row_i}"] = (
            f"=IFERROR(Q{a_row_excel_num['Dry Bulk traffic - Import']}"
            f"+Q{a_row_excel_num['Dry Bulk traffic - Export']}"
            f"+Q{a_row_excel_num['Break Bulk traffic - Import']}"
            f"+Q{a_row_excel_num['Break Bulk traffic - Export']}"
            f"+Q{a_row_excel_num['Liquid - Import']}"
            f"+Q{a_row_excel_num['Liquid - Export']},Q{b_row_excel_num['Avg. Berth stay']}),0)"
        )
        b_row_excel_num["Ship Output per Day (Tonnes)"] = row_i
        row_i += 1

        for label, unit, fmt in [
            ("No. of Rakes handled", "Nos", "0"),
            ("Total Rail traffic", "TEUs", "0"),
            ("% wrt to Total Thoughput", "%", "0.0%"),
        ]:
            row_i += 1
            ws[f"B{row_i}"] = ""
            ws[f"C{row_i}"] = label
            ws[f"D{row_i}"] = unit
            row_i += 0  # placeholder rows left blank (no data source)

        row_i += 1
        # ---------------- Section C ----------------
        ws[f"B{row_i}"] = "C)"
        ws[f"C{row_i}"] = "PRODUCTIVITY PARAMETERS (Derived - Not to be filled)"
        ws[f"B{row_i}"].font = section_font
        ws[f"C{row_i}"].font = section_font
        row_i += 1

        c_defs = [
            ("Avg. Pre-berthing Waiting Time-Total", "Hrs.", "0.000000",
             lambda i: f"=IFERROR(({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}+{A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)})/{A('Vessels Sailed', i)},0)"),
            ("Avg. Pre-berthing Waiting Time-Port A/c.", "Hrs.", "0.000000",
             lambda i: f"=IFERROR({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}/{A('Vessels Sailed', i)},0)"),
            ("Avg. Pre-berthing Waiting Time-Non-Port A/c.", "Hrs.", "0.000000",
             lambda i: f"=IFERROR({A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)}/{A('Vessels Sailed', i)},0)"),
            ("Avg. Berth stay", "Hrs.", "0.000000",
             lambda i: f"=IFERROR({A('Total Berth Stay of all vessels (For Berth Productivity)', i)}/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Total", "Hrs.", "0.000000",
             lambda i: f"=IFERROR(({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}+{A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)}+{A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Port A/c.", "Hrs.", "0.000000",
             lambda i: f"=IFERROR(({A('Pre_berthing Waiting Time-on Port a/c (Total)', i)}+{A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Non- Port A/c.", "Hrs.", "0.000000",
             lambda i: f"=IFERROR({A('Pre_berthing Waiting Time-on Non-Port a/c (Total)', i)}/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Total", "Hrs.", "0.000000",
             lambda i: f"=IFERROR(({A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/{A('Vessels Sailed', i)},0)"),
            ("Avg. Turn around time - Pilot Boarding to De-boarding-Port A/c.", "Hrs.", "0.000000",
             lambda i: f"=IFERROR(({A('Total Berth Stay of all vessels (For Berth Productivity)', i)}+{A('Vessel Inward movement (Total)', i)}+{A('Vessel Outward movement (Total)', i)})/{A('Vessels Sailed', i)},0)"),
        ]
        c_row_excel_num = {}
        for sr, (label, unit, fmt, formula_fn) in enumerate(c_defs, start=1):
            values = [formula_fn(i) for i in range(12)]
            _write_row(ws, row_i, sr, label, unit, values, fmt, thin_border, is_formula_row=True)
            ws[f"Q{row_i}"] = f"=AVERAGE(E{row_i}:P{row_i})"
            c_row_excel_num[label] = row_i
            row_i += 1

        def C(label, month_i):
            return f"{get_column_letter(5 + month_i)}{c_row_excel_num[label]}"

        trt_days_values = [
            f"=IFERROR({B('Avg. Turn around time - Total', i)}*{A('Vessels Sailed', i)}*1000/{B('Total traffic throughputs (Tons)', i)},0)"
            for i in range(12)
        ]
        _write_row(ws, row_i, len(c_defs) + 1, "Avg. TRT for 1000 tonnes", "Days", trt_days_values, "0.000000", thin_border, is_formula_row=True)
        ws[f"Q{row_i}"] = f"=AVERAGE(E{row_i}:P{row_i})"
        trt_days_row = row_i
        row_i += 1

        trt_hrs_values = [f"={get_column_letter(5 + i)}{trt_days_row}*24" for i in range(12)]
        _write_row(ws, row_i, len(c_defs) + 2, "Avg. TRT for 1000 tonnes", "Hrs.", trt_hrs_values, "0.000000", thin_border, is_formula_row=True)
        ws[f"Q{row_i}"] = f"=AVERAGE(E{row_i}:P{row_i})"
        row_i += 1

        for label, unit in [("Avg. TRT for 1000 TEUs", "Days"), ("Avg. TRT for 1000 TEUs", "Hrs.")]:
            row_i += 1
            ws[f"C{row_i}"] = label
            ws[f"D{row_i}"] = unit

        # ---------------- Reasons block (static legend) ----------------
        row_i += 2
        ws[f"C{row_i}"] = "Reasons to be Considered"
        ws[f"C{row_i}"].font = bold
        row_i += 1
        ws[f"C{row_i}"] = "Port A/c."
        ws[f"D{row_i}"] = "Non-Port  A/c."
        ws[f"C{row_i}"].font = header_font
        ws[f"D{row_i}"].font = header_font
        row_i += 1
        for i in range(max(len(REASONS_PORT), len(REASONS_NON_PORT))):
            if i < len(REASONS_PORT):
                ws[f"C{row_i}"] = REASONS_PORT[i]
            if i < len(REASONS_NON_PORT):
                ws[f"D{row_i}"] = REASONS_NON_PORT[i]
            row_i += 1

        # ---- column widths -------------------------------------------------
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 48
        ws.column_dimensions["D"].width = 10
        for c in range(5, 18):
            ws.column_dimensions[get_column_letter(c)].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Report-11_Bulk_Terminal_Performance_{fin_year}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ReportDataError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Unexpected server error: {e}"}), 500